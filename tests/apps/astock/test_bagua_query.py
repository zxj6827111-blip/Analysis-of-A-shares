# -*- coding: utf-8 -*-
"""Tests for bagua single-stock query service."""

from __future__ import annotations

import re
from pathlib import Path
from types import SimpleNamespace

import pytest

from wtpy.apps.astock.bagua.calculator import BaguaCalculator
from wtpy.apps.astock.data.tdx_reader import DayBar
from wtpy.apps.astock.service import bagua_query as bq


JSON_PATH = (
    Path(__file__).resolve().parents[3]
    / "wtpy"
    / "apps"
    / "astock"
    / "bagua"
    / "bagua_384.json"
)


@pytest.fixture(autouse=True)
def _no_tushare_network(monkeypatch):
    """禁止导出测试触网：Tushare 元数据拉取一律降级，不写真实缓存。

    需要验证网络行为的测试会在测试体内显式重新 mock 该函数覆盖此默认值。
    同时清空模块级元数据缓存，避免测试间泄漏导致断言不确定。
    """

    def _fail(*_a, **_k):
        raise RuntimeError("Tushare network disabled in tests")

    monkeypatch.setattr(bq, "_fetch_symbol_meta_from_tushare", _fail)
    monkeypatch.setattr(bq, "_SYMBOL_META_CACHE", {})
    yield


def test_normalize_period_and_code():
    assert bq.normalize_period("日") == "DAY"
    assert bq.normalize_period("按月") == "MONTH"
    assert bq.normalize_period("WEEK") == "WEEK"
    assert bq.normalize_query_code("600000") == "SSE.STK.600000"
    assert bq.normalize_query_code("sh600000") == "SSE.STK.600000"
    assert bq.normalize_query_code("000001") == "SZSE.STK.000001"
    assert bq.normalize_query_code("600000.SH") == "SSE.STK.600000"
    assert bq.normalize_query_code("000001.SZ") == "SZSE.STK.000001"
    assert bq.normalize_query_code("920001.BJ") == "BSE.STK.920001"
    assert bq.normalize_query_code("000001.SH") == "SSE.IDX.000001"
    assert bq.normalize_query_code("399001.SZ") == "SZSE.IDX.399001"
    assert bq.display_code("SSE.STK.600000") == "sh600000"


def test_parse_ymd():
    assert bq._parse_ymd("2024-01-15") == 20240115
    assert bq._parse_ymd(20240115) == 20240115
    assert bq._parse_ymd("2024/01/15") == 20240115


def test_find_day_bar_fallback():
    bars = [
        DayBar(20240102, 10, 11, 9, 10.5, 1, 1),
        DayBar(20240103, 10.5, 11, 10, 10.8, 1, 1),
        DayBar(20240105, 10.8, 11.2, 10.5, 11, 1, 1),
    ]
    bar, exact = bq._find_day_bar(bars, 20240103)
    assert exact and bar.date == 20240103
    bar2, exact2 = bq._find_day_bar(bars, 20240104)  # holiday
    assert not exact2 and bar2.date == 20240103


def test_query_bagua_with_mock_bars(monkeypatch):
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")

    bars = [
        DayBar(
            date=20240103,
            open=6.27,
            high=7.33,
            low=5.90,
            close=5.90,
            amount=1.0,
            volume=1.0,
        ),
    ]

    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=Path("."),
        tdx_root=Path("."),
        forecast_root=Path("."),
        forecast_weekly_dir=Path("."),
        universe_path=Path("."),
    )
    monkeypatch.setattr(bq, "load_day_bars", lambda _cfg, _code: bars)

    out = bq.query_bagua(cfg, code="600000", date="2024-01-03", period="DAY")
    assert out["ok"] is True
    assert out["period"] == "DAY"
    assert out["code"] == "sh600000"
    assert out["bar"]["open"] == 6.27
    assert out["bagua"]["upper_id"] == 7
    assert out["bagua"]["lower_id"] == 6
    assert out["bagua"]["yao_order"] == 3
    assert out["summary"]["yao_order"] == 3
    assert "山水" in (out["summary"]["full_name"] or out["bagua"].get("full_name") or "")


def test_calculator_same_as_query(monkeypatch):
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")
    calc = BaguaCalculator.from_json(JSON_PATH)
    r = calc.calculate(open_price=6.27, high_price=7.33, low_price=5.90, close_price=5.90)
    bars = [
        DayBar(20240103, 6.27, 7.33, 5.90, 5.90, 1, 1),
    ]
    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=Path("."),
        tdx_root=Path("."),
        forecast_root=Path("."),
        forecast_weekly_dir=Path("."),
        universe_path=Path("."),
    )
    monkeypatch.setattr(bq, "load_day_bars", lambda *_a, **_k: bars)
    out = bq.query_bagua(cfg, code="sz000001", date=20240103, period="DAY")
    assert out["bagua"]["full_name"] == r.full_name
    assert out["bagua"]["yao_name"] == r.yao_name
    assert "name" in out
    assert "display" in out


def test_source_match_pairs():
    from wtpy.apps.astock.service.bagua_query import SourceDisabledError

    with pytest.raises(SourceDisabledError, match="已停用"):
        bq._source_match_pairs("tdx_front")
    pairs = bq._source_match_pairs("tushare_qfq")
    assert pairs[0] == ("internal", "composite_tushare_factor_qfq")
    assert pairs[1] == ("tushare", "qfq")
    # legacy internal/tushare_factor_qfq derived sets are no longer readable
    assert ("internal", "tushare_factor_qfq") not in pairs
    raw_pairs = bq._source_match_pairs("raw")
    assert raw_pairs[0] == ("internal", "composite_none")
    assert ("tushare", "none") in raw_pairs
    assert all(p[0] != "local_vendor" for p in raw_pairs)


def test_query_bagua_tushare_qfq_from_warehouse(monkeypatch):
    """Tushare QFQ resolves via the warehouse multi-dataset scan."""
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")

    bars_ts = [DayBar(20240103, 10.1, 11.1, 9.6, 10.6, 1, 1)]

    def _fake_load(_cfg, _code, source_key, asof=None):
        return bars_ts, {
            "dataset_id": "internal_composite_tushare_factor_qfq_1d_new",
            "dataset_source": "internal",
            "dataset_adjustment": "composite_tushare_factor_qfq",
            "dataset_status": "ready",
            "covers_asof": True,
            "candidate_datasets": 3,
        }

    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=Path("."),
        tdx_root=Path("."),
        market_data_root=Path("."),
        forecast_root=Path("."),
        forecast_weekly_dir=Path("."),
        universe_path=Path("."),
        adj_root=Path("."),
    )
    monkeypatch.setattr(bq, "_load_dataset_bars", _fake_load)

    out_ts = bq.query_bagua(
        cfg, code="600000", date="2024-01-03", period="DAY", adjust="tushare_qfq"
    )
    assert out_ts["ok"] is True
    assert out_ts["adjust"] == "tushare_qfq"
    assert out_ts["bar"]["close"] == 10.6
    assert out_ts["adjust_meta"]["dataset_source"] == "internal"
    assert "Tushare" in (out_ts["algorithm"].get("price_format") or "")


def test_query_bagua_raw_from_warehouse(monkeypatch):
    """未复权走仓库正式L2（internal/composite_none）优先。"""
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")

    bars_raw = [DayBar(20240103, 9.0, 9.5, 8.8, 9.2, 1, 1)]

    def _fake_load(_cfg, _code, source_key, asof=None):
        assert source_key == "raw"
        return bars_raw, {
            "dataset_id": "internal_composite_none_1d_20260726",
            "dataset_source": "internal",
            "dataset_adjustment": "composite_none",
            "dataset_status": "ready",
            "covers_asof": True,
            "candidate_datasets": 1,
        }

    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=Path("."),
        tdx_root=Path("."),
        market_data_root=Path("."),
        forecast_root=Path("."),
        forecast_weekly_dir=Path("."),
        universe_path=Path("."),
        adj_root=Path("."),
    )
    monkeypatch.setattr(bq, "_load_dataset_bars", _fake_load)

    out = bq.query_bagua(cfg, code="600000", date="2024-01-03", period="DAY", adjust="raw")
    assert out["ok"] is True
    assert out["adjust"] == "raw"
    assert out["adjust_meta"]["dataset_source"] == "internal"
    assert out["adjust_meta"]["dataset_adjustment"] == "composite_none"
    assert "正式L2" in (out["algorithm"].get("price_format") or "")


def test_bagua_plane_session_indexes_once(tmp_path):
    """Session lists manifests once; multi-symbol loads share the index."""
    from wtpy.apps.astock.data.dataset_store import (
        DatasetManifest,
        DatasetStore,
        SymbolRecord,
    )
    from wtpy.apps.astock.data.providers.base import MarketBar

    store = DatasetStore(tmp_path / "market_data")
    codes = ["SSE.STK.600000", "SSE.STK.600004", "SZSE.STK.000001"]
    records = []
    total_rows = 0
    # >= 120 rows and a >60-day span so the orphan-window gate passes
    dates = list(range(20240102, 20240102 + 130))
    for code in codes:
        bars = [
            MarketBar(
                symbol=code,
                trade_date=date,
                period="1d",
                open=10.0,
                high=11.0,
                low=9.0,
                close=10.5,
                volume=1000.0,
                amount=10000.0,
            )
            for date in dates
        ]
        sha = store.store_bars(code, bars)
        records.append(
            SymbolRecord(
                symbol=code,
                blob_sha256=sha,
                row_count=len(bars),
                quality="ok",
                first_date=dates[0],
                last_date=dates[-1],
            )
        )
        total_rows += len(bars)
    # Tushare-only policy: the composite is only eligible as the FORMAL L1
    # (exact dataset id once the atomic product pair exists), so the session
    # test must publish the formal L2 + L1 pair with tushare_only_v1 lineage:
    # L2 carries its two parents (base + supplement) and L1 its raw + factor
    # parents so the strict fail-closed lineage validation accepts the pair.
    def _parent(dataset_id, source, adjustment, *, factor=False):
        store.publish(DatasetManifest(
            dataset_id=dataset_id,
            source=source,
            adjustment=adjustment,
            period="1d",
            status="building",
            dataset_type="factor" if factor else "bars",
            data_cutoff_date=dates[-1],
            symbol_count=0,
            row_count=0,
        ))

    _parent("tushare_none_1d_pair_base", "tushare", "none")
    _parent("tushare_none_1d_pair_supp", "tushare", "none")
    _parent("tushare_adjfactor_1d_pair", "tushare", "adj_factor", factor=True)
    l2_id = "internal_composite_none_1d_session_t1"
    l2_manifest = DatasetManifest(
        dataset_id=l2_id,
        source="internal",
        adjustment="composite_none",
        period="1d",
        status="building",
        data_cutoff_date=dates[-1],
        symbols=records,
        symbol_count=len(records),
        row_count=total_rows,
        provenance={
            "data_policy": "tushare_only_v1",
            "base_source": "tushare",
            "supplement_source": "tushare",
            "parents": [
                {"dataset_id": "tushare_none_1d_pair_base", "role": "base"},
                {"dataset_id": "tushare_none_1d_pair_supp", "role": "supplement"},
            ],
        },
    )
    store.publish(l2_manifest)
    m = DatasetManifest(
        dataset_id="internal_composite_tushare_factor_qfq_1d_session_t1",
        source="internal",
        adjustment="composite_tushare_factor_qfq",
        period="1d",
        status="building",
        data_cutoff_date=dates[-1],
        symbols=records,
        symbol_count=len(records),
        row_count=total_rows,
        raw_dataset_id=l2_id,
        factor_dataset_id="tushare_adjfactor_1d_pair",
        provenance={"data_policy": "tushare_only_v1"},
    )
    store.publish(m)

    cfg = SimpleNamespace(market_data_root=tmp_path / "market_data")
    session = bq.BaguaPlaneSession(cfg, "tushare_qfq")
    assert session._saw_any_pair is True
    assert len(session._indexed) >= 1
    # the formal product pair pins the exact composite as the only candidate
    assert session.formal_l1_id == "internal_composite_tushare_factor_qfq_1d_session_t1"

    bars0, meta0 = session.load_symbol("SSE.STK.600000")
    assert len(bars0) == 130
    assert meta0.get("session_indexed") is True
    assert meta0.get("dataset_id") == "internal_composite_tushare_factor_qfq_1d_session_t1"

    bars1, meta1 = session.load_symbol("SZSE.STK.000001", asof=20240103)
    assert len(bars1) == 130
    assert meta1.get("covers_asof") is True

    # Shared load_day_bars_for_plane path with session + date trim
    bars2, meta2 = bq.load_day_bars_for_plane(
        cfg, "SSE.STK.600004", "tushare_qfq", start=20240103, end=20240104, session=session
    )
    assert len(bars2) == 2
    assert meta2.get("session_indexed") is True

    with pytest.raises(FileNotFoundError):
        session.load_symbol("SSE.STK.999999")



def test_bagua_plane_session_prefers_freshest_stale_candidate(tmp_path):
    """A stale composite surface must not hide a newer Tushare increment."""
    from wtpy.apps.astock.data.dataset_store import (
        DatasetManifest,
        DatasetStore,
        SymbolRecord,
    )
    from wtpy.apps.astock.data.providers.base import MarketBar

    store = DatasetStore(tmp_path / "market_data")
    code = "SZSE.STK.300475"

    def publish(dataset_id, source, adjustment, dates, *, reverse_record_dates=False):
        bars = [
            MarketBar(
                symbol=code,
                trade_date=date,
                period="1d",
                open=10.0,
                high=11.0,
                low=9.0,
                close=float(date % 100),
                volume=1000.0,
                amount=10000.0,
            )
            for date in dates
        ]
        sha = store.store_bars(code, bars)
        first_date, last_date = min(dates), max(dates)
        if reverse_record_dates:
            first_date, last_date = last_date, first_date
        record = SymbolRecord(
            symbol=code,
            blob_sha256=sha,
            row_count=len(bars),
            quality="ok",
            first_date=first_date,
            last_date=last_date,
        )
        manifest = DatasetManifest(
            dataset_id=dataset_id,
            source=source,
            adjustment=adjustment,
            period="1d",
            status="building",
            data_cutoff_date=max(dates),
            symbols=[record],
            symbol_count=1,
            row_count=len(bars),
        )
        store.publish(manifest)

    # >= 120 rows with a >60-day span so the orphan gate passes
    stale_dates = list(range(20240101, 20240101 + 200))  # 20240101..20240300
    fresh_dates = list(range(20240101 + 100, 20240101 + 300))  # 20240201..20240400
    publish(
        "internal_composite_none_1d_20240300_test",
        "internal",
        "composite_none",
        stale_dates,
    )
    publish(
        "tushare_none_1d_20240400_test",
        "tushare",
        "none",
        fresh_dates,
        reverse_record_dates=True,
    )

    cfg = SimpleNamespace(market_data_root=tmp_path / "market_data")
    session = bq.BaguaPlaneSession(cfg, "raw")

    # Tushare-only policy bootstrap: without a formal product pair the legacy
    # internal composite (no tushare_only_v1 lineage) is INELIGIBLE — it must
    # never hide the newer complete tushare/none increment.
    assert session.formal_l2_id is None

    # latest query: freshest real data wins within the role
    fresh_bars, fresh_meta = session.load_symbol(code, asof=20241201)
    assert fresh_meta["dataset_source"] == "tushare"
    assert fresh_meta["dataset_id"] == "tushare_none_1d_20240400_test"
    assert fresh_meta["bootstrap_fallback"] is True
    assert fresh_meta["symbol_effective_last_date"] == 20240400
    assert fresh_bars[-1].date == 20240400

    # historical asof: the legacy composite is not a product candidate; the
    # complete tushare/none bootstrap surface covers the query date instead
    historical_bars, historical_meta = session.load_symbol(code, asof=20240201)
    assert historical_meta["dataset_source"] == "tushare"
    assert historical_meta["dataset_id"] == "tushare_none_1d_20240400_test"
    assert historical_meta["covers_asof"] is True
    assert historical_bars[-1].date == 20240400


def test_bagua_plane_session_caches_manifest_signals(tmp_path):
    """Per-manifest history signals are computed once at index build; _score
    reads the cached signals (no per-symbol manifest rescans)."""
    from wtpy.apps.astock.data.dataset_store import (
        DatasetManifest,
        DatasetStore,
        SymbolRecord,
    )
    from wtpy.apps.astock.data.providers.base import MarketBar
    from wtpy.apps.astock.data.tushare_product import manifest_history_signals

    store = DatasetStore(tmp_path / "market_data")
    code = "SSE.STK.600000"
    dates = list(range(20240102, 20240102 + 130))
    bars = [
        MarketBar(
            symbol=code, trade_date=d, period="1d",
            open=10.0, high=11.0, low=9.0, close=10.5,
            volume=1000.0, amount=10000.0,
        )
        for d in dates
    ]
    sha = store.store_bars(code, bars)
    m = DatasetManifest(
        dataset_id="tushare_none_1d_sigcache_t1",
        source="tushare",
        adjustment="none",
        period="1d",
        status="building",
        data_cutoff_date=dates[-1],
        symbols=[SymbolRecord(
            symbol=code, blob_sha256=sha, row_count=len(bars), quality="ok",
            first_date=dates[0], last_date=dates[-1],
        )],
        symbol_count=1,
        row_count=len(bars),
    )
    store.publish(m)

    cfg = SimpleNamespace(market_data_root=tmp_path / "market_data")
    session = bq.BaguaPlaneSession(cfg, "raw")
    # the cached entry exists and matches a direct computation
    sig = session._manifest_sig.get(m.dataset_id)
    assert sig is not None
    assert sig == manifest_history_signals(m)
    assert sig.median_rows == 130
    # loads still resolve through the cached signals
    bars_out, meta = session.load_symbol(code, asof=20240201)
    assert meta["dataset_id"] == m.dataset_id
    assert len(bars_out) == 130


def test_bagua_plane_session_short_index_etf_surface_indexed(tmp_path):
    """A newly listed ETF/index tushare/none surface (short window, few rows)
    must NOT be dropped wholesale by the orphan-window gate (it would hide
    the only warehouse bars for that symbol), while a short stock surface
    stays excluded."""
    from wtpy.apps.astock.data.dataset_store import (
        DatasetManifest,
        DatasetStore,
        SymbolRecord,
    )
    from wtpy.apps.astock.data.providers.base import MarketBar

    store = DatasetStore(tmp_path / "market_data")

    def publish(dataset_id, symbol, dates):
        bars = [
            MarketBar(
                symbol=symbol, trade_date=d, period="1d",
                open=10.0, high=11.0, low=9.0, close=float(d % 100),
                volume=1000.0, amount=10000.0,
            )
            for d in dates
        ]
        sha = store.store_bars(symbol, bars)
        store.publish(DatasetManifest(
            dataset_id=dataset_id,
            source="tushare",
            adjustment="none",
            period="1d",
            status="building",
            data_cutoff_date=max(dates),
            symbols=[SymbolRecord(
                symbol=symbol, blob_sha256=sha, row_count=len(bars),
                quality="ok", first_date=min(dates), last_date=max(dates),
            )],
            symbol_count=1,
            row_count=len(bars),
        ))

    # newly listed ETF / index: 16 rows, short span -> orphan window
    publish("tushare_none_1d_etf_new_t1", "SSE.ETF.510300",
            list(range(20260701, 20260701 + 16)))
    publish("tushare_none_1d_idx_new_t1", "SSE.IDX.000001",
            list(range(20260701, 20260701 + 16)))
    # short stock surface must stay excluded by the orphan gate
    publish("tushare_none_1d_stock_short_t1", "SSE.STK.600000",
            list(range(20260701, 20260701 + 16)))

    cfg = SimpleNamespace(market_data_root=tmp_path / "market_data")
    session = bq.BaguaPlaneSession(cfg, "raw")
    indexed_ids = {m.dataset_id for m, _, _ in session._indexed}
    assert "tushare_none_1d_etf_new_t1" in indexed_ids
    assert "tushare_none_1d_idx_new_t1" in indexed_ids
    assert "tushare_none_1d_stock_short_t1" not in indexed_ids

    bars, meta = session.load_symbol("sh510300", asof=20260710)
    assert meta["dataset_id"] == "tushare_none_1d_etf_new_t1"
    assert meta["covers_asof"] is True
    assert len(bars) == 16
    bars2, meta2 = session.load_symbol("sh000001", asof=20260710)
    assert meta2["dataset_id"] == "tushare_none_1d_idx_new_t1"
    assert meta2["covers_asof"] is True


def test_bagua_plane_session_index_bare_code_collision(tmp_path):
    """sh000001 (上证指数) must never resolve to SZSE.STK.000001 (平安银行).

    Both canonical symbols share the bare variant "000001"; the qualified
    variant (SSE.IDX.000001 / 000001.SH / sh000001) must win even when the
    same-coded stock sits in a dataset that ranks higher.
    """
    from wtpy.apps.astock.data.dataset_store import (
        DatasetManifest,
        DatasetStore,
        SymbolRecord,
    )
    from wtpy.apps.astock.data.providers.base import MarketBar

    store = DatasetStore(tmp_path / "market_data")

    def publish(dataset_id, symbol, dates):
        bars = [
            MarketBar(
                symbol=symbol,
                trade_date=date,
                period="1d",
                open=10.0,
                high=11.0,
                low=9.0,
                close=float(date % 100),
                volume=1000.0,
                amount=10000.0,
            )
            for date in dates
        ]
        sha = store.store_bars(symbol, bars)
        manifest = DatasetManifest(
            dataset_id=dataset_id,
            source="tushare",
            adjustment="none",
            period="1d",
            status="ready",
            data_cutoff_date=max(dates),
            symbols=[
                SymbolRecord(
                    symbol=symbol,
                    blob_sha256=sha,
                    row_count=len(bars),
                    quality="ok",
                    first_date=min(dates),
                    last_date=max(dates),
                )
            ],
            symbol_count=1,
            row_count=len(bars),
        )
        store.publish(manifest)

    # Stock dataset is ready and holds SZSE.STK.000001 — the collision trap.
    publish("tushare_none_1d_stock_test", "SZSE.STK.000001", list(range(20240102, 20240102 + 120)))
    publish("tushare_none_1d_index_test", "SSE.IDX.000001", list(range(20240102, 20240102 + 120)))

    cfg = SimpleNamespace(market_data_root=tmp_path / "market_data")
    session = bq.BaguaPlaneSession(cfg, "raw")

    bars, meta = session.load_symbol("sh000001", asof=20240105)
    assert meta["dataset_id"] == "tushare_none_1d_index_test"
    assert bars[0].date == 20240102
    assert bars[-1].date == 20240102 + 119

    # 000001.SH form and canonical form resolve the index too.
    bars2, meta2 = session.load_symbol("000001.SH", asof=20240105)
    assert meta2["dataset_id"] == "tushare_none_1d_index_test"
    bars3, meta3 = session.load_symbol("SSE.IDX.000001", asof=20240105)
    assert meta3["dataset_id"] == "tushare_none_1d_index_test"

    # The same-coded stock (sz000001) still resolves to the stock.
    bars4, meta4 = session.load_symbol("sz000001", asof=20240105)
    assert meta4["dataset_id"] == "tushare_none_1d_stock_test"


def test_bagua_plane_session_index_asof_before_inception_no_stock_leak(tmp_path):
    """Index queried before its inception must fail, not leak to the stock.

    When the qualified variant (SSE.IDX.000001) matches a record that does not
    cover the query date, we must NOT fall back to the bare code — otherwise
    an out-of-range historical index query returns 平安银行's bars.
    """
    from wtpy.apps.astock.data.dataset_store import (
        DatasetManifest,
        DatasetStore,
        SymbolRecord,
    )
    from wtpy.apps.astock.data.providers.base import MarketBar

    store = DatasetStore(tmp_path / "market_data")

    def publish(dataset_id, symbol, dates):
        bars = [
            MarketBar(
                symbol=symbol,
                trade_date=date,
                period="1d",
                open=10.0,
                high=11.0,
                low=9.0,
                close=float(date % 100),
                volume=1000.0,
                amount=10000.0,
            )
            for date in dates
        ]
        sha = store.store_bars(symbol, bars)
        manifest = DatasetManifest(
            dataset_id=dataset_id,
            source="tushare",
            adjustment="none",
            period="1d",
            status="ready",
            data_cutoff_date=max(dates),
            symbols=[
                SymbolRecord(
                    symbol=symbol,
                    blob_sha256=sha,
                    row_count=len(bars),
                    quality="ok",
                    first_date=min(dates),
                    last_date=max(dates),
                )
            ],
            symbol_count=1,
            row_count=len(bars),
        )
        store.publish(manifest)

    publish("tushare_none_1d_stock_test", "SZSE.STK.000001", list(range(20240102, 20240102 + 120)))
    # Index exists but starts later than the queried date.
    publish("tushare_none_1d_index_test", "SSE.IDX.000001", list(range(20240201, 20240201 + 120)))

    cfg = SimpleNamespace(market_data_root=tmp_path / "market_data")
    session = bq.BaguaPlaneSession(cfg, "raw")

    with pytest.raises(FileNotFoundError):
        session.load_symbol("sh000001", asof=20240105)

    # Same-code stock queries in-range still work.
    bars, meta = session.load_symbol("sz000001", asof=20240105)
    assert meta["dataset_id"] == "tushare_none_1d_stock_test"


def test_bagua_plane_signal_family_match_helpers():
    """Document product family gates used by backtest reuse (inline mirror)."""
    def matches(plane, src, adj):
        src = (src or "").strip()
        adj = (adj or "").strip()
        if plane == "tdx_front":
            return src == "tdxquant" and adj in ("", "front")
        if plane == "tushare_qfq":
            if src == "tushare" and adj in ("", "qfq", "tushare_factor_qfq", "composite_tushare_factor_qfq"):
                return True
            if src == "internal" and adj in (
                "tushare_factor_qfq", "composite_tushare_factor_qfq", "qfq", "asof_qfq",
            ):
                return True
        return False

    assert matches("tdx_front", "tdxquant", "front")
    assert matches("tdx_front", "tdxquant", "")
    assert not matches("tdx_front", "tushare", "qfq")
    assert matches("tushare_qfq", "tushare", "qfq")
    assert matches("tushare_qfq", "internal", "tushare_factor_qfq")


def test_batch_query_bagua_multi_codes(monkeypatch, tmp_path):
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")

    bars = [DayBar(20240103, 6.27, 7.33, 5.90, 5.90, 1.0, 1.0)]

    def _fake_load(_cfg, _code, source_key, asof=None):
        return bars, {
            "dataset_id": "mock",
            "dataset_source": "tdxquant",
            "dataset_adjustment": "front",
            "dataset_status": "ready",
            "covers_asof": True,
            "candidate_datasets": 1,
        }

    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=tmp_path,
        tdx_root=tmp_path,
        market_data_root=tmp_path / "md",
        forecast_root=tmp_path,
        forecast_weekly_dir=tmp_path,
        universe_path=tmp_path / "universe.json",
        adj_root=tmp_path,
    )
    monkeypatch.setattr(bq, "_load_dataset_bars", _fake_load)
    # Force no session (market_data missing) so path uses _load_dataset_bars
    monkeypatch.setattr(
        bq,
        "BaguaPlaneSession",
        lambda *_a, **_k: (_ for _ in ()).throw(FileNotFoundError("no md")),
    )

    out = bq.batch_query_bagua(
        cfg,
        codes=["600000", "000001", "600000"],
        date="2024-01-03",
        period="DAY",
        adjust="tushare_qfq",
    )
    assert out["ok"] is True
    assert out["requested"] == 2  # de-duped
    assert out["ok_count"] == 2
    assert out["error_count"] == 0
    assert len(out["results"]) == 2
    assert out["results"][0]["bagua"]["upper_id"] == 7


def test_batch_query_bagua_all_stocks_limit(monkeypatch, tmp_path):
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")

    bars = [DayBar(20240103, 10.0, 11.0, 9.0, 10.5, 1.0, 1.0)]
    monkeypatch.setattr(
        bq,
        "_load_dataset_bars",
        lambda *_a, **_k: (
            bars,
            {
                "dataset_id": "mock",
                "dataset_source": "tdxquant",
                "dataset_adjustment": "front",
                "dataset_status": "ready",
                "covers_asof": True,
                "candidate_datasets": 1,
            },
        ),
    )
    monkeypatch.setattr(
        bq,
        "BaguaPlaneSession",
        lambda *_a, **_k: (_ for _ in ()).throw(FileNotFoundError("no md")),
    )
    def _fake_resolve(cfg, codes=None, *, all_stocks=False):
        assert all_stocks is True
        return ["SSE.STK.600000", "SZSE.STK.000001", "SSE.STK.600519"]

    monkeypatch.setattr(bq, "_resolve_batch_codes", _fake_resolve)

    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=tmp_path,
        tdx_root=tmp_path,
        market_data_root=tmp_path / "md",
        forecast_root=tmp_path,
        forecast_weekly_dir=tmp_path,
        universe_path=tmp_path / "universe.json",
        adj_root=tmp_path,
    )
    out = bq.batch_query_bagua(
        cfg,
        all_stocks=True,
        date=20240103,
        period="DAY",
        adjust="tushare_qfq",
        limit=2,
    )
    assert out["requested"] == 2
    assert out["ok_count"] == 2
    assert out["all_stocks"] is True


def test_export_bagua_multi_period_xlsx(monkeypatch, tmp_path):
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")

    bars = [DayBar(20240103, 6.27, 7.33, 5.90, 5.90, 1.0, 1.0)]
    monkeypatch.setattr(
        bq,
        "_load_dataset_bars",
        lambda *_a, **_k: (
            bars,
            {
                "dataset_id": "mock",
                "dataset_source": "tdxquant",
                "dataset_adjustment": "front",
                "dataset_status": "ready",
                "covers_asof": True,
                "candidate_datasets": 1,
            },
        ),
    )
    monkeypatch.setattr(
        bq,
        "BaguaPlaneSession",
        lambda *_a, **_k: (_ for _ in ()).throw(FileNotFoundError("no md")),
    )

    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=tmp_path,
        tdx_root=tmp_path,
        market_data_root=tmp_path / "md",
        forecast_root=tmp_path,
        forecast_weekly_dir=tmp_path,
        universe_path=tmp_path / "universe.json",
        adj_root=tmp_path,
    )
    monkeypatch.setattr(bq, "load_rizhu_map", lambda _p=None: {"600000": "甲子", "000001": "乙丑"})
    path = bq.export_bagua_multi_period_xlsx(
        cfg,
        date="2024-01-03",
        periods=["WEEK", "MONTH"],
        adjust="tushare_qfq",
        codes=["600000", "000001"],
        all_stocks=False,
    )
    assert path.exists()
    import openpyxl

    wb = openpyxl.load_workbook(path)
    assert "meta" in wb.sheetnames
    assert "stock-all" in wb.sheetnames
    ws = wb["stock-all"]
    headers = [c.value for c in ws[1]]
    assert headers[:8] == ["code", "name", "week_end", "open", "high", "low", "close", "日柱"]
    assert re.match(r"^周卦周线-组合\(\d{4}-W\d{1,2}\)$", headers[8]), headers[8]
    assert headers[9] == "爻辞解释"
    assert re.match(r"^月卦月线-组合\(\d{4}-\d{2}\)$", headers[10]), headers[10]
    assert headers[11] == "爻辞解释"
    assert ws.max_row >= 2
    # code / 日柱 columns
    assert ws.cell(2, 1).value in ("600000", "000001")
    assert ws.cell(2, 8).value in ("甲子", "乙丑", None, "")


def test_export_month_defaults_to_prev_month(monkeypatch, tmp_path):
    """MONTH 列默认取查询月份的上一个月（回归）。

    date=2024-01-15 时 asof_map = {"WEEK": 20240115, "MONTH": 20231231}：
    周卦取查询周（2024-W02），月卦取上一月（2023-12）。
    12 月与 1 月 OHLC 不同（山水蒙 vs 天水讼），若月卦误用当月数据，
    月卦组合列将与周卦列一致而不是 2023-12 的卦象。
    """
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")

    # 2023-12 全部交易日 + 2024-01 月上旬（到 1/12，即 2024 年第 2 周）
    dec_days = [
        20231201, 20231204, 20231205, 20231206, 20231207, 20231208,
        20231211, 20231212, 20231213, 20231214, 20231215,
        20231218, 20231219, 20231220, 20231221, 20231222,
        20231225, 20231226, 20231227, 20231228, 20231229,
    ]
    jan_days = [
        20240102, 20240103, 20240104, 20240105,
        20240108, 20240109, 20240110, 20240111, 20240112,
    ]
    bars = [DayBar(d, 6.27, 7.33, 5.90, 5.90, 1.0, 1.0) for d in dec_days] + [
        DayBar(d, 10.0, 11.0, 9.0, 10.5, 1.0, 1.0) for d in jan_days
    ]

    monkeypatch.setattr(
        bq,
        "_load_dataset_bars",
        lambda *_a, **_k: (
            bars,
            {
                "dataset_id": "mock",
                "dataset_source": "tdxquant",
                "dataset_adjustment": "front",
                "dataset_status": "ready",
                "covers_asof": True,
                "candidate_datasets": 1,
            },
        ),
    )
    monkeypatch.setattr(
        bq,
        "BaguaPlaneSession",
        lambda *_a, **_k: (_ for _ in ()).throw(FileNotFoundError("no md")),
    )
    monkeypatch.setattr(bq, "load_rizhu_map", lambda _p=None: {"600000": "甲子"})

    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=tmp_path,
        tdx_root=tmp_path,
        market_data_root=tmp_path / "md",
        forecast_root=tmp_path,
        forecast_weekly_dir=tmp_path,
        universe_path=tmp_path / "universe.json",
        adj_root=tmp_path,
    )

    # 基准：直接用上一月最后一天查 MONTH，导出文件的月卦列应与其一致
    ref = bq.query_bagua(
        cfg, code="600000", date="2023-12-31", period="MONTH", adjust="tushare_qfq"
    )
    assert ref["ok"] is True
    ref_full = ref["summary"]["full_name"]
    assert ref_full
    # 导出组合已剥离卦符（U+4DC0–U+4DFF），基准 full_name 同步剥离后再比较
    ref_full_plain = re.sub(r"[\u4dc0-\u4dff]", "", ref_full)
    expected_month_combo = bq._bagua_combo(ref)

    path = bq.export_bagua_multi_period_xlsx(
        cfg,
        date="2024-01-15",
        periods=["WEEK", "MONTH"],
        adjust="tushare_qfq",
        codes=["600000"],
        all_stocks=False,
    )
    assert path.exists()

    import openpyxl

    wb = openpyxl.load_workbook(path)
    assert "meta" in wb.sheetnames
    assert "stock-all" in wb.sheetnames
    ws = wb["stock-all"]
    headers = [c.value for c in ws[1]]
    assert len(headers) == 13
    assert headers[:8] == ["code", "name", "week_end", "open", "high", "low", "close", "日柱"]
    # 周卦列在前(9)、月卦列在后(11)，标签分别含查询周 ISO 周与上一月
    assert headers[8].startswith("周卦周线-组合(") and "2024-W02" in headers[8], headers[8]
    assert headers[9] == "爻辞解释"
    assert headers[10].startswith("月卦月线-组合(") and "2023-12" in headers[10], headers[10]
    assert headers[11] == "爻辞解释"
    # 末列为数据状态（失败行写 data_status/error_reason，正常行为空）
    assert headers[12] == "数据状态"

    # 月卦数据确实取 2023-12：组合与 2023-12-31 的 MONTH 查询完全一致，
    # 且与周卦列（2024-W02 的天水讼）不同
    week_combo = ws.cell(2, 9).value or ""
    month_combo = ws.cell(2, 11).value or ""
    assert week_combo, "周卦组合不应为空"
    assert month_combo, "月卦组合不应为空"
    assert month_combo.startswith(ref_full_plain), f"{month_combo!r} 应以 {ref_full_plain!r} 开头"
    assert month_combo == expected_month_combo, f"{month_combo!r} != {expected_month_combo!r}"
    assert week_combo != month_combo

    # 周卦 week_end 落在 2024-01 第 2 周（2024-01-08 ~ 2024-01-14）
    week_end = str(ws.cell(2, 3).value or "")
    assert "2024-01-08" <= week_end <= "2024-01-14", week_end

    # meta 记录 month_asof = 20231231
    meta = {r[0]: r[1] for r in wb["meta"].iter_rows(min_row=2, values_only=True)}
    assert meta.get("month_asof") == 20231231
    assert meta.get("query_date") == 20240115


def test_export_all_stocks_two_sheets_and_no_gua_symbol(monkeypatch, tmp_path):
    """全市场导出：同一 Excel 两个 sheet（stock-all / etf-all），卦象组合不含卦符。"""
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")

    stock_bars = [DayBar(20240103, 6.27, 7.33, 5.90, 5.90, 1.0, 1.0)]
    etf_days = [
        20231204, 20231205, 20231206, 20231207, 20231208,
        20231211, 20231212, 20231213, 20231214, 20231215,
        20231218, 20231219, 20231220, 20231221, 20231222,
        20231225, 20231226, 20231227, 20231228, 20231229,
        20240102, 20240103, 20240104, 20240105,
    ]
    # 两只 ETF 用不同的 OHLC，确保各自独立算出卦象（而非共用同一行数据）
    etf_bars_by_code = {
        "SSE.ETF.510300": [DayBar(d, 4.05, 4.12, 3.98, 4.06, 1.0, 1.0) for d in etf_days],
        "SZSE.ETF.159915": [DayBar(d, 2.35, 2.44, 2.31, 2.40, 1.0, 1.0) for d in etf_days],
    }

    def _fake_load(cfg, std_code, source_key, asof=None):
        if std_code.startswith(("SSE.ETF.", "SZSE.ETF.")):
            bars = etf_bars_by_code.get(std_code) or list(etf_bars_by_code.values())[0]
        else:
            bars = stock_bars
        return (
            bars,
            {
                "dataset_id": "mock",
                "dataset_source": "tdxquant",
                "dataset_adjustment": "front",
                "dataset_status": "ready",
                "covers_asof": True,
                "candidate_datasets": 1,
            },
        )

    monkeypatch.setattr(bq, "_load_dataset_bars", _fake_load)
    monkeypatch.setattr(
        bq,
        "BaguaPlaneSession",
        lambda *_a, **_k: (_ for _ in ()).throw(FileNotFoundError("no md")),
    )
    monkeypatch.setattr(
        bq,
        "_resolve_batch_codes",
        lambda cfg, codes=None, *, all_stocks=False: ["SSE.STK.600000"],
    )
    monkeypatch.setattr(
        bq, "list_etf_std_codes", lambda cfg: ["SSE.ETF.510300", "SZSE.ETF.159915"]
    )

    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=tmp_path,
        tdx_root=tmp_path,
        market_data_root=tmp_path / "md",
        forecast_root=tmp_path,
        forecast_weekly_dir=tmp_path,
        universe_path=tmp_path / "universe.json",
        adj_root=tmp_path,
    )
    path = bq.export_bagua_multi_period_xlsx(
        cfg,
        date="2024-01-15",
        periods=["WEEK", "MONTH"],
        adjust="tushare_qfq",
        all_stocks=True,
    )
    assert path.exists()

    import openpyxl

    wb = openpyxl.load_workbook(path)
    assert "stock-all" in wb.sheetnames
    assert "etf-all" in wb.sheetnames
    assert "meta" in wb.sheetnames
    ws_stock = wb["stock-all"]
    ws_etf = wb["etf-all"]
    assert ws_stock.max_row >= 2
    assert ws_etf.max_row >= 2
    assert len([c.value for c in ws_etf[1]]) == 13
    # 两个 sheet 的所有单元格均不含卦符（U+4DC0–U+4DFF）
    for ws in (ws_stock, ws_etf):
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None and re.search(r"[\u4dc0-\u4dff]", str(cell)):
                    pytest.fail(f"卦符残留: {cell!r}")
    # ETF 行名称兜底解析（resolve_stock_name 无数据时为空字符串）
    assert ws_etf.cell(2, 1).value in ("510300", "159915")

    # ---- 补强：etf-all 两行都写入真实卦象，且不含 stock-all 的股票代码 ----
    etf_rows = list(ws_etf.iter_rows(min_row=2, values_only=True))
    assert len(etf_rows) == 2, f"etf-all 应写入两行，实际 {len(etf_rows)}"
    etf_codes = [r[0] for r in etf_rows]
    assert set(etf_codes) == {"510300", "159915"}, etf_codes
    # 组合列（周卦=第9列、月卦=第11列）均写入真实卦象，非 error 空行
    assert all(r[8] for r in etf_rows), f"etf-all 周卦组合列存在空行: {etf_rows}"
    assert all(r[10] for r in etf_rows), f"etf-all 月卦组合列存在空行: {etf_rows}"
    # 两只 ETF 数据不同 -> 两行卦象组合互不相同（各自独立计算）
    assert etf_rows[0][8] != etf_rows[1][8], "两只 ETF 卦象不应相同"
    # etf-all 不得混入 stock-all 的股票代码
    stock_codes = {
        r[0] for r in ws_stock.iter_rows(min_row=2, values_only=True) if r[0]
    }
    assert stock_codes.isdisjoint(etf_codes), f"etf-all 混入股票代码: {stock_codes & set(etf_codes)}"

    meta = {r[0]: r[1] for r in wb["meta"].iter_rows(min_row=2, values_only=True)}
    assert meta.get("stock_count") == 1
    assert meta.get("etf_count") == 2
    assert meta.get("sheets") == "stock-all,etf-all"


def test_export_etf_only_codes_skips_empty_stock_sheet(monkeypatch, tmp_path):
    """手动 codes 全为 ETF：导出只有 etf-all + meta 两个 sheet（无空 stock-all）。"""
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")

    etf_days = [
        20231204, 20231205, 20231206, 20231207, 20231208,
        20231211, 20231212, 20231213, 20231214, 20231215,
        20231218, 20231219, 20231220, 20231221, 20231222,
        20231225, 20231226, 20231227, 20231228, 20231229,
        20240102, 20240103, 20240104, 20240105,
    ]
    etf_bars = [DayBar(d, 4.05, 4.12, 3.98, 4.06, 1.0, 1.0) for d in etf_days]
    monkeypatch.setattr(
        bq,
        "_load_dataset_bars",
        lambda *_a, **_k: (
            etf_bars,
            {
                "dataset_id": "mock",
                "dataset_source": "tdxquant",
                "dataset_adjustment": "front",
                "dataset_status": "ready",
                "covers_asof": True,
                "candidate_datasets": 1,
            },
        ),
    )
    monkeypatch.setattr(
        bq,
        "BaguaPlaneSession",
        lambda *_a, **_k: (_ for _ in ()).throw(FileNotFoundError("no md")),
    )
    # 不应调用 _resolve_batch_codes（没有股票 codes）；若被调用会断言失败
    def _unexpected_resolve(cfg, codes=None, *, all_stocks=False):
        raise AssertionError("全 ETF codes 不应触发股票解析")

    monkeypatch.setattr(bq, "_resolve_batch_codes", _unexpected_resolve)

    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=tmp_path,
        tdx_root=tmp_path,
        market_data_root=tmp_path / "md",
        forecast_root=tmp_path,
        forecast_weekly_dir=tmp_path,
        universe_path=tmp_path / "universe.json",
        adj_root=tmp_path,
    )
    path = bq.export_bagua_multi_period_xlsx(
        cfg,
        date="2024-01-15",
        periods=["WEEK", "MONTH"],
        adjust="tushare_qfq",
        codes=["sh510300"],
        all_stocks=False,
    )
    assert path.exists()

    import openpyxl

    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {"meta", "etf-all"}, wb.sheetnames
    ws_etf = wb["etf-all"]
    assert ws_etf.max_row >= 2
    etf_rows = list(ws_etf.iter_rows(min_row=2, values_only=True))
    assert len(etf_rows) == 1
    assert etf_rows[0][0] == "510300"
    assert etf_rows[0][8], "周卦组合列不应为空"
    meta = {r[0]: r[1] for r in wb["meta"].iter_rows(min_row=2, values_only=True)}
    assert meta.get("stock_count") == 0
    assert meta.get("etf_count") == 1


def test_export_mixed_codes_with_invalid_dropped(monkeypatch, tmp_path):
    """手工 codes 混入无法识别的代码（如 "garbage"）时静默丢弃，导出仍成功（回归）。

    修复前：无效代码进入股票池后 ``_resolve_batch_codes`` 抛
    "no valid stock codes"，导致整个导出失败（ETF+无效代码混输无法导出）。
    修复后：进入股票池前先经 ``normalize_query_code`` 校验，失败则丢弃。
    ``_resolve_batch_codes`` 在手工 codes 路径收到的是已过滤后的有效代码
    （仅字符串规范化、无文件系统访问），这里让它真实执行以覆盖完整链路。
    """
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")

    stock_bars = [DayBar(20240103, 6.27, 7.33, 5.90, 5.90, 1.0, 1.0)]
    etf_days = [
        20231204, 20231205, 20231206, 20231207, 20231208,
        20231211, 20231212, 20231213, 20231214, 20231215,
        20231218, 20231219, 20231220, 20231221, 20231222,
        20231225, 20231226, 20231227, 20231228, 20231229,
        20240102, 20240103, 20240104, 20240105,
    ]
    etf_bars = [DayBar(d, 4.05, 4.12, 3.98, 4.06, 1.0, 1.0) for d in etf_days]

    def _fake_load(cfg, std_code, source_key, asof=None):
        if std_code == "SSE.ETF.510300":
            bars = etf_bars
        else:
            bars = stock_bars
        return (
            bars,
            {
                "dataset_id": "mock",
                "dataset_source": "tdxquant",
                "dataset_adjustment": "front",
                "dataset_status": "ready",
                "covers_asof": True,
                "candidate_datasets": 1,
            },
        )

    monkeypatch.setattr(bq, "_load_dataset_bars", _fake_load)
    monkeypatch.setattr(
        bq,
        "BaguaPlaneSession",
        lambda *_a, **_k: (_ for _ in ()).throw(FileNotFoundError("no md")),
    )
    monkeypatch.setattr(bq, "load_rizhu_map", lambda _p=None: {"600000": "甲子", "510300": "丙午"})

    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=tmp_path,
        tdx_root=tmp_path,
        market_data_root=tmp_path / "md",
        forecast_root=tmp_path,
        forecast_weekly_dir=tmp_path,
        universe_path=tmp_path / "universe.json",
        adj_root=tmp_path,
    )
    # 混输：ETF + 无效代码 + 股票。修复前会抛 "no valid stock codes"。
    path = bq.export_bagua_multi_period_xlsx(
        cfg,
        date="2024-01-15",
        periods=["WEEK", "MONTH"],
        adjust="tushare_qfq",
        codes=["sh510300", "garbage", "600000"],
        all_stocks=False,
    )
    assert path.exists()

    import openpyxl

    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {"meta", "stock-all", "etf-all"}, wb.sheetnames
    ws_stock = wb["stock-all"]
    ws_etf = wb["etf-all"]
    stock_rows = list(ws_stock.iter_rows(min_row=2, values_only=True))
    etf_rows = list(ws_etf.iter_rows(min_row=2, values_only=True))
    # stock-all 有 600000 行、etf-all 有 510300 行，各仅一行（无多余行）
    assert len(stock_rows) == 1, f"stock-all 应只有 600000 一行，实际 {len(stock_rows)}"
    assert len(etf_rows) == 1, f"etf-all 应只有 510300 一行，实际 {len(etf_rows)}"
    assert stock_rows[0][0] == "600000"
    assert etf_rows[0][0] == "510300"
    # "garbage" 未出现在任何 sheet 的 code 列
    for ws in (ws_stock, ws_etf):
        codes_col = [str(r[0] or "") for r in ws.iter_rows(min_row=2, values_only=True)]
        assert "garbage" not in codes_col, f"{ws.title} code 列残留无效代码: {codes_col}"
    # 全表（含 name 等列）也不应残留 garbage 字样
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None and "garbage" in str(cell).lower():
                    pytest.fail(f"无效代码残留: {cell!r}")

    meta = {r[0]: r[1] for r in wb["meta"].iter_rows(min_row=2, values_only=True)}
    assert meta.get("requested") == 2
    assert meta.get("stock_count") == 1
    assert meta.get("etf_count") == 1
    assert meta.get("sheets") == "stock-all,etf-all"

    # 更严格的回归场景：ETF + 仅无效代码（股票池全部无效）。
    # 修复前 "garbage" 会无过滤进入股票池，_resolve_batch_codes 对全无效
    # 池抛 ValueError("no valid stock codes")，导致整个导出失败。
    path2 = bq.export_bagua_multi_period_xlsx(
        cfg,
        date="2024-01-15",
        periods=["WEEK", "MONTH"],
        adjust="tushare_qfq",
        codes=["sh510300", "garbage"],
        all_stocks=False,
    )
    assert path2.exists()

    wb2 = openpyxl.load_workbook(path2)
    assert set(wb2.sheetnames) == {"meta", "etf-all"}, wb2.sheetnames
    etf_rows2 = list(wb2["etf-all"].iter_rows(min_row=2, values_only=True))
    assert len(etf_rows2) == 1
    assert etf_rows2[0][0] == "510300"
    for row in wb2["etf-all"].iter_rows(values_only=True):
        for cell in row:
            if cell is not None and "garbage" in str(cell).lower():
                pytest.fail(f"无效代码残留: {cell!r}")
    meta2 = {r[0]: r[1] for r in wb2["meta"].iter_rows(min_row=2, values_only=True)}
    assert meta2.get("stock_count") == 0
    assert meta2.get("etf_count") == 1
    assert meta2.get("sheets") == "etf-all"


# ---- 日柱补齐：上市日期推算 60 甲子 ----

def test_rizhu_from_list_date_known_values():
    """60 甲子推算与 日柱(1).xlsx 已知值对拍（口径=上市日当天干支）。"""
    cases = [
        (18991222, "甲子"),  # 锚点
        (19911104, "戊寅"),  # 已知对拍值
        (19910403, "癸卯"),  # 000001 平安银行
        (19991110, "丙寅"),  # 600000 浦发银行
        (20050223, "戊寅"),  # 510050 上证50ETF
        (20120528, "己丑"),  # 510300 沪深300ETF
        (20201116, "癸亥"),  # 588000 科创50ETF
        (20250613, "癸丑"),  # 688795 摩尔线程
        (20250620, "庚申"),  # 600930 华电新能
    ]
    for ymd, expect in cases:
        assert bq._rizhu_from_list_date(ymd) == expect, (ymd, expect)


def test_ensure_rizhu_coverage_cache_hit_no_fetch(monkeypatch, tmp_path):
    """缓存命中时不再调 Tushare，直接推算补齐缺失代码。"""
    monkeypatch.setattr(bq, "_SYMBOL_META_CACHE", {})
    monkeypatch.setattr(bq, "_rizhu_list_dates_cache_path", lambda: tmp_path / "c.json")
    monkeypatch.setattr(
        bq,
        "_load_symbol_meta_cache",
        lambda: ({"600930": 20250620}, {"510300": 20120528}, {}, {}),
    )
    called = {"n": 0}

    def boom(*_a, **_k):
        called["n"] += 1
        raise AssertionError("不应调用 Tushare")

    monkeypatch.setattr(bq, "_fetch_symbol_meta_from_tushare", boom)
    got = bq.ensure_rizhu_coverage(
        SimpleNamespace(), ["600930", "510300", "600000"], {"600000": "甲子"}
    )
    assert got == {"600930": "庚申", "510300": "己丑"}
    assert called["n"] == 0


def test_ensure_rizhu_coverage_fetch_and_persist(monkeypatch, tmp_path):
    """缓存缺失时拉 Tushare 并落盘，之后走缓存。"""
    monkeypatch.setattr(bq, "_SYMBOL_META_CACHE", {})
    cache_file = tmp_path / "c.json"
    monkeypatch.setattr(bq, "_rizhu_list_dates_cache_path", lambda: cache_file)
    monkeypatch.setattr(
        bq,
        "_fetch_symbol_meta_from_tushare",
        lambda _cfg: ({"600930": 20250620}, {"510300": 20120528}, {}, {}),
    )
    got = bq.ensure_rizhu_coverage(SimpleNamespace(), ["600930", "510300"], {})
    assert got == {"600930": "庚申", "510300": "己丑"}
    # 已持久化，二次调用不再触发网络
    monkeypatch.setattr(
        bq, "_fetch_symbol_meta_from_tushare",
        lambda _cfg: (_ for _ in ()).throw(AssertionError("不应二次拉取")),
    )
    got2 = bq.ensure_rizhu_coverage(SimpleNamespace(), ["600930", "510300"], {})
    assert got2 == {"600930": "庚申", "510300": "己丑"}


def test_ensure_rizhu_coverage_tushare_failure_degrades(monkeypatch, tmp_path):
    """Tushare 不可用时静默降级：已缓存的部分返回，其余放弃，不抛异常。"""
    monkeypatch.setattr(bq, "_SYMBOL_META_CACHE", {})
    monkeypatch.setattr(bq, "_rizhu_list_dates_cache_path", lambda: tmp_path / "c.json")
    monkeypatch.setattr(
        bq,
        "_load_symbol_meta_cache",
        lambda: ({"600930": 20250620}, {}, {}, {}),
    )
    monkeypatch.setattr(
        bq, "_fetch_symbol_meta_from_tushare",
        lambda _cfg: (_ for _ in ()).throw(RuntimeError("network down")),
    )
    got = bq.ensure_rizhu_coverage(
        SimpleNamespace(), ["600930", "510300", "000001"], {"000001": "癸卯"}
    )
    assert got == {"600930": "庚申"}  # 仅缓存覆盖的


def test_export_fills_rizhu_for_new_stock_and_etf(monkeypatch, tmp_path):
    """导出时次新股与 ETF 的日柱列被补齐（Excel 表外代码）。"""
    if not JSON_PATH.exists():
        pytest.skip("bagua_384.json missing")
    bars = [DayBar(20240103, 6.27, 7.33, 5.90, 5.90, 1.0, 1.0)]
    monkeypatch.setattr(
        bq,
        "_load_dataset_bars",
        lambda *_a, **_k: (
            bars,
            {
                "dataset_id": "mock",
                "dataset_source": "tdxquant",
                "dataset_adjustment": "front",
                "dataset_status": "ready",
                "covers_asof": True,
                "candidate_datasets": 1,
            },
        ),
    )
    monkeypatch.setattr(
        bq,
        "BaguaPlaneSession",
        lambda *_a, **_k: (_ for _ in ()).throw(FileNotFoundError("no md")),
    )
    monkeypatch.setattr(bq, "load_rizhu_map", lambda _p=None: {"000001": "癸卯"})
    monkeypatch.setattr(bq, "_SYMBOL_META_CACHE", {})
    monkeypatch.setattr(bq, "_rizhu_list_dates_cache_path", lambda: tmp_path / "c.json")
    monkeypatch.setattr(
        bq,
        "_fetch_symbol_meta_from_tushare",
        lambda _cfg: (
            {"600930": 20250620},
            {"510300": 20120528},
            {"600930": "华电新能"},
            {"510300": "沪深300ETF"},
        ),
    )
    cfg = SimpleNamespace(
        bagua_json=JSON_PATH,
        storage_root=tmp_path,
        tdx_root=tmp_path,
        market_data_root=tmp_path / "md",
        forecast_root=tmp_path,
        forecast_weekly_dir=tmp_path,
        universe_path=tmp_path / "universe.json",
        adj_root=tmp_path,
    )
    path = bq.export_bagua_multi_period_xlsx(
        cfg,
        date="2024-01-03",
        periods=["WEEK", "MONTH"],
        adjust="tushare_qfq",
        codes=["600930", "sh510300"],
        all_stocks=False,
    )
    assert path.exists()
    import openpyxl

    wb = openpyxl.load_workbook(path)
    rizhu_by_code = {}
    name_by_code = {}
    for sn in ("stock-all", "etf-all"):
        for row in wb[sn].iter_rows(min_row=2, values_only=True):
            rizhu_by_code[row[0]] = row[7]  # 日柱列
            name_by_code[row[0]] = row[1]  # name 列
    assert rizhu_by_code.get("600930") == "庚申", rizhu_by_code
    assert rizhu_by_code.get("510300") == "己丑", rizhu_by_code
    # name 列也被同一份 Tushare 元数据补齐
    assert name_by_code.get("600930") == "华电新能", name_by_code
    assert name_by_code.get("510300") == "沪深300ETF", name_by_code
    meta = {r[0]: r[1] for r in wb["meta"].iter_rows(min_row=2, values_only=True)}
    assert meta.get("rizhu_note", "").startswith("Excel 日柱表优先"), meta
    assert meta.get("name_note", "").startswith("name 列优先本地"), meta


def test_ensure_name_coverage_fetches_and_skips_existing(monkeypatch, tmp_path):
    """name 补齐：已有名字的代码跳过，缺的按 Tushare 元数据补齐。"""
    monkeypatch.setattr(bq, "_SYMBOL_META_CACHE", {})
    monkeypatch.setattr(bq, "_rizhu_list_dates_cache_path", lambda: tmp_path / "c.json")
    monkeypatch.setattr(
        bq,
        "_fetch_symbol_meta_from_tushare",
        lambda _cfg: (
            {"600930": 20250620},
            {"510300": 20120528},
            {"600930": "华电新能"},
            {"510300": "沪深300ETF"},
        ),
    )
    got = bq.ensure_name_coverage(
        SimpleNamespace(), ["600930", "510300", "600000"], {"600000": "浦发银行"}
    )
    assert got == {"600930": "华电新能", "510300": "沪深300ETF"}
    assert "600000" not in got  # 已有名字，跳过


def test_load_symbol_meta_cache_v1_schema_compat(monkeypatch, tmp_path):
    """旧 schema v1 缓存（只有 list_date 无 name）仍可加载，name 缺省为空。"""
    import json

    monkeypatch.setattr(bq, "_SYMBOL_META_CACHE", {})
    cache_file = tmp_path / "c.json"
    cache_file.write_text(
        json.dumps(
            {"schema_version": 1, "stocks": {"600000": 19991110}, "etfs": {}}
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(bq, "_rizhu_list_dates_cache_path", lambda: cache_file)
    stocks, etfs, snames, enames = bq._load_symbol_meta_cache()
    assert stocks == {"600000": 19991110}
    assert snames == {}
    assert enames == {}


def test_export_etf_pool_uses_newest_surface_only(tmp_path, monkeypatch):
    """导出 ETF 池只取最新就绪面：旧快照里的 LOF 污染不得再进池。

    旧实现是对全部 ready tushare/none manifest 求并集——上游 universe
    修复后，历史快照残留的 LOF 污染（.ETF. 标签但实为 LOF）仍会永久
    占据导出池。改为按 cutoff/created_at 取最新单面后，新干净 manifest
    落地即自动清除污染。
    """
    from wtpy.apps.astock.config import get_default_config
    from wtpy.apps.astock.data.dataset_store import (
        DatasetManifest,
        DatasetStore,
        SymbolRecord,
    )
    from wtpy.apps.astock.data.providers.base import MarketBar

    def _bars(sym, dates, base):
        return [
            MarketBar(symbol=sym, trade_date=d, period="1d", open=base,
                      high=base + 0.1, low=base - 0.1, close=base + 0.05,
                      volume=1000.0, amount=100000.0,
                      source="tushare", adjustment="none")
            for d in dates
        ]

    store = DatasetStore(tmp_path)

    def _publish(ds_id, spec, cutoff, created_at):
        recs = []
        for sym, dates, base in spec:
            sha = store.store_bars(sym, _bars(sym, dates, base))
            recs.append(SymbolRecord(symbol=sym, blob_sha256=sha,
                                     first_date=dates[0], last_date=dates[-1],
                                     row_count=len(dates), quality="ok"))
        m = DatasetManifest(dataset_id=ds_id, source="tushare",
                            adjustment="none", period="1d",
                            data_cutoff_date=cutoff, snapshot_date=cutoff,
                            provider_version="test", status="ready",
                            created_at=created_at)
        m.symbols = recs
        m.symbol_count = len(recs)
        m.row_count = sum(r.row_count for r in recs)
        m.expected_symbol_count = len(recs)
        m.imported_symbol_count = len(recs)
        m.coverage_ratio = 1.0
        store.publish(m)

    old_dates = [20240101 + i for i in range(10)]
    new_dates = [20240101 + i for i in range(20)]
    # 旧面：含 510300 与 LOF 污染 161725；新面：只有干净的 510300
    _publish("tushare_none_1d_old_surface",
             [("SSE.ETF.510300", old_dates, 4.0),
              ("SZSE.ETF.161725", old_dates, 0.8)],
             cutoff=old_dates[-1], created_at="2024-01-11T18:00:00")
    _publish("tushare_none_1d_new_surface",
             [("SSE.ETF.510300", new_dates, 4.0)],
             cutoff=new_dates[-1], created_at="2024-01-21T18:00:00")

    monkeypatch.setenv("MARKET_DATA_ROOT", str(tmp_path))
    cfg = get_default_config()
    pool = bq._enumerate_export_etf_pool(cfg)
    assert pool == ["SSE.ETF.510300"]


def test_export_etf_pool_rejects_pollution_and_partial_surfaces(
    tmp_path, monkeypatch
):
    """历史伪 .ETF. 标签 + 同 cutoff 污染面 + 新 cutoff 单标的面 三重防线。

    复刻服务器实测形态：
      - 旧污染面：cutoff=20240120，含 510300 与伪标签 161725/150001（LOF）
      - 新干净面：同 cutoff、创建更晚、只有真 ETF → 必须选新干净面
        （created_at 排在数量之前，否则污染大面永远获胜）
      - 局部残片：cutoff 更新但只有 1 只 → 完整度门槛排除
    """
    from wtpy.apps.astock.config import get_default_config
    from wtpy.apps.astock.data.dataset_store import (
        DatasetManifest,
        DatasetStore,
        SymbolRecord,
    )
    from wtpy.apps.astock.data.providers.base import MarketBar

    def _bars(sym, dates, base):
        return [
            MarketBar(symbol=sym, trade_date=d, period="1d", open=base,
                      high=base + 0.1, low=base - 0.1, close=base + 0.05,
                      volume=1000.0, amount=100000.0,
                      source="tushare", adjustment="none")
            for d in dates
        ]

    store = DatasetStore(tmp_path)

    def _publish(ds_id, spec, cutoff, created_at):
        recs = []
        for sym, dates, base in spec:
            sha = store.store_bars(sym, _bars(sym, dates, base))
            recs.append(SymbolRecord(symbol=sym, blob_sha256=sha,
                                     first_date=dates[0], last_date=dates[-1],
                                     row_count=len(dates), quality="ok"))
        m = DatasetManifest(dataset_id=ds_id, source="tushare",
                            adjustment="none", period="1d",
                            data_cutoff_date=cutoff, snapshot_date=cutoff,
                            provider_version="test", status="ready",
                            created_at=created_at)
        m.symbols = recs
        m.symbol_count = len(recs)
        m.row_count = sum(r.row_count for r in recs)
        m.expected_symbol_count = len(recs)
        m.imported_symbol_count = len(recs)
        m.coverage_ratio = 1.0
        store.publish(m)

    old_dates = [20240101 + i for i in range(10)]
    new_dates = [20240101 + i for i in range(20)]
    cutoff = new_dates[-1]
    # 污染面：品种复验后仍有 **4 个有效 ETF**（>干净面的 3 个）——只有让
    # 两个面都通过完整度门槛，才能真正验证排序是 created_at 优先于数量。
    # 若错误地改回 (cutoff, count, created_at)，本测试必然失败。
    _publish("tushare_none_1d_polluted_surface",
             [("SSE.ETF.510300", new_dates, 4.0),
              ("SSE.ETF.510050", new_dates, 2.1),
              ("SSE.ETF.561830", new_dates, 1.2),
              ("SSE.ETF.511990", new_dates, 6.0),
              ("SZSE.ETF.161725", old_dates, 0.8),   # LOF 伪标签
              ("SZSE.ETF.150001", old_dates, 0.7)],  # LOF 伪标签
             cutoff=cutoff, created_at="2024-01-21T18:00:00")
    # 干净面：3 个真 ETF，创建更晚——必须凭 created_at 胜出
    _publish("tushare_none_1d_clean_surface",
             [("SSE.ETF.510300", new_dates, 4.0),
              ("SZSE.ETF.159915", new_dates, 2.5),
              ("SZSE.ETF.158012", new_dates, 1.3)],
             cutoff=cutoff, created_at="2024-01-21T20:00:00")
    # 局部残片：cutoff 更高但仅 1 只 → 完整度门槛排除
    _publish("tushare_none_1d_partial_surface",
             [("SSE.ETF.510300", new_dates + [20240121, 20240122], 4.0)],
             cutoff=20240122, created_at="2024-01-23T18:00:00")

    monkeypatch.setenv("MARKET_DATA_ROOT", str(tmp_path))
    cfg = get_default_config()
    pool = bq._enumerate_export_etf_pool(cfg)
    # 选中同 cutoff 但创建更晚的干净面；伪标签不出现
    assert pool == ["SSE.ETF.510300", "SZSE.ETF.158012", "SZSE.ETF.159915"]
    assert "SZSE.ETF.161725" not in pool
    assert "SZSE.ETF.150001" not in pool
    assert "SSE.ETF.511990" not in pool  # 污染面整体落选，非逐只拼合


def test_export_etf_pool_prefers_authoritative_pointer(tmp_path, monkeypatch):
    """指针有效时直接采用权威面，即使启发式会选更新 cutoff 的其他面。"""
    from wtpy.apps.astock.config import get_default_config
    from wtpy.apps.astock.data.dataset_store import (
        DatasetManifest,
        DatasetStore,
        SymbolRecord,
    )
    from wtpy.apps.astock.data.io_util import atomic_write_json
    from wtpy.apps.astock.data.providers.base import MarketBar

    old_dates = [20240101 + i for i in range(10)]
    newer_dates = [20240105 + i for i in range(20)]

    def _bars(sym, dates, base):
        return [
            MarketBar(symbol=sym, trade_date=d, period="1d", open=base,
                      high=base + 0.1, low=base - 0.1, close=base + 0.05,
                      volume=1000.0, amount=100000.0,
                      source="tushare", adjustment="none")
            for d in dates
        ]

    store = DatasetStore(tmp_path)

    def _publish(ds_id, spec, cutoff, created_at):
        recs = []
        for sym, dates, base in spec:
            sha = store.store_bars(sym, _bars(sym, dates, base))
            recs.append(SymbolRecord(symbol=sym, blob_sha256=sha,
                                     first_date=dates[0], last_date=dates[-1],
                                     row_count=len(dates), quality="ok"))
        m = DatasetManifest(dataset_id=ds_id, source="tushare",
                            adjustment="none", period="1d",
                            data_cutoff_date=cutoff, snapshot_date=cutoff,
                            provider_version="test", status="ready",
                            created_at=created_at)
        m.symbols = recs
        m.symbol_count = len(recs)
        m.row_count = sum(r.row_count for r in recs)
        m.expected_symbol_count = len(recs)
        m.imported_symbol_count = len(recs)
        m.coverage_ratio = 1.0
        store.publish(m)
        return m

    # 权威面：cutoff 较旧但经完整 universe 同步验证（指针指向它）
    target = _publish("tushare_none_1d_verified_full",
                      [("SSE.ETF.510300", old_dates, 4.0),
                       ("SZSE.ETF.158012", old_dates, 1.3)],
                      cutoff=old_dates[-1],
                      created_at="2024-01-11T18:00:00")
    # 启发式会因更高 cutoff 选中的局部面（仅 1 只）
    _publish("tushare_none_1d_partial_newer",
             [("SSE.ETF.510300", newer_dates, 4.0)],
             cutoff=newer_dates[-1], created_at="2024-01-26T18:00:00")

    pointer = {
        "dataset_id": target.dataset_id,
        "manifest_sha256": target.manifest_sha256,
        "data_cutoff_date": int(target.data_cutoff_date or 0),
        "coverage": 0.95,
    }
    atomic_write_json(tmp_path / "etf_surface_pointer.json", pointer)

    monkeypatch.setenv("MARKET_DATA_ROOT", str(tmp_path))
    cfg = get_default_config()
    assert bq._enumerate_export_etf_pool(cfg) == [
        "SSE.ETF.510300", "SZSE.ETF.158012"
    ]

    # 指针 sha 失配（manifest 被替换）→ 回退启发式，不盲用失效指针
    bad = dict(pointer, manifest_sha256="deadbeef")
    atomic_write_json(tmp_path / "etf_surface_pointer.json", bad)
    assert bq._enumerate_export_etf_pool(cfg) == ["SSE.ETF.510300"]

    # fail-closed：缺失 manifest_sha256 的指针一律视为无效（权威面必须
    # 可验证内容身份，不允许无 sha 的"口头权威"）
    atomic_write_json(
        tmp_path / "etf_surface_pointer.json",
        {"dataset_id": target.dataset_id, "data_cutoff_date": 20240110},
    )
    assert bq._enumerate_export_etf_pool(cfg) == ["SSE.ETF.510300"]


def test_export_etf_pool_rejects_tampered_manifest_content(tmp_path, monkeypatch):
    """篡改 manifest 文件内容但保留声明 sha → 指针面必须被拒。

    复核复现的攻击面：只比对 manifest 对象的 manifest_sha256 属性与
    指针值，挡不住"改磁盘文件内容、不动声明字段"的篡改——声明 sha
    与文件实算 canonical hash 不一致时，指针与启发式候选都必须
    fail-closed。此处直接改写磁盘上的 symbols 列表注入一只不存在的
    ETF，验证导出池不会带出它。
    """
    import json as _json

    from wtpy.apps.astock.config import get_default_config
    from wtpy.apps.astock.data.dataset_store import DatasetStore
    from wtpy.apps.astock.data.io_util import atomic_write_json
    from wtpy.apps.astock.data.providers.base import MarketBar

    dates = [20240101 + i for i in range(10)]
    store = DatasetStore(tmp_path)

    def _publish(ds_id, symbols):
        recs = []
        for sym in symbols:
            bars = [
                MarketBar(symbol=sym, trade_date=d, period="1d", open=1.0,
                          high=1.1, low=0.9, close=1.05, volume=1000.0,
                          amount=100000.0, source="tushare",
                          adjustment="none")
                for d in dates
            ]
            from wtpy.apps.astock.data.dataset_store import SymbolRecord

            sha = store.store_bars(sym, bars)
            recs.append(SymbolRecord(symbol=sym, blob_sha256=sha,
                                     first_date=dates[0],
                                     last_date=dates[-1],
                                     row_count=len(dates), quality="ok"))
        from wtpy.apps.astock.data.dataset_store import DatasetManifest

        m = DatasetManifest(dataset_id=ds_id, source="tushare",
                            adjustment="none", period="1d",
                            data_cutoff_date=dates[-1],
                            snapshot_date=dates[-1], provider_version="test",
                            status="ready", created_at="2024-01-11T18:00:00")
        m.symbols = recs
        m.symbol_count = len(recs)
        m.row_count = sum(r.row_count for r in recs)
        m.expected_symbol_count = len(recs)
        m.imported_symbol_count = len(recs)
        m.coverage_ratio = 1.0
        store.publish(m)
        return m

    target = _publish("tushare_none_1d_pointer_target",
                      ["SSE.ETF.510300", "SZSE.ETF.158012"])
    _publish("tushare_none_1d_fallback",
             ["SSE.ETF.510300", "SZSE.ETF.159915"])

    pointer = {
        "dataset_id": target.dataset_id,
        "manifest_sha256": target.manifest_sha256,
        "data_cutoff_date": int(target.data_cutoff_date or 0),
        "coverage": 0.95,
    }
    atomic_write_json(tmp_path / "etf_surface_pointer.json", pointer)

    # 篡改：把指针面磁盘 manifest 里的 158012 换成不存在的 199999，
    # manifest_sha256 声明字段原样保留
    mpath = tmp_path / "manifests" / f"{target.dataset_id}.json"
    payload = _json.loads(mpath.read_text(encoding="utf-8"))
    tampered_text = mpath.read_text(encoding="utf-8").replace(
        "SZSE.ETF.158012", "SZSE.ETF.199999"
    )
    assert tampered_text != mpath.read_text(encoding="utf-8"), "篡改必须生效"
    assert _json.loads(tampered_text)["manifest_sha256"] == \
        payload["manifest_sha256"], "声明 sha 必须未变（攻击前提）"
    mpath.write_text(tampered_text, encoding="utf-8")

    monkeypatch.setenv("MARKET_DATA_ROOT", str(tmp_path))
    cfg = get_default_config()
    # 指针面被拒 -> 回退启发式选中未篡改的 fallback 面；
    # 关键断言：篡改注入的 199999 绝不能进入导出池
    assert bq._enumerate_export_etf_pool(cfg) == [
        "SSE.ETF.510300", "SZSE.ETF.159915"
    ]
