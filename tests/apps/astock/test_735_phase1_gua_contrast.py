# -*- coding: utf-8 -*-
"""735 phase1: gua_filter contrast (service-layer filter, not PortfolioBacktester).

Note: gua_filter / bagua allowlist is applied in BacktestService (and study.attach
+ filter_rules) before events reach PortfolioBacktester. Engine itself does not
interpret gua. These tests lock:
1) same Fri->Mon open buy / Wed open sell schedule as 735 smoke
2) filter_events_by_gua_filter disabled vs mock drop filter -> trade/event counts differ
"""
from __future__ import annotations

import tests.apps.astock.conftest  # noqa: F401

from wtpy.apps.astock.bagua.filter_rules import (
    GuaFilter,
    filter_events_by_gua_filter,
)
from wtpy.apps.astock.config import AStockConfig, CostConfig
from wtpy.apps.astock.data.calendar import TradeCalendar
from wtpy.apps.astock.data.tdx_reader import DayBar
from wtpy.apps.astock.strategy import EXIT_REASON_WEEKDAY_EXIT, PortfolioBacktester
from wtpy.apps.astock.study import SignalEvent


def _cfg():
    cfg = AStockConfig()
    cfg.initial_capital = 1_000_000
    cfg.max_weight = 1.0
    cfg.lot_size = 100
    cfg.costs = CostConfig(0, 0, 0, 0, note="test zero costs")
    return cfg


def _calendar_two_weeks():
    return [
        20240105,  # Fri
        20240108,
        20240109,
        20240110,
        20240111,
        20240112,
        20240115,
        20240116,
    ]


def _bars(code: str):
    dates = _calendar_two_weeks()
    bars = {
        code: [
            DayBar(d, 10.0 + i * 0.2, 12, 9, 10.5, 1, 1000) for i, d in enumerate(dates)
        ]
    }
    bars[code][1] = DayBar(20240108, 11.0, 12, 10, 11.5, 1, 1000)
    return bars


class _MockEv:
    """Minimal event for filter_rules (bagua attachment shape)."""

    def __init__(self, std_code, date, bagua):
        self.std_code = std_code
        self.date = date
        self.period = "DAY"
        self.rule_id = "735"
        self.bagua = bagua


def test_gua_filter_only_in_service_layer_not_portfolio_backtester():
    """PortfolioBacktester has no gua_filter arg — filtering is upstream."""
    import inspect

    sig = inspect.signature(PortfolioBacktester.run)
    assert "gua_filter" not in sig.parameters
    assert "with_bagua" not in sig.parameters


def test_735_schedule_gua_disabled_vs_mock_drop_trade_counts_differ():
    """Same technical schedule; mock gua filter drops signal -> fewer trades."""
    code = "SSE.STK.600000"
    dates = _calendar_two_weeks()
    bars = _bars(code)
    cal = TradeCalendar(dates)
    bt = PortfolioBacktester(_cfg(), cal, bars)

    # Events with bagua labels as service would attach before filtering
    all_events = [
        SignalEvent(code, 20240105, "DAY", "735"),
    ]
    # Attach bagua dict for filter path (as attach_bagua would)
    all_events[0].bagua = {
        "state_id": "01-1",
        "gua_order": 1,
        "yao_order": 1,
        "action_signal": "持有",
        "main_hexagram_id": 1,
    }

    gf_off = GuaFilter(enabled=False, selection_mode="none")
    gf_drop = GuaFilter(
        enabled=True,
        selection_mode="exact_line",
        selected_state_ids=["99-9"],  # no match -> drop all
    )

    kept_off = filter_events_by_gua_filter(all_events, gf_off)
    kept_drop = filter_events_by_gua_filter(all_events, gf_drop)
    assert len(kept_off) == 1
    assert len(kept_drop) == 0

    common_kw = dict(
        hold=1,
        entry_lag=1,
        period="DAY",
        formal_ok=True,
        _skip_zero_replay=True,
        buy_weekday=1,
        exit_weekday=3,  # Wed open sell (735 matrix cell)
        buy_on="open",
        sell_on="open",
        signal_weekdays=[5],
    )
    res_off = bt.run(list(kept_off), **common_kw)
    res_drop = bt.run(list(kept_drop), **common_kw)

    buys_off = [f for f in res_off.fills if f.side == "BUY"]
    buys_drop = [f for f in res_drop.fills if f.side == "BUY"]
    sells_off = [f for f in res_off.fills if f.side == "SELL"]

    assert len(buys_off) == 1 and buys_off[0].date == 20240108
    assert len(sells_off) == 1 and sells_off[0].date == 20240110
    assert sells_off[0].reason == EXIT_REASON_WEEKDAY_EXIT
    assert len(buys_drop) == 0
    assert len(res_drop.fills) == 0
    assert len(buys_off) != len(buys_drop)


def test_result_config_includes_full_cost_fields():
    """BacktestResult.config.costs must carry full CostConfig for artifact writers."""
    code = "SSE.STK.600000"
    dates = _calendar_two_weeks()
    cfg = _cfg()
    cfg.costs = CostConfig(
        commission_rate=0.0003,
        min_commission=5.0,
        stamp_tax_rate=0.001,
        slippage=0.0,
        note="Example costs only; not user real trading costs.",
    )
    bt = PortfolioBacktester(cfg, TradeCalendar(dates), _bars(code))
    res = bt.run(
        [SignalEvent(code, 20240105, "DAY", "735")],
        hold=1,
        entry_lag=1,
        period="DAY",
        formal_ok=True,
        _skip_zero_replay=True,
        buy_weekday=1,
        exit_weekday=3,
        buy_on="open",
        sell_on="open",
        signal_weekdays=[5],
    )
    costs = (res.config or {}).get("costs") or {}
    for k in (
        "commission_rate",
        "min_commission",
        "stamp_tax_rate",
        "slippage",
        "note",
    ):
        assert k in costs, k
    assert costs["commission_rate"] == 0.0003
    assert costs["min_commission"] == 5.0
    assert costs["stamp_tax_rate"] == 0.001
    assert "Example" in (costs.get("note") or "")


def test_write_backtest_csv_run_meta_and_excel_costs(tmp_path):
    """run_meta.json + Excel summary include full cost fields (P1.7)."""
    from wtpy.apps.astock.reports import write_backtest_csv
    import json

    code = "SSE.STK.600000"
    dates = _calendar_two_weeks()
    cfg = _cfg()
    cfg.costs = CostConfig(
        commission_rate=0.0003,
        min_commission=5.0,
        stamp_tax_rate=0.001,
        slippage=0.01,
        note="traceability-test",
    )
    bt = PortfolioBacktester(cfg, TradeCalendar(dates), _bars(code))
    res = bt.run(
        [SignalEvent(code, 20240105, "DAY", "735")],
        hold=1,
        entry_lag=1,
        period="DAY",
        formal_ok=True,
        _skip_zero_replay=True,
        buy_weekday=1,
        exit_weekday=3,
        buy_on="open",
        sell_on="open",
        signal_weekdays=[5],
    )
    meta = {
        "title": "735 cost meta test",
        "config": res.config,
        "costs": res.config.get("costs"),
        "period": "DAY",
        "hold": 1,
    }
    paths = write_backtest_csv(tmp_path, res, meta=meta)
    meta_path = paths["meta"]
    data = json.loads(meta_path.read_text(encoding="utf-8"))
    assert "costs" in data
    for k in (
        "commission_rate",
        "min_commission",
        "stamp_tax_rate",
        "slippage",
        "note",
    ):
        assert k in data["costs"], k
    assert data["costs"]["note"] == "traceability-test"
    assert data["costs"]["slippage"] == 0.01
    # Excel may be summary.xlsx or failed
    xlsx = paths["excel"]
    assert xlsx.exists()
    if xlsx.suffix == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(xlsx)
        ws = wb.active
        labels = {str(ws.cell(r, 1).value): ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
        assert "手续费率 commission_rate" in labels
        assert labels["手续费率 commission_rate"] == 0.0003
        assert labels["最低佣金 min_commission"] == 5.0
        assert labels["印花税率 stamp_tax_rate"] == 0.001
        assert labels["滑点 slippage"] == 0.01
        assert labels["成本说明 note"] == "traceability-test"
