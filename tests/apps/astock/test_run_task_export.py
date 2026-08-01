# -*- coding: utf-8 -*-
"""Backtest task-record Excel export."""

from __future__ import annotations

import io
import json
from pathlib import Path

import pytest

import tests.apps.astock.conftest  # noqa: F401

from wtpy.apps.astock.api import create_app
from wtpy.apps.astock.config import AStockConfig
from wtpy.apps.astock.service.run_task_export import (
    normalize_export_run_ids,
    write_run_task_excel,
)


def _write_run(
    root: Path,
    run_id: str,
    *,
    indicator_name: str,
    buy_weekday: int | None,
    exit_weekday: int | None,
    buy_on: str,
    sell_on: str,
    with_bagua: bool = False,
    total_return: float = 0.1261,
    annual_return: float = 0.0842,
    max_drawdown: float = -0.0612,
    win_rate: float = 0.6829,
    n_round_trips: int = 137,
) -> None:
    run_dir = root / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    gua_filter = (
        {
            "enabled": True,
            "history_summary": {
                "short": "䷗ 地雷复·初九、䷭ 地风升·初六",
                "tooltip_lines": ["䷗ 地雷复·初九", "䷭ 地风升·初六"],
            },
        }
        if with_bagua
        else {"enabled": False}
    )
    meta = {
        "run_id": run_id,
        "status": "ok",
        "title": indicator_name,
        "indicator_names": [indicator_name],
        "period": "DAY",
        "start": 20240101,
        "end": 20241231,
        "schedule_mode": "weekday" if buy_weekday is not None else "tn",
        "buy_weekday": buy_weekday,
        "exit_weekday": exit_weekday,
        "entry_lag": 1,
        "hold": 5,
        "buy_on": buy_on,
        "sell_on": sell_on,
        "signal_weekdays": [5],
        "account_mode": "portfolio",
        "with_bagua": with_bagua,
        "gua_filter": gua_filter,
        "bagua_filter_label": (
            "精确爻象命中已选的2条状态。" if with_bagua else None
        ),
        "metrics": {
            "total_return": total_return,
            "annual_return": annual_return,
            "max_drawdown": max_drawdown,
            "win_rate": win_rate,
            "n_round_trips": n_round_trips,
        },
    }
    (run_dir / "run_meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (run_dir / "metrics.json").write_text(
        json.dumps(meta["metrics"], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


@pytest.fixture()
def cfg(tmp_path: Path) -> AStockConfig:
    cfg = AStockConfig()
    cfg.output_root = tmp_path / "outputs"
    cfg.output_root.mkdir(parents=True, exist_ok=True)
    return cfg


def test_normalize_export_run_ids_rejects_traversal():
    assert normalize_export_run_ids(["bt_a", "bt_a", "bt_b"]) == ["bt_a", "bt_b"]
    with pytest.raises(ValueError, match="invalid run_id"):
        normalize_export_run_ids(["../run_meta.json"])


def test_write_run_task_excel_matches_reference_layout(cfg: AStockConfig, tmp_path: Path):
    _write_run(
        Path(cfg.output_root),
        "bt_735",
        indicator_name="735指标+",
        buy_weekday=1,
        exit_weekday=4,
        buy_on="open",
        sell_on="close",
    )
    _write_run(
        Path(cfg.output_root),
        "bt_5day",
        indicator_name="5日外",
        buy_weekday=None,
        exit_weekday=None,
        buy_on="open",
        sell_on="open",
        with_bagua=True,
    )

    path = write_run_task_excel(
        cfg,
        ["bt_735", "bt_5day"],
        tmp_path / "回测数据记录.xlsx",
    )

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False)
    assert wb.sheetnames == ["回测任务", "任务参数"]
    ws = wb["回测任务"]
    assert {"A1:A2", "B1:G1"}.issubset(
        {str(item) for item in ws.merged_cells.ranges}
    )
    assert [ws.cell(2, col).value for col in range(2, 9)] == [
        "策略条件",
        "买入日",
        "买入时点",
        "卖出日",
        "卖出时点",
        "卦象筛选",
        "回测摘要",
    ]
    assert [ws.cell(3, col).value for col in range(1, 7)] == [
        1,
        "735指标+",
        "周一",
        "开盘",
        "周四",
        "收盘",
    ]
    assert ws["G3"].value == "未启用"
    assert "绩效：总收益 +12.61%" in ws["H3"].value
    assert "最大回撤 6.12%（回撤较低）" in ws["H3"].value
    assert "样本：完成交易 137 笔，达到基础观察阈值" in ws["H3"].value
    # 复验优先级列已生成，且符合门槛的最高收益行使用鲜明样式标记。
    assert ws["I1"].value == "复验优先级"
    assert ws["I3"].value == "★ 本批次最优\n优先复验"
    assert ws["H3"].font.bold is True
    assert str(ws["H3"].font.color.rgb).endswith("FF0000")
    assert ws["I3"].font.bold is True
    assert str(ws["A3"].fill.fgColor.rgb).endswith("FFF200")
    assert str(ws["I3"].fill.fgColor.rgb).endswith("FF0000")
    assert str(ws["I3"].font.color.rgb).endswith("FFFFFF")
    assert 76 <= ws.row_dimensions[3].height <= 104
    assert 76 <= ws.row_dimensions[4].height <= 104
    assert ws["C4"].value == "T+1"
    assert ws["E4"].value == "持有5日"
    assert ws["G4"].value == "䷗ 地雷复·初九、䷭ 地风升·初六"
    assert wb["任务参数"]["B2"].value == "bt_735"
    assert wb["任务参数"]["K2"].number_format == "0.00%"


def test_best_highlight_skips_low_sample_and_high_drawdown(
    cfg: AStockConfig, tmp_path: Path
):
    _write_run(
        Path(cfg.output_root),
        "bt_low_sample",
        indicator_name="低样本高收益",
        buy_weekday=1,
        exit_weekday=3,
        buy_on="open",
        sell_on="close",
        total_return=0.90,
        annual_return=0.35,
        max_drawdown=-0.08,
        n_round_trips=15,
    )
    _write_run(
        Path(cfg.output_root),
        "bt_high_drawdown",
        indicator_name="高回撤高收益",
        buy_weekday=1,
        exit_weekday=3,
        buy_on="open",
        sell_on="close",
        total_return=0.80,
        annual_return=0.32,
        max_drawdown=-0.31,
        n_round_trips=200,
    )
    _write_run(
        Path(cfg.output_root),
        "bt_eligible",
        indicator_name="符合门槛",
        buy_weekday=1,
        exit_weekday=3,
        buy_on="open",
        sell_on="close",
        total_return=0.30,
        annual_return=0.16,
        max_drawdown=-0.10,
        n_round_trips=120,
    )

    path = write_run_task_excel(
        cfg,
        ["bt_low_sample", "bt_high_drawdown", "bt_eligible"],
        tmp_path / "best_gate.xlsx",
    )

    from openpyxl import load_workbook

    ws = load_workbook(path, data_only=True)["回测任务"]
    assert "样本过少" in ws["H3"].value
    assert ws["I3"].value == "暂不评级（样本过少）"
    assert ws["I4"].value == "暂缓复验（回撤较高）"
    assert ws["I5"].value == "★ 本批次最优\n优先复验"
    assert str(ws["A5"].fill.fgColor.rgb).endswith("FFF200")
    assert str(ws["I5"].fill.fgColor.rgb).endswith("FF0000")


def test_run_task_export_api_download(cfg: AStockConfig):
    pytest.importorskip("fastapi")
    from fastapi.testclient import TestClient
    from openpyxl import load_workbook

    _write_run(
        Path(cfg.output_root),
        "bt_api",
        indicator_name="735指标+",
        buy_weekday=1,
        exit_weekday=2,
        buy_on="open",
        sell_on="open",
    )

    client = TestClient(create_app(cfg))
    response = client.get(
        "/api/v1/backtests/bt_api/artifacts/%E5%9B%9E%E6%B5%8B%E6%95%B0%E6%8D%AE%E8%AE%B0%E5%BD%95.xlsx"
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(io.BytesIO(response.content), data_only=True)
    assert wb["回测任务"]["B3"].value == "735指标+"
    assert wb["回测任务"]["E3"].value == "周二"
