# -*- coding: utf-8 -*-
"""Excel trade sheet columns must match TRADE_TRIP_FIELDS order (no shift)."""
from pathlib import Path
import tempfile

from openpyxl import load_workbook

from wtpy.apps.astock.report_price_schema import TRADE_TRIP_FIELDS
from wtpy.apps.astock.reports import pair_round_trips, write_excel_summary
from wtpy.apps.astock.strategy_models import BacktestResult, Fill


def test_excel_trade_sheet_columns_aligned_with_schema():
    fills = [
        Fill(
            date=20240102,
            std_code="SSE.STK.600000",
            side="BUY",
            price=10.0,
            shares=100,
            amount=1000.0,
            commission=1.0,
            stamp_tax=0.0,
            reason="entry",
            raw_price=10.0,
            adjusted_reference_price=9.5,
            standard_qfq_reference_price=9.8,
            adjustment_factor=0.98,
            adjustment_scale=0.98,
        ),
        Fill(
            date=20240110,
            std_code="SSE.STK.600000",
            side="SELL",
            price=11.0,
            shares=100,
            amount=1100.0,
            commission=1.0,
            stamp_tax=1.0,
            reason="time_exit",
            raw_price=11.0,
            adjusted_reference_price=10.5,
            standard_qfq_reference_price=10.8,
            adjustment_factor=0.99,
            adjustment_scale=0.99,
        ),
    ]
    trips = pair_round_trips(fills)
    result = BacktestResult(
        run_id="excel_col_align",
        config={},
        fills=fills,
        equity_curve=[],
        metrics={"total_return": 0.1, "max_drawdown": 0.05, "n_trades": 1},
        notes=[],
    )
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "summary.xlsx"
        write_excel_summary(
            path,
            result,
            trips=trips,
            meta={
                "repro": {
                    "start": 20240101,
                    "end": 20240131,
                    "period": "DAY",
                    "hold": 5,
                    "entry_lag": 1,
                }
            },
        )
        ws = load_workbook(path)["交易明细"]
        headers = [c.value for c in ws[2]]
        row = [c.value for c in ws[3]]

    assert len(headers) == len(TRADE_TRIP_FIELDS) == len(row)
    # display alias for pct columns only
    for i, field in enumerate(TRADE_TRIP_FIELDS):
        if field in ("毛收益率", "净收益率"):
            assert headers[i] == field + "%"
        else:
            assert headers[i] == field, (i, field, headers[i])

    m = dict(zip(headers, row))
    assert str(m["买入价"]).startswith("10")
    assert str(m["买入价_起点锚定研究参考"]).startswith("9.5")
    assert str(m["买入价_普通前复权参考"]).startswith("9.8")
    assert str(m["买入复权因子"]).startswith("0.98")
    assert str(m["买入复权比例"]).startswith("0.98")
    assert "2024" in str(m["卖出日期"])
    assert str(m["卖出价"]).startswith("11")
    assert str(m["卖出价_普通前复权参考"]).startswith("10.8")
    assert m["数量"] == 100
    assert m["状态"] == "已平仓"
    # columns after 买入复权比例 must not be shifted left into factor/scale
    assert m["卖出复权比例"] is not None
    assert str(m["卖出复权比例"]).startswith("0.99")
