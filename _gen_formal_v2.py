# -*- coding: utf-8 -*-
import sys, json
from pathlib import Path
from types import ModuleType

ROOT = Path(r"E:\Software Development\wtpy-master")
sys.path.insert(0, str(ROOT))
for k in list(sys.modules):
    if k.startswith("wtpy.apps.astock"):
        del sys.modules[k]

def ensure(n, p):
    if n in sys.modules:
        return
    m = ModuleType(n)
    m.__path__ = [str(p)]
    m.__package__ = n
    sys.modules[n] = m

for n, p in [
    ("wtpy", ROOT / "wtpy"),
    ("wtpy.apps", ROOT / "wtpy" / "apps"),
    ("wtpy.apps.astock", ROOT / "wtpy" / "apps" / "astock"),
    ("wtpy.apps.astock.data", ROOT / "wtpy" / "apps" / "astock" / "data"),
    ("wtpy.apps.astock.indicators", ROOT / "wtpy" / "apps" / "astock" / "indicators"),
    ("wtpy.apps.astock.bagua", ROOT / "wtpy" / "apps" / "astock" / "bagua"),
]:
    ensure(n, p)

from wtpy.apps.astock.cli import main, _astock_code_sha
from wtpy.apps.astock.config import get_default_config
from wtpy.apps.astock.data.catalog import file_sha_or_empty, selected_universe_sha
from wtpy.apps.astock.indicators.registry import IndicatorRegistry
from wtpy.apps.astock.indicators.tn6_importer import (
    load_source_map,
    resolve_formula_audit,
    file_sha256,
)

# Freeze live sha AFTER all code edits
live = _astock_code_sha()
print("LIVE_CODE_SHA", live)

cfg = get_default_config()
print("csv", len(list((cfg.storage_root / "csv" / "day").rglob("*.csv"))))
print("npz", len(list((cfg.storage_root / "npz" / "day").rglob("*.npz"))))
print("manifest", json.loads(cfg.manifest_path.read_text(encoding="utf-8")).get("count"))
print("universe", json.loads(cfg.universe_path.read_text(encoding="utf-8")).get("count"))
print("GLOBAL_M", file_sha_or_empty(cfg.manifest_path))
print("GLOBAL_U", file_sha_or_empty(cfg.universe_path))

# verify 735 confirmation live
ind = Path(cfg.indicator_dir)
tn6 = list(ind.glob("*735*.tn6"))[0]
src = list(ind.glob("*735*.txt"))[0]
pkg = file_sha256(tn6)
ss = file_sha256(src)
print("live_pkg", pkg)
print("live_src", ss)
entry = load_source_map(cfg.mapping_path)[pkg]
# ensure package_file set
if not entry.get("package_file"):
    entry["package_file"] = str(tn6.resolve())
    from wtpy.apps.astock.indicators.tn6_importer import save_source_map
    m = load_source_map(cfg.mapping_path)
    m[pkg] = entry
    save_source_map(cfg.mapping_path, m)
    entry = m[pkg]
audit = resolve_formula_audit(entry, package_sha256=pkg)
print("audit", json.dumps(audit, ensure_ascii=False))

reg = IndicatorRegistry.bootstrap(cfg.indicator_dir, cfg.mapping_path)
s735 = [s for s in reg.list() if s.package_sha256 == pkg][0]
print("id", s735.id)

codes = "sh600000,sz000001,sh600519,sz000002,sh601318,sz300750,sh600036,sz000858,sh601166,sz002415"

# Formal confirmed
rc = main([
    "backtest", "--indicator", s735.id, "--period", "DAY", "--hold", "5",
    "--codes", codes, "--start", "20200101", "--end", "20241231",
    "--stop-loss", "0.03", "--take-profit", "0.08",
    "--run-id", "pool10_735_formal_v2",
])
print("formal_v2_rc", rc)

# Research adjusted unconfirmed: use temp mapping by temporarily not using project confirm -
# For production research_adjusted_v4 we need unconfirmed fixture.
# Create isolated storage clone is heavy; instead use CLI with a copy of mapping?
# Spec: pool10_735_research_adjusted_v4 with temporary unconfirmed fixture.
# Use --storage temp with selection data from project? Easier: call main with project storage
# but pair-unconfirmed would overwrite confirm - forbidden.
# Implement by running with a temp map via storage that has bars symlink-like copy of 10 csv?
# Minimal: import-data codes into temp storage from tdx, pair without confirm, write synthetic factors.

import tempfile, shutil
tmp = Path(tempfile.mkdtemp(prefix="astock_research_adj_"))
# use project tdx
rc_imp = main([
    "--storage", str(tmp / "storage"),
    "import-data", "--codes", codes, "--skip-dsb", "--skip-factors",
])
print("tmp import", rc_imp)
# copy adjustment json for the 10 codes from project if exist
src_adj = cfg.adj_root
dst_adj = tmp / "storage" / "adjustments"
dst_adj.mkdir(parents=True, exist_ok=True)
for p in src_adj.glob("*.json"):
    shutil.copy2(p, dst_adj / p.name)
# pair 735 into temp map
rc_pair = main([
    "--storage", str(tmp / "storage"),
    "pair-735",
])
print("tmp pair", rc_pair)
# research unconfirmed adjusted
rc_ra = main([
    "--storage", str(tmp / "storage"),
    "backtest",
    "--indicator", s735.id if False else "tn6_735金叉及趋势",
    "--period", "DAY", "--hold", "5",
    "--codes", codes, "--start", "20200101", "--end", "20241231",
    "--research-unconfirmed-formula",
    "--stop-loss", "0.03", "--take-profit", "0.08",
    "--run-id", "pool10_735_research_adjusted_v4",
])
print("research_adj_rc", rc_ra)
# copy output into project outputs if generated under project output_root
# with --storage only, output_root still project outputs/astock by config design?
from wtpy.apps.astock.config import get_default_config as g
c_tmp = g(storage_root=tmp / "storage")
print("tmp output root", c_tmp.output_root)
# if run landed in project outputs due to default output_root not overridden:
for name in ["pool10_735_formal_v2", "pool10_735_research_adjusted_v4"]:
    root = cfg.output_root / name
    if not root.exists():
        root = c_tmp.output_root / name
    print("exists", name, root.exists(), root)
    if root.exists():
        meta = json.loads((root / "run_meta.json").read_text(encoding="utf-8"))
        m = json.loads((root / "metrics.json").read_text(encoding="utf-8"))
        repro = meta.get("repro") or {}
        print("====", name)
        print(" status", meta.get("status"), "price_mode", repro.get("price_mode"))
        print(" prov", repro.get("formula_provenance"), repro.get("source_pair_status"), "formal_allowed", repro.get("formal_backtest_allowed"))
        print(" confirmed_by", repro.get("confirmed_by"))
        print(" ret", m.get("total_return"), "buys", m.get("n_buys"), "sells", m.get("n_sells"), "open", m.get("n_open_positions"))
        print(" reasons", m.get("sell_reason_counts"))
        print(" code_match", repro.get("astock_code_sha") == live, (repro.get("astock_code_sha") or "")[:16])
        print(" g_m", (repro.get("global_manifest_sha") or "")[:16], "sel", (repro.get("selected_universe_sha") or "")[:16], "gcount", repro.get("global_universe_count"), "scount", repro.get("selected_codes_count"))
        print(" xlsx", (root / "summary.xlsx").exists(), "failed", (root / "summary.xlsx.failed").exists())
        # compare excel data_quality vs run_meta for key fields
        try:
            from openpyxl import load_workbook
            wb = load_workbook(root / "summary.xlsx")
            ws = wb["data_quality"]
            dq = {r[0].value: r[1].value for r in ws.iter_rows(min_row=2, max_col=2) if r[0].value}
            keys = ["astock_code_sha", "price_mode", "formula_provenance", "source_pair_status", "global_manifest_sha", "selected_universe_sha"]
            for k in keys:
                if k in repro:
                    print(" dq", k, dq.get(k) == repro.get(k) or str(dq.get(k)) == str(repro.get(k)), dq.get(k), repro.get(k))
        except Exception as e:
            print("excel check err", e)

print("LIVE", live)
