"""CSV / Excel report writers with reproducibility metadata.

Excel goal: **one** human-readable workbook (交易回测明细.xlsx preferred as summary.xlsx)
with 汇总 + 交易明细 (signal → buy → sell → P&L), inspired by
outputs/astock/735_八卦过滤交易明细_用户白名单.xlsx but without bagua-only focus.
"""

from __future__ import annotations

from dataclasses import asdict, is_dataclass

import csv
import json
from collections import defaultdict, deque
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from .data.io_util import atomic_write_json
from .strategy import BacktestResult, Fill
from .study import ForwardStats, SignalEvent


def ensure_dir(path: Path) -> Path:
    path = Path(path)
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_signals_csv(path: Path, events: Sequence[SignalEvent]) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "std_code",
        "date",
        "period",
        "indicator_id",
        "value",
        "is_dwm",
        "bagua_full_name",
        "bagua_yao_name",
        "bagua_gua_order",
        "bagua_judgement",
        "bagua_core_gang",
        "bagua_state_id",
        "bagua_action_signal",
        "bagua_biangua",
        "bagua_rule_version",
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for e in events:
            bg = e.bagua or {}
            w.writerow(
                {
                    "std_code": e.std_code,
                    "date": e.date,
                    "period": e.period,
                    "indicator_id": e.indicator_id,
                    "value": e.value,
                    "is_dwm": int(bool(e.is_dwm)),
                    "bagua_full_name": bg.get("full_name", ""),
                    "bagua_yao_name": bg.get("yao_name", ""),
                    "bagua_gua_order": bg.get("gua_order", ""),
                    "bagua_judgement": bg.get("market_judgement", ""),
                    "bagua_core_gang": bg.get("core_gang", ""),
                    "bagua_state_id": bg.get("state_id", ""),
                    "bagua_action_signal": bg.get("action_signal", ""),
                    "bagua_biangua": bg.get("biangua") or bg.get("changed_hexagram_name") or "",
                    "bagua_rule_version": "",
                }
            )
    return path


def write_stats_csv(path: Path, stats: Sequence[ForwardStats]) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "key",
        "n",
        "win_rate",
        "mean_return",
        "median_return",
        "mfe",
        "mae",
        "period",
        "indicator_id",
        "gua",
        "yao",
        "is_dwm",
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for s in stats:
            w.writerow(s.to_dict())
    return path


def _code_short(std_code: str) -> str:
    if not std_code:
        return ""
    parts = str(std_code).split(".")
    return parts[-1] if parts else str(std_code)


def _fmt_date(d: Any) -> str:
    if d is None or d == "":
        return ""
    try:
        n = int(d)
        return f"{n // 10000:04d}-{(n // 100) % 100:02d}-{n % 100:02d}"
    except Exception:
        return str(d)


def _fmt_pct(x: Any) -> Any:
    if x is None:
        return ""
    try:
        return round(float(x) * 100.0, 4)
    except Exception:
        return x


def _fmt_num(x: Any, nd: int = 4) -> Any:
    if x is None or x == "":
        return ""
    try:
        return round(float(x), nd)
    except Exception:
        return x


def pair_round_trips(
    fills: Sequence[Fill],
    events: Optional[Sequence[SignalEvent]] = None,
) -> List[Dict[str, Any]]:
    """Match BUY->SELL per stock (FIFO) and attach nearest prior signal date."""
    # (date, indicator_id, full_name, yao, judge, state_id, action, biangua, gua_order)
    sig_by_code: Dict[str, List[Tuple]] = defaultdict(list)
    if events:
        for e in events:
            try:
                bg = e.bagua or {}
                if not isinstance(bg, dict):
                    bg = {}
                sid = bg.get("state_id") or ""
                if not sid and bg.get("gua_order") is not None and bg.get("yao_order") is not None:
                    sid = f"{int(bg['gua_order']):02d}-{int(bg['yao_order'])}"
                sig_by_code[e.std_code].append(
                    (
                        int(e.date),
                        str(e.indicator_id or ""),
                        str(bg.get("full_name") or bg.get("gua_name") or ""),
                        str(bg.get("yao_name") or bg.get("line_name") or ""),
                        str(bg.get("market_judgement") or bg.get("market_summary") or "").replace("\n", " | "),
                        str(sid or ""),
                        str(bg.get("action_signal") or ""),
                        str(bg.get("biangua") or bg.get("changed_hexagram_name") or ""),
                        bg.get("gua_order") if bg.get("gua_order") is not None else "",
                    )
                )
            except Exception:
                continue
        for k in sig_by_code:
            sig_by_code[k].sort(key=lambda x: x[0])

    def find_signal(code: str, buy_date: int):
        arr = sig_by_code.get(code) or []
        best = None
        for row in arr:
            d = row[0]
            if d <= buy_date:
                best = row
            else:
                break
        if best:
            return best
        return (None, "", "", "", "", "", "", "", "")

    # queue of open buy lots: (fill, remaining_shares)
    open_buys: Dict[str, deque] = defaultdict(deque)
    trips: List[Dict[str, Any]] = []
    seq = 0

    for f in fills:
        side = (f.side or "").upper()
        code = f.std_code
        if side == "BUY":
            open_buys[code].append((f, int(f.shares or 0)))
            continue
        if side != "SELL":
            continue

        remain = int(f.shares or 0)
        sell_price = float(f.price or 0.0)
        sell_date = int(f.date)
        sell_comm = float(f.commission or 0.0)
        sell_tax = float(f.stamp_tax or 0.0)
        sell_shares_total = max(int(f.shares or 0), 1)
        reason = f.reason or ""
        parts: List[Tuple[Fill, int]] = []

        while remain > 0 and open_buys[code]:
            b, left = open_buys[code][0]
            if left <= 0:
                open_buys[code].popleft()
                continue
            take = min(left, remain)
            parts.append((b, take))
            left -= take
            remain -= take
            if left <= 0:
                open_buys[code].popleft()
            else:
                open_buys[code][0] = (b, left)

        if not parts:
            seq += 1
            trips.append(
                {
                    "序号": seq,
                    "证券代码": code,
                    "代码": _code_short(code),
                    "信号日期": "",
                    "指标": "",
                    "买入日期": "",
                    "买入价": "",
                    "卖出日期": sell_date,
                    "卖出价": sell_price,
                    "数量": int(f.shares or 0),
                    "买入金额": "",
                    "卖出金额": float(f.amount or 0.0),
                    "买入手续费": "",
                    "卖出手续费及印花税": sell_comm + sell_tax,
                    "毛利润": "",
                    "净利润": "",
                    "毛收益率": "",
                    "净收益率": "",
                    "是否盈利": "",
                    "卖出原因": reason,
                    "状态": "仅有卖出/未匹配买入",
                }
            )
            continue

        for b, take in parts:
            seq += 1
            buy_date = int(b.date)
            buy_price = float(b.price or 0.0)
            b_sh = max(int(b.shares or 0), 1)
            ratio = take / float(b_sh)
            buy_amount = float(b.amount or 0.0) * ratio
            if not buy_amount:
                buy_amount = buy_price * take
            buy_fee = float(b.commission or 0.0) * ratio
            sell_amount = sell_price * take
            sell_cost = (sell_comm + sell_tax) * (take / float(sell_shares_total))
            gross = sell_amount - buy_amount
            net = gross - buy_fee - sell_cost
            gross_ret = (gross / buy_amount) if buy_amount else None
            net_ret = (net / buy_amount) if buy_amount else None
            sig_row = find_signal(code, buy_date)
            sig_d, ind, gua, yao, judge, st_id, act_sig, biangua, gua_ord = (list(sig_row) + [""] * 9)[:9]
            trips.append(
                {
                    "序号": seq,
                    "证券代码": code,
                    "代码": _code_short(code),
                    "信号日期": sig_d if sig_d is not None else "",
                    "指标": ind,
                    "卦名": gua,
                    "爻位": yao,
                    "卦序": gua_ord,
                    "变卦": biangua or "未配置",
                    "操作信号": act_sig,
                    "state_id": st_id,
                    "卦象简判": judge,
                    "买入日期": buy_date,
                    "买入价": buy_price,
                    "卖出日期": sell_date,
                    "卖出价": sell_price,
                    "数量": take,
                    "买入金额": buy_amount,
                    "卖出金额": sell_amount,
                    "买入手续费": buy_fee,
                    "卖出手续费及印花税": sell_cost,
                    "毛利润": gross,
                    "净利润": net,
                    "毛收益率": gross_ret,
                    "净收益率": net_ret,
                    "是否盈利": (
                        "盈利" if net > 0 else ("亏损" if net < 0 else "持平")
                    ),
                    "卖出原因": reason,
                    "状态": "已平仓",
                }
            )

    for code, q in open_buys.items():
        while q:
            b, left = q.popleft()
            if left <= 0:
                continue
            seq += 1
            buy_date = int(b.date)
            buy_price = float(b.price or 0.0)
            b_sh = max(int(b.shares or 0), 1)
            ratio = left / float(b_sh)
            buy_amount = float(b.amount or 0.0) * ratio or (buy_price * left)
            buy_fee = float(b.commission or 0.0) * ratio
            sig_row = find_signal(code, buy_date)
            sig_d, ind, gua, yao, judge, st_id, act_sig, biangua, gua_ord = (list(sig_row) + [""] * 9)[:9]
            trips.append(
                {
                    "序号": seq,
                    "证券代码": code,
                    "代码": _code_short(code),
                    "信号日期": sig_d if sig_d is not None else "",
                    "指标": ind,
                    "卦名": gua,
                    "爻位": yao,
                    "卦序": gua_ord,
                    "变卦": biangua or "未配置",
                    "操作信号": act_sig,
                    "state_id": st_id,
                    "卦象简判": judge,
                    "买入日期": buy_date,
                    "买入价": buy_price,
                    "卖出日期": "",
                    "卖出价": "",
                    "数量": left,
                    "买入金额": buy_amount,
                    "卖出金额": "",
                    "买入手续费": buy_fee,
                    "卖出手续费及印花税": "",
                    "毛利润": "",
                    "净利润": "",
                    "毛收益率": "",
                    "净收益率": "",
                    "是否盈利": "",
                    "卖出原因": "",
                    "状态": "未平仓",
                }
            )

    trips.sort(
        key=lambda r: (
            int(r["买入日期"] or 0),
            str(r["证券代码"]),
            int(r["卖出日期"] or 0),
        )
    )
    for i, r in enumerate(trips, 1):
        r["序号"] = i
    return trips


def write_backtest_csv(
    out_dir: Path,
    result: BacktestResult,
    *,
    meta: Optional[dict] = None,
    study_stats: Optional[Sequence[ForwardStats]] = None,
    events: Optional[Sequence[SignalEvent]] = None,
) -> Dict[str, Path]:
    out_dir = ensure_dir(out_dir)
    fills_path = out_dir / "fills.csv"
    eq_path = out_dir / "equity.csv"
    metrics_path = out_dir / "metrics.json"
    meta_path = out_dir / "run_meta.json"
    trades_path = out_dir / "trades.csv"

    with open(fills_path, "w", newline="", encoding="utf-8-sig") as f:
        fields = [
            "date",
            "std_code",
            "side",
            "price",
            "shares",
            "amount",
            "commission",
            "stamp_tax",
            "reason",
        ]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for x in result.fills:
            w.writerow(
                {
                    "date": x.date,
                    "std_code": x.std_code,
                    "side": x.side,
                    "price": x.price,
                    "shares": x.shares,
                    "amount": x.amount,
                    "commission": x.commission,
                    "stamp_tax": x.stamp_tax,
                    "reason": x.reason,
                }
            )

    with open(eq_path, "w", newline="", encoding="utf-8-sig") as f:
        fields = ["date", "cash", "market_value", "equity"]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for e in result.equity_curve:
            w.writerow(
                {
                    "date": e.date,
                    "cash": e.cash,
                    "market_value": e.market_value,
                    "equity": e.equity,
                }
            )

    trips = pair_round_trips(result.fills, events=events)
    trade_fields = [
        "序号",
        "证券代码",
        "代码",
        "信号日期",
        "指标",
        "卦名",
        "爻位",
        "卦序",
        "变卦",
        "操作信号",
        "state_id",
        "卦象简判",
        "买入日期",
        "买入价",
        "卖出日期",
        "卖出价",
        "数量",
        "买入金额",
        "卖出金额",
        "买入手续费",
        "卖出手续费及印花税",
        "毛利润",
        "净利润",
        "毛收益率",
        "净收益率",
        "是否盈利",
        "卖出原因",
        "状态",
    ]
    with open(trades_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=trade_fields, extrasaction="ignore")
        w.writeheader()
        for row in trips:
            w.writerow(row)

    atomic_write_json(metrics_path, result.metrics)

    # Do not embed full fills/equity in run_meta.json (can be tens of MB and block UI at 96%).
    full_meta = {
        "run_id": result.run_id,
        "status": result.status,
        "metrics": result.metrics,
        "notes": result.notes,
        "n_fills": len(result.fills),
        "n_equity_points": len(result.equity_curve),
        "n_trade_rows": len(trips),
        "fills_path": "fills.csv",
        "equity_path": "equity.csv",
        "trades_path": "trades.csv",
        "signals_path": "signals.csv",
    }
    if meta:
        full_meta["repro"] = meta
        if isinstance(meta.get("config"), dict):
            full_meta["config"] = meta["config"]
        for k in (
            "title",
            "indicator_names",
            "indicator_ids",
            "hold",
            "entry_lag",
            "period",
            "with_bagua",
            "gua_filter",
            "n_signals_before_bagua",
            "n_signals_after_bagua",
            "buy_weekday",
            "exit_weekday",
            "buy_on",
            "sell_on",
            "signal_weekdays",
            "schedule_mode",
            "account_mode",
            "bagua_filter_mode",
            "bagua_filter_label",
            "start",
            "end",
            "costs",
        ):
            if k in meta and k not in full_meta:
                full_meta[k] = meta[k]
    # Cost traceability (P1.7): always surface CostConfig fields on run_meta
    costs_src = None
    if isinstance(full_meta.get("costs"), dict):
        costs_src = full_meta["costs"]
    elif isinstance(full_meta.get("config"), dict) and isinstance(full_meta["config"].get("costs"), dict):
        costs_src = full_meta["config"]["costs"]
    elif isinstance(getattr(result, "config", None), dict) and isinstance(result.config.get("costs"), dict):
        costs_src = result.config["costs"]
        full_meta.setdefault("config", dict(result.config))
    if costs_src is not None:
        full_meta["costs"] = {
            "commission_rate": costs_src.get("commission_rate"),
            "min_commission": costs_src.get("min_commission"),
            "stamp_tax_rate": costs_src.get("stamp_tax_rate"),
            "slippage": costs_src.get("slippage"),
            "note": costs_src.get("note", ""),
        }
        cfg = full_meta.get("config")
        if isinstance(cfg, dict) and "costs" not in cfg:
            cfg = dict(cfg)
            cfg["costs"] = dict(full_meta["costs"])
            full_meta["config"] = cfg
    # keep a tiny sample for debugging only
    full_meta["fills_sample"] = [
        asdict(f) if is_dataclass(f) else str(f) for f in result.fills[:20]
    ]
    atomic_write_json(meta_path, full_meta)

    if study_stats:
        write_stats_csv(out_dir / "study_stats.csv", study_stats)

    xlsx_path = out_dir / "summary.xlsx"
    failed_path = out_dir / "summary.xlsx.failed"
    try:
        write_excel_summary(
            xlsx_path,
            result,
            meta=meta,
            study_stats=study_stats,
            events=events,
            trips=trips,
        )
        if failed_path.exists():
            failed_path.unlink()
    except Exception as e:
        if xlsx_path.exists():
            try:
                xlsx_path.unlink()
            except Exception:
                pass
        failed_path.write_text(str(e), encoding="utf-8")
        xlsx_path = failed_path

    return {
        "fills": fills_path,
        "equity": eq_path,
        "metrics": metrics_path,
        "meta": meta_path,
        "trades": trades_path,
        "excel": xlsx_path,
    }


def write_excel_summary(
    path: Path,
    result: BacktestResult,
    *,
    meta: Optional[dict] = None,
    study_stats: Optional[Sequence[ForwardStats]] = None,
    events: Optional[Sequence[SignalEvent]] = None,
    trips: Optional[List[Dict[str, Any]]] = None,
) -> Path:
    """Single workbook: 汇总 + 交易明细 (+ optional thin 说明)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    if trips is None:
        trips = pair_round_trips(result.fills, events=events)

    repro = meta or {}
    m = result.metrics or {}
    _costs_cfg = None
    if isinstance(repro.get("costs"), dict):
        _costs_cfg = repro["costs"]
    elif isinstance(repro.get("config"), dict) and isinstance(repro["config"].get("costs"), dict):
        _costs_cfg = repro["config"]["costs"]
    elif isinstance(getattr(result, "config", None), dict) and isinstance(result.config.get("costs"), dict):
        _costs_cfg = result.config["costs"]
    _costs_cfg = _costs_cfg or {}

    closed = [t for t in trips if t.get("状态") == "已平仓"]
    open_n = sum(1 for t in trips if t.get("状态") == "未平仓")
    win_n = sum(1 for t in closed if t.get("是否盈利") == "盈利")
    loss_n = sum(1 for t in closed if t.get("是否盈利") == "亏损")
    flat_n = sum(1 for t in closed if t.get("是否盈利") == "持平")
    net_sum = sum(float(t["净利润"]) for t in closed if t.get("净利润") != "")
    gross_sum = sum(float(t["毛利润"]) for t in closed if t.get("毛利润") != "")
    trip_wr = (win_n / len(closed)) if closed else None

    wb = Workbook()

    # ---- 汇总 ----
    ws = wb.active
    ws.title = "汇总"
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    section_fill = PatternFill("solid", fgColor="D6EAF8")
    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    ws.append(["回测结果汇总（单文件）"])
    ws["A1"].font = title_font
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws.merge_cells("A1:B1")

    rows_sum = [
        ("run_id", result.run_id),
        ("状态", result.status),
        ("指标", ",".join(repro.get("indicator_ids") or [])),
        ("八卦过滤", repro.get("bagua_filter_label") or ("否" if not repro.get("with_bagua") else "是")),
        ("八卦白名单", ",".join(repro.get("bagua_allowlist") or []) if repro.get("bagua_allowlist") else ""),
        ("卦象规则版本", (repro.get("gua_filter") or {}).get("rule_version") if isinstance(repro.get("gua_filter"), dict) else ""),
        ("卦象选择模式", (repro.get("gua_filter") or {}).get("selection_mode") if isinstance(repro.get("gua_filter"), dict) else ""),
        ("过滤前信号数", repro.get("n_signals_before_bagua") if repro.get("n_signals_before_bagua") is not None else ""),
        ("过滤后信号数", repro.get("n_signals_after_bagua") if repro.get("n_signals_after_bagua") is not None else ""),
        ("周期", repro.get("period") or ""),
        ("持有天数/期数 hold", (
            "（由平仓星期覆盖；hold 仅复现）"
            if repro.get("exit_weekday") is not None
            else (repro.get("hold") if repro.get("hold") is not None else "")
        )),
        ("账户模式", (
            "通达信对照·单票独立资金" if str(repro.get("account_mode") or "").lower() in
            ("per_symbol", "tdx", "per_stock") else "组合账户·共享资金"
        )),
        ("日程模式 schedule_mode", repro.get("schedule_mode") or ("weekday" if repro.get("buy_weekday") is not None or repro.get("exit_weekday") is not None else "tn")),
        ("信号星期", (
            "、".join(str(x) for x in (repro.get("signal_weekdays") or []))
            if repro.get("signal_weekdays")
            else "不限制"
        )),
        ("买入日（星期）", repro.get("buy_weekday") if repro.get("buy_weekday") is not None else "—（用 entry_lag）"),
        ("平仓日（星期）", repro.get("exit_weekday") if repro.get("exit_weekday") is not None else "—（用 hold）"),
        ("买入/卖出时点", f"{repro.get('buy_on') or 'open'} / {repro.get('sell_on') or 'open'}"),
        ("入场滞后 entry_lag", (
            f"{repro.get('entry_lag') or ''}（星期买入覆盖步进时仅作复现字段）"
            if repro.get("buy_weekday") is not None
            else (repro.get("entry_lag") or "")
        )),
        ("回测区间", f"{repro.get('start') or ''} ~ {repro.get('end') or ''}"),
        ("股票池数量", repro.get("selected_codes_count") or ""),
        ("价格模式", repro.get("price_mode") or ""),
        ("止损", repro.get("stop_loss_pct") if repro.get("stop_loss_pct") is not None else ""),
        ("止盈", repro.get("take_profit_pct") if repro.get("take_profit_pct") is not None else ""),
        ("", ""),
        ("【组合层指标】", ""),
        ("总收益率(含成本)", _fmt_pct(m.get("total_return"))),
        ("等权平均单票收益(通达信口径)", _fmt_pct(m.get("mean_symbol_return")) if m.get("mean_symbol_return") is not None else ""),
        ("单票账户数", m.get("n_symbol_accounts") if m.get("n_symbol_accounts") is not None else ""),
        ("盈利股票占比", _fmt_pct(m.get("pct_symbols_profitable")) if m.get("pct_symbols_profitable") is not None else ""),
        ("年化收益率%", _fmt_pct(m.get("annual_return"))),
        ("最大回撤%", _fmt_pct(m.get("max_drawdown"))),
        ("波动率%", _fmt_pct(m.get("volatility"))),
        ("夏普", _fmt_num(m.get("sharpe"), 4)),
        ("期末权益", _fmt_num(m.get("final_equity"), 2)),
        ("交易日数", m.get("n_days")),
        ("买入次数", m.get("n_buys")),
        ("卖出次数", m.get("n_sells")),
        ("完整回合数", m.get("n_round_trips")),
        ("组合胜率%", _fmt_pct(m.get("win_rate"))),
        ("换手", _fmt_num(m.get("turnover"), 4)),
        ("总成本", _fmt_num(m.get("cost_total"), 2)),
        ("", ""),
        ("【成本口径 CostConfig】", ""),
        ("手续费率 commission_rate", _costs_cfg.get("commission_rate", "")),
        ("最低佣金 min_commission", _costs_cfg.get("min_commission", "")),
        ("印花税率 stamp_tax_rate", _costs_cfg.get("stamp_tax_rate", "")),
        ("滑点 slippage", _costs_cfg.get("slippage", "")),
        ("成本说明 note", _costs_cfg.get("note", "")),
        ("未平仓数量", m.get("n_open_positions")),
        ("未平仓市值", _fmt_num(m.get("open_market_value"), 2)),
        ("不计成本收益%", _fmt_pct(m.get("zero_cost_return"))),
        ("成本拖累%", _fmt_pct(m.get("cost_impact"))),
        ("", ""),
        ("【逐笔回合统计】", ""),
        ("明细笔数(含未平)", len(trips)),
        ("已平仓笔数", len(closed)),
        ("盈利笔数", win_n),
        ("亏损笔数", loss_n),
        ("持平笔数", flat_n),
        ("未平仓笔数", open_n),
        ("回合胜率%", _fmt_pct(trip_wr)),
        ("合计毛利润", _fmt_num(gross_sum, 2)),
        ("合计净利润", _fmt_num(net_sum, 2)),
        ("平均净收益率%", _fmt_pct((net_sum / sum(float(t['买入金额']) for t in closed if t.get('买入金额'))) if closed and sum(float(t.get('买入金额') or 0) for t in closed) else None)),
        ("", ""),
        ("【说明】", ""),
        (
            "字段含义",
            "信号日期=指标触发日（收盘确认）。"
            "买卖日程在交易日序列上计算：有买入/平仓星期时，取信号/买入之后第一个该 ISO 星期的可交易日（节假日顺延）；"
            "无星期配置时用 entry_lag（信号后第 N 个交易日入场）与 hold（持有 N 期后强平）。"
            "开盘/收盘由 buy_on/sell_on 决定。净利润=卖出金额-买入金额-买卖费用；"
            "组合收益受仓位/拒单影响，与逐笔简单加总可能不一致。",
        ),
        ("免责", "研究用途，非投资建议；费率/涨跌停/停牌规则为系统近似。"),
    ]
    for k, v in rows_sum:
        ws.append([k, v])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin
        if row[0].value and str(row[0].value).startswith("【"):
            row[0].fill = section_fill
            row[0].font = header_font
            row[1].fill = section_fill

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72

    # notes
    if result.notes:
        ws.append(["", ""])
        ws.append(["系统备注", ""])
        for n in result.notes:
            ws.append(["note", n])

    # ---- 交易明细 ----
    ws2 = wb.create_sheet("交易明细")
    headers = [
        "序号",
        "证券代码",
        "代码",
        "信号日期",
        "指标",
        "卦名",
        "爻位",
        "卦序",
        "变卦",
        "操作信号",
        "state_id",
        "卦象简判",
        "买入日期",
        "买入价",
        "卖出日期",
        "卖出价",
        "数量",
        "买入金额",
        "卖出金额",
        "买入手续费",
        "卖出手续费及印花税",
        "毛利润",
        "净利润",
        "毛收益率%",
        "净收益率%",
        "是否盈利",
        "卖出原因",
        "状态",
    ]
    excel_cap = 3000
    trips_for_excel = list(trips)
    excel_truncated = False
    if len(trips_for_excel) > excel_cap:
        excel_truncated = True
        trips_for_excel = trips_for_excel[:excel_cap]

    # banner
    _bw = repro.get("buy_weekday")
    _ew = repro.get("exit_weekday")
    _sched = repro.get("schedule_mode") or (
        "weekday" if _bw is not None or _ew is not None else "tn"
    )
    if _sched == "weekday":
        _sched_txt = (
            f"schedule=weekday buy_wd={_bw} exit_wd={_ew} "
            f"buy_on={repro.get('buy_on') or 'open'} sell_on={repro.get('sell_on') or 'open'}"
        )
    else:
        _sched_txt = (
            f"schedule=tn entry_lag={repro.get('entry_lag')} hold={repro.get('hold')} "
            f"buy_on={repro.get('buy_on') or 'open'} sell_on={repro.get('sell_on') or 'open'}"
        )
    ws2.append(
        [
            f"run_id={result.run_id} | 区间 {repro.get('start') or ''}~{repro.get('end') or ''} | "
            f"period={repro.get('period')} | {_sched_txt} | "
            f"{repro.get('bagua_filter_label') or '无八卦过滤'} | "
            f"已平{len(closed)} 盈{win_n} 亏{loss_n} 未平{open_n} | 合计净利润≈{_fmt_num(net_sum, 2)} | "
            f"明细行{len(trips)}"
            + (f"（Excel仅预览前{excel_cap}行，完整见 trades.csv）" if excel_truncated else "")
        ]
    )
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws2["A1"].font = Font(bold=True, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws2.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws2.cell(2, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E86C1")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    green = PatternFill("solid", fgColor="D5F5E3")
    red = PatternFill("solid", fgColor="FADBD8")
    gray = PatternFill("solid", fgColor="F2F3F4")

    # openpyxl is very slow for tens of thousands of styled rows (UI stuck at 96%).
    # Full FIFO list is always in trades.csv; Excel keeps a preview + summary.
    for t in trips_for_excel:
        row = [
            t.get("序号"),
            t.get("证券代码"),
            t.get("代码"),
            _fmt_date(t.get("信号日期")),
            t.get("指标"),
            t.get("卦名"),
            t.get("爻位"),
            t.get("卦序"),
            t.get("变卦"),
            t.get("操作信号"),
            t.get("state_id"),
            t.get("卦象简判"),
            _fmt_date(t.get("买入日期")),
            _fmt_num(t.get("买入价"), 4),
            _fmt_date(t.get("卖出日期")),
            _fmt_num(t.get("卖出价"), 4) if t.get("卖出价") != "" else "",
            t.get("数量"),
            _fmt_num(t.get("买入金额"), 2),
            _fmt_num(t.get("卖出金额"), 2) if t.get("卖出金额") != "" else "",
            _fmt_num(t.get("买入手续费"), 4),
            _fmt_num(t.get("卖出手续费及印花税"), 4) if t.get("卖出手续费及印花税") != "" else "",
            _fmt_num(t.get("毛利润"), 2) if t.get("毛利润") != "" else "",
            _fmt_num(t.get("净利润"), 2) if t.get("净利润") != "" else "",
            _fmt_pct(t.get("毛收益率")) if t.get("毛收益率") != "" else "",
            _fmt_pct(t.get("净收益率")) if t.get("净收益率") != "" else "",
            t.get("是否盈利"),
            t.get("卖出原因"),
            t.get("状态"),
        ]
        ws2.append(row)
        r_idx = ws2.max_row
        fill = None
        if t.get("状态") == "未平仓":
            fill = gray
        elif t.get("是否盈利") == "盈利":
            fill = green
        elif t.get("是否盈利") == "亏损":
            fill = red
        for c in range(1, len(headers) + 1):
            cell = ws2.cell(r_idx, c)
            cell.border = thin
            if fill:
                cell.fill = fill

    widths = [6, 16, 10, 12, 18, 14, 10, 36, 12, 10, 12, 10, 8, 12, 12, 12, 14, 10, 10, 10, 10, 8, 16, 10]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws2.max_row}"
    ws2.freeze_panes = "A3"

    # ---- 说明（简短一页，不拆多表）----
    ws3 = wb.create_sheet("说明")
    for line in [
        "本工作簿为单次回测的一份总表，研究用途，非投资建议。",
        "「汇总」：组合层收益/回撤/买卖次数等。",
        "「交易明细」：按 FIFO 将买入与卖出配对；信号日期取不晚于买入日的最近信号。",
        "【交易日程 · 双层模型】前台用「信号星期 / 买入日星期 / 平仓日星期 / 开盘·收盘」配置；",
        "后台一律在 A 股交易日日历上求解（T+N 体系）：有 buy_weekday 时用 next_weekday_trading_day（覆盖 entry_lag 步进），",
        "有 exit_weekday 时同理覆盖 hold；未设星期时仍用 entry_lag + hold。节假日自动顺延到之后第一个可交易日。",
        "「按星期」不是另一套经济逻辑，而是 T+N 在星期锚定约束下的日历求解。",
        "启用八卦时默认可按「最佳3爻」等方案过滤；明细列「卦名/爻位/卦象简判」来自信号日 OHLC 标注。",
        "交易明细超过 3000 行时，Excel 仅预览前 3000 行；完整明细见同目录 trades.csv。",
        "毛利润 = 卖出金额 - 买入金额；净利润 = 毛利润 - 买入手续费 - 卖出手续费及印花税。",
        "收益率分母为买入金额。卖出原因含 hold_expired / stop_loss / take_profit 等。",
        "hold_expired：时间止损或星期平仓日强制平仓，成交价按 sell_on（开/收盘）；止损/止盈为触发后下一可交易日开盘价。",
        "组合层总收益受仓位权重、资金不足拒单等影响，与明细净利润简单加总可能不完全一致。",
        f"run_id={result.run_id}",
        f"indicator_ids={repro.get('indicator_ids')}",
        f"schedule_mode={repro.get('schedule_mode')} buy_weekday={repro.get('buy_weekday')} exit_weekday={repro.get('exit_weekday')}",
        f"entry_lag={repro.get('entry_lag')} hold={repro.get('hold')} buy_on={repro.get('buy_on')} sell_on={repro.get('sell_on')}",
        f"price_mode={repro.get('price_mode')} research_unadjusted={repro.get('research_unadjusted')}",
    ]:
        ws3.append([line])
    ws3.column_dimensions["A"].width = 100

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp)
    tmp.replace(path)
    return path
