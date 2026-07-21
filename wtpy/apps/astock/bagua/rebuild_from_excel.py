# -*- coding: utf-8 -*-
"""Rebuild bagua_384.json from the authoritative Excel (with 操作信号)."""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

TRIGRAMS = {
    1: {"id": 1, "name": "乾", "alias": "天", "symbol": "☰"},
    2: {"id": 2, "name": "兑", "alias": "泽", "symbol": "☱"},
    3: {"id": 3, "name": "离", "alias": "火", "symbol": "☲"},
    4: {"id": 4, "name": "震", "alias": "雷", "symbol": "☳"},
    5: {"id": 5, "name": "巽", "alias": "风", "symbol": "☴"},
    6: {"id": 6, "name": "坎", "alias": "水", "symbol": "☵"},
    7: {"id": 7, "name": "艮", "alias": "山", "symbol": "☶"},
    8: {"id": 8, "name": "坤", "alias": "地", "symbol": "☷"},
}

YAO_ORDER_MAP = {
    "初九": 1,
    "初六": 1,
    "九二": 2,
    "六二": 2,
    "九三": 3,
    "六三": 3,
    "九四": 4,
    "六四": 4,
    "九五": 5,
    "六五": 5,
    "上九": 6,
    "上六": 6,
}

SELF_MAP = {
    "乾为天": (1, 1),
    "坤为地": (8, 8),
    "震为雷": (4, 4),
    "艮为山": (7, 7),
    "坎为水": (6, 6),
    "离为火": (3, 3),
    "巽为风": (5, 5),
    "兑为泽": (2, 2),
}

DEFAULT_RULE_VERSION = "gua_rules_v20260721"

# Known aliases: short biangua name in Excel → full gua_name style
BIANGUA_ALIASES: Dict[str, str] = {}


def _alias_to_id() -> Dict[str, int]:
    return {v["alias"]: k for k, v in TRIGRAMS.items()}


def parse_ul(name: str) -> Tuple[int, int]:
    if name in SELF_MAP:
        return SELF_MAP[name]
    alias_to_id = _alias_to_id()
    return alias_to_id[name[0]], alias_to_id[name[1]]


def make_state_id(gua_order: int, yao_order: int) -> str:
    return f"{int(gua_order):02d}-{int(yao_order)}"


def rebuild_knowledge_from_excel(
    xlsx_path: Path | str,
    out_json: Path | str,
    *,
    rule_version: str = DEFAULT_RULE_VERSION,
) -> dict:
    """Rebuild bagua_384.json from Excel authority (supports 9-col with 操作信号)."""
    import openpyxl

    xlsx_path = Path(xlsx_path)
    out_json = Path(out_json)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    sha = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()

    headers = [ws.cell(1, c).value for c in range(1, 12)]
    # Expected: 卦象卦名,卦辞原文,卦核心总纲,爻位,变卦,爻辞原文,个股行情简判,备注&实操总结[,操作信号]
    has_action = any(h and "操作" in str(h) for h in headers)

    entries: List[dict] = []
    current: Optional[dict] = None
    gua_idx = 0
    r = 2
    while True:
        vals = [ws.cell(r, c).value for c in range(1, 10)]
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        gname, gci, gang, yao_pos, biangua, yao_ci, pan, note, action = vals
        if gname:
            full = str(gname).strip()
            if current is None or full != current["full_name"]:
                gua_idx += 1
                m = re.match(r"^(.)(.+)$", full)
                symbol = m.group(1) if m else ""
                rest = m.group(2) if m else full
                current = {
                    "gua_order": gua_idx,
                    "full_name": full,
                    "gua_symbol": symbol,
                    "gua_name": rest,
                    "gua_ci": str(gci).strip() if gci else "",
                    "core_gang": str(gang).strip() if gang else "",
                }
            else:
                if gci and str(gci).strip():
                    current["gua_ci"] = str(gci).strip()
                if gang and str(gang).strip():
                    current["core_gang"] = str(gang).strip()
        elif current is None:
            raise ValueError(f"missing hexagram name at row {r}")
        else:
            if gci and str(gci).strip():
                current["gua_ci"] = str(gci).strip()
            if gang and str(gang).strip():
                current["core_gang"] = str(gang).strip()

        yao = str(yao_pos).strip() if yao_pos else ""
        yao_order = YAO_ORDER_MAP.get(yao)
        if yao_order is None:
            raise ValueError(f"unknown yao at row {r}: {yao!r}")
        u, lo = parse_ul(current["gua_name"])
        bg_name = str(biangua).strip() if biangua else ""
        act = str(action).strip() if (has_action and action) else ""
        pan_s = str(pan).strip() if pan else ""
        yao_ci_s = str(yao_ci).strip() if yao_ci else ""
        note_s = str(note).strip() if note else ""
        state_id = make_state_id(current["gua_order"], yao_order)
        entries.append(
            {
                "excel_row": r,
                "state_id": state_id,
                "main_hexagram_id": current["gua_order"],
                "gua_order": current["gua_order"],
                "gua_symbol": current["gua_symbol"],
                "hexagram_symbol": current["gua_symbol"],
                "gua_name": current["gua_name"],
                "main_hexagram_name": current["gua_name"],
                "full_name": current["full_name"],
                "gua_ci": current["gua_ci"],
                "core_gang": current["core_gang"],
                "yao_order": yao_order,
                "line_index": yao_order,
                "yao_name": yao,
                "line_name": yao,
                "biangua": bg_name,
                "changed_hexagram_name": bg_name,
                "changed_hexagram_id": None,
                "yao_ci": yao_ci_s,
                "line_text": yao_ci_s,
                "market_judgement": pan_s,
                "market_summary": pan_s,
                "note": note_s,
                "action_signal": act,
                "upper": u,
                "lower": lo,
                "upper_name": TRIGRAMS[u]["name"],
                "lower_name": TRIGRAMS[lo]["name"],
                "upper_alias": TRIGRAMS[u]["alias"],
                "lower_alias": TRIGRAMS[lo]["alias"],
                "upper_symbol": TRIGRAMS[u]["symbol"],
                "lower_symbol": TRIGRAMS[lo]["symbol"],
            }
        )
        r += 1

    by_go: Dict[int, List[dict]] = {}
    for e in entries:
        by_go.setdefault(e["gua_order"], []).append(e)
    for go, rows in by_go.items():
        gang = next((x["core_gang"] for x in rows if x["core_gang"]), "")
        gci = next((x["gua_ci"] for x in rows if x["gua_ci"]), "")
        for x in rows:
            x["core_gang"] = gang
            x["gua_ci"] = gci

    # Resolve changed_hexagram_id by matching biangua short name against gua_name / full_name
    name_to_id: Dict[str, int] = {}
    for go, rows in by_go.items():
        nm = rows[0]["gua_name"]
        name_to_id[nm] = go
        # also short stems
        for alias in (nm, nm.replace("为", "")):
            name_to_id.setdefault(alias, go)
        # common short forms in Excel: 姤、同人、履 …
        # match if biangua is substring of gua_name or vice versa
    for e in entries:
        bg = e.get("biangua") or ""
        if not bg:
            e["changed_hexagram_id"] = None
            continue
        cid = name_to_id.get(bg)
        if cid is None:
            for nm, gid in name_to_id.items():
                if bg in nm or nm in bg:
                    cid = gid
                    break
        e["changed_hexagram_id"] = cid

    sigs = Counter(e["action_signal"] for e in entries if e["action_signal"])
    empty_bg = sum(1 for e in entries if not e["biangua"])

    kb = {
        "source_file": xlsx_path.name,
        "source_sha256": sha,
        "source_path": str(xlsx_path.resolve()),
        "rule_version": rule_version,
        "trigrams": TRIGRAMS,
        "entries": entries,
        "count_gua": len(by_go),
        "count_yao": len(entries),
        "action_signal_counts": dict(sigs),
        "empty_biangua_count": empty_bg,
        "headers": headers,
    }
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(kb, ensure_ascii=False, indent=2), encoding="utf-8")
    return kb


def validate_knowledge(kb: dict) -> Dict[str, Any]:
    entries = kb.get("entries") or []
    issues: List[str] = []
    by_go: Dict[int, List[dict]] = defaultdict(list)
    for e in entries:
        by_go[int(e["gua_order"])].append(e)
    if len(entries) != 384:
        issues.append(f"expected 384 states, got {len(entries)}")
    if len(by_go) != 64:
        issues.append(f"expected 64 main hexagrams, got {len(by_go)}")
    sids = [e.get("state_id") for e in entries]
    if len(sids) != len(set(sids)):
        issues.append("duplicate state_id")
    for go, rows in sorted(by_go.items()):
        if len(rows) != 6:
            issues.append(f"gua_order={go} has {len(rows)} lines, expected 6")
        orders = sorted(int(r["yao_order"]) for r in rows)
        if orders != [1, 2, 3, 4, 5, 6]:
            issues.append(f"gua_order={go} bad yao_order {orders}")
    missing_sig = sum(1 for e in entries if not e.get("action_signal"))
    empty_bg = sum(1 for e in entries if not e.get("biangua"))
    return {
        "ok": len(issues) == 0,
        "issues": issues,
        "count_entries": len(entries),
        "count_gua": len(by_go),
        "unique_state_ids": len(set(sids)),
        "action_signal_counts": dict(Counter(e.get("action_signal") or "" for e in entries)),
        "empty_biangua": empty_bg,
        "missing_action_signal": missing_sig,
        "rule_version": kb.get("rule_version"),
        "source_file": kb.get("source_file"),
        "source_sha256": kb.get("source_sha256"),
    }


def default_excel_path(repo_root: Optional[Path] = None) -> Path:
    root = repo_root or Path(__file__).resolve().parents[4]
    # parents: bagua -> astock -> apps -> wtpy -> repo
    candidates = [
        root / "指标" / "64卦384爻行情简判完整版-已添加操作信号.xlsx",
        root / "指标" / "64卦384爻行情简判完整版-已添加操作信号(1).xlsx",
    ]
    for c in candidates:
        if c.exists():
            return c
    # fuzzy
    ind = root / "指标"
    if ind.is_dir():
        for p in ind.glob("*操作信号*.xlsx"):
            return p
        for p in ind.glob("*384*.xlsx"):
            return p
    raise FileNotFoundError("no bagua excel found under 指标/")


def main(argv: Optional[List[str]] = None) -> int:
    argv = list(argv or sys.argv[1:])
    bagua_dir = Path(__file__).resolve().parent
    out_json = bagua_dir / "bagua_384.json"
    report_path = bagua_dir / "gua_validation_report.json"
    xlsx = Path(argv[0]) if argv else default_excel_path(Path(__file__).resolve().parents[4])
    # fix root: file is wtpy/apps/astock/bagua -> parents[3]=wtpy, parents[4]=repo
    try:
        xlsx = Path(argv[0]) if argv else default_excel_path()
    except FileNotFoundError:
        # try alternate parent depth
        for depth in (3, 4, 5):
            try:
                xlsx = default_excel_path(Path(__file__).resolve().parents[depth])
                break
            except Exception:
                continue
        else:
            raise
    print("excel:", xlsx)
    kb = rebuild_knowledge_from_excel(xlsx, out_json)
    report = validate_knowledge(kb)
    report["out_json"] = str(out_json)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
