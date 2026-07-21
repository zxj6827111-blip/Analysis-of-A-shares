"""Bagua OHLC calculator and knowledge base."""

from __future__ import annotations

import hashlib
import json
import re
from dataclasses import asdict, dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


TRIGRAMS = {
    1: {"id": 1, "name": "乾", "alias": "天", "symbol": "☰"},
    2: {"id": 2, "name": "兑", "alias": "泽", "symbol": "☱"},
    3: {"id": 3, "name": "离", "alias": "火", "symbol": "☲"},
    4: {"id": 4, "name": "震", "alias": "雷", "symbol": "☳"},
    5: {"id": 5, "name": "巽", "alias": "风", "symbol": "☴"},
    6: {"id": 6, "name": "坎", "alias": "水", "symbol": "☵"},
    7: {"id": 7, "name": "艮", "alias": "山", "symbol": "☶"},
    8: {"id": 8, "name": "坤", "alias": "地", "symbol": "☷"},
}

YAO_ORDER_MAP = {
    "初九": 1,
    "初六": 1,
    "九二": 2,
    "六二": 2,
    "九三": 3,
    "六三": 3,
    "九四": 4,
    "六四": 4,
    "九五": 5,
    "六五": 5,
    "上九": 6,
    "上六": 6,
}


@dataclass
class BaguaResult:
    open_price: str
    high_price: str
    low_price: str
    close_price: str
    open_digit_sum: int
    high_digit_sum: int
    low_digit_sum: int
    close_digit_sum: int
    upper_id: int
    lower_id: int
    upper_name: str
    lower_name: str
    upper_alias: str
    lower_alias: str
    upper_symbol: str
    lower_symbol: str
    yao_order: int
    yao_name: str
    gua_order: Optional[int]
    gua_symbol: str
    gua_name: str
    full_name: str
    gua_ci: str
    core_gang: str
    yao_ci: str
    market_judgement: str
    biangua: str
    note: str
    state_id: str = ""
    action_signal: str = ""
    main_hexagram_id: Optional[int] = None
    main_hexagram_name: str = ""
    changed_hexagram_name: str = ""
    changed_hexagram_id: Optional[int] = None
    market_summary: str = ""
    line_index: Optional[int] = None
    line_name: str = ""
    line_text: str = ""
    hexagram_symbol: str = ""

    def to_dict(self) -> dict:
        d = asdict(self)
        if not d.get("state_id") and d.get("gua_order") is not None and d.get("yao_order") is not None:
            d["state_id"] = f"{int(d['gua_order']):02d}-{int(d['yao_order'])}"
        if not d.get("market_summary"):
            d["market_summary"] = d.get("market_judgement") or ""
        if not d.get("changed_hexagram_name"):
            d["changed_hexagram_name"] = d.get("biangua") or ""
        if not d.get("main_hexagram_name"):
            d["main_hexagram_name"] = d.get("gua_name") or ""
        if d.get("main_hexagram_id") is None and d.get("gua_order") is not None:
            d["main_hexagram_id"] = d.get("gua_order")
        if not d.get("line_name"):
            d["line_name"] = d.get("yao_name") or ""
        if d.get("line_index") is None:
            d["line_index"] = d.get("yao_order")
        if not d.get("line_text"):
            d["line_text"] = d.get("yao_ci") or ""
        if not d.get("hexagram_symbol"):
            d["hexagram_symbol"] = d.get("gua_symbol") or ""
        return d


def format_price_2(price) -> str:
    d = Decimal(str(price)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{d:.2f}"


def digit_sum_price(price) -> int:
    s = format_price_2(price)
    total = 0
    for ch in s:
        if ch.isdigit():
            total += int(ch)
    return total


def mod_map(value: int, modulus: int) -> int:
    r = value % modulus
    return modulus if r == 0 else r


class BaguaKnowledge:
    REQUIRED_FIELDS = (
        "gua_order",
        "gua_symbol",
        "gua_name",
        "full_name",
        "gua_ci",
        "core_gang",
        "yao_order",
        "yao_name",
        "yao_ci",
        "market_judgement",
        "upper",
        "lower",
    )

    def __init__(self, path: Path | str):
        self.path = Path(path)
        raw = json.loads(self.path.read_text(encoding="utf-8"))
        self.source_file = raw.get("source_file")
        self.source_sha256 = raw.get("source_sha256")
        self.source_path = raw.get("source_path")
        self.trigrams = raw.get("trigrams") or TRIGRAMS
        self.entries: List[dict] = raw.get("entries") or []
        self._by_key: Dict[Tuple[int, int, int], dict] = {}
        self._by_gua_order: Dict[int, List[dict]] = {}
        self._by_ul: Dict[Tuple[int, int], List[dict]] = {}
        for e in self.entries:
            key = (int(e["upper"]), int(e["lower"]), int(e["yao_order"]))
            self._by_key[key] = e
            go = int(e["gua_order"])
            self._by_gua_order.setdefault(go, []).append(e)
            ul = (int(e["upper"]), int(e["lower"]))
            self._by_ul.setdefault(ul, []).append(e)

    def validate(self) -> List[str]:
        issues: List[str] = []
        if len(self.entries) != 384:
            issues.append(f"expected 384 yaos, got {len(self.entries)}")
        if len(self._by_gua_order) != 64:
            issues.append(f"expected 64 gua_order, got {len(self._by_gua_order)}")
        expected_orders = set(range(1, 65))
        got_orders = set(self._by_gua_order.keys())
        if got_orders != expected_orders:
            issues.append(
                f"gua_order set mismatch missing={sorted(expected_orders-got_orders)[:5]} "
                f"extra={sorted(got_orders-expected_orders)[:5]}"
            )
        for go, rows in self._by_gua_order.items():
            if len(rows) != 6:
                issues.append(f"gua_order={go} has {len(rows)} yaos, expected 6")
                continue
            orders = sorted(int(r["yao_order"]) for r in rows)
            if orders != [1, 2, 3, 4, 5, 6]:
                issues.append(f"gua_order={go} bad yao_order {orders}")
            # group-level field consistency
            for fld in ("full_name", "gua_name", "gua_symbol", "gua_ci", "core_gang", "upper", "lower"):
                vals = {str(r.get(fld) or "") for r in rows}
                if len(vals) != 1:
                    issues.append(f"gua_order={go} inconsistent {fld}: {vals}")
            if not rows[0].get("core_gang"):
                issues.append(f"gua_order={go} empty core_gang")
            if not rows[0].get("gua_ci"):
                issues.append(f"gua_order={go} empty gua_ci")
        for i, e in enumerate(self.entries):
            for f in self.REQUIRED_FIELDS:
                if e.get(f) in (None, ""):
                    issues.append(f"entry[{i}] missing {f} ({e.get('full_name')})")
                    break
        return issues

    def lookup(self, upper: int, lower: int, yao_order: int) -> Optional[dict]:
        return self._by_key.get((upper, lower, yao_order))

    def excel_consistency_check(self, xlsx_path: Path | str) -> List[str]:
        """Row-by-row compare against Excel authority (A1:H385)."""
        import openpyxl

        xlsx_path = Path(xlsx_path)
        issues: List[str] = []
        if not xlsx_path.exists():
            return [f"excel missing: {xlsx_path}"]
        digest = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()
        if self.source_sha256 and digest != self.source_sha256:
            issues.append(
                f"source sha mismatch json={self.source_sha256} excel={digest}"
            )
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        # rebuild expected names/yao/judgement with same forward-fill as rebuild
        # Compare each entry excel_row if present, else sequential
        for idx, e in enumerate(self.entries):
            r = int(e.get("excel_row") or (idx + 2))
            a = ws.cell(r, 1).value
            d = ws.cell(r, 4).value
            f = ws.cell(r, 6).value
            g = ws.cell(r, 7).value
            if a is not None and str(a).strip() != e.get("full_name"):
                issues.append(f"row {r} full_name excel={a!r} json={e.get('full_name')!r}")
            if d is not None and str(d).strip() != e.get("yao_name"):
                issues.append(f"row {r} yao_name excel={d!r} json={e.get('yao_name')!r}")
            if f is not None and str(f).strip() != e.get("yao_ci"):
                issues.append(f"row {r} yao_ci mismatch at {r}")
            if g is not None and str(g).strip() != e.get("market_judgement"):
                issues.append(f"row {r} market_judgement mismatch at {r}")
        return issues


class BaguaCalculator:
    def __init__(self, knowledge: Optional[BaguaKnowledge] = None):
        self.knowledge = knowledge

    @classmethod
    def from_json(cls, path: Path | str) -> "BaguaCalculator":
        return cls(BaguaKnowledge(path))

    def calculate(
        self,
        *,
        open_price,
        high_price,
        low_price,
        close_price,
    ) -> BaguaResult:
        o_s = format_price_2(open_price)
        h_s = format_price_2(high_price)
        l_s = format_price_2(low_price)
        c_s = format_price_2(close_price)

        o_sum = digit_sum_price(o_s)
        h_sum = digit_sum_price(h_s)
        l_sum = digit_sum_price(l_s)
        c_sum = digit_sum_price(c_s)

        upper_id = mod_map(o_sum, 8)
        lower_id = mod_map(c_sum, 8)
        yao_order = mod_map(h_sum + l_sum, 6)

        u = TRIGRAMS[upper_id]
        lo = TRIGRAMS[lower_id]

        gua_order = None
        gua_symbol = ""
        gua_name = f"{u['alias']}{lo['alias']}"
        full_name = ""
        gua_ci = ""
        core_gang = ""
        yao_ci = ""
        market_judgement = ""
        biangua = ""
        note = ""
        yao_name = ""

        entry = None
        if self.knowledge:
            entry = self.knowledge.lookup(upper_id, lower_id, yao_order)
        state_id = ""
        action_signal = ""
        main_hexagram_id = None
        main_hexagram_name = ""
        changed_hexagram_name = ""
        changed_hexagram_id = None
        market_summary = ""
        line_index = None
        line_name = ""
        line_text = ""
        hexagram_symbol = ""
        if entry:
            gua_order = entry.get("gua_order")
            gua_symbol = entry.get("gua_symbol") or ""
            gua_name = entry.get("gua_name") or gua_name
            full_name = entry.get("full_name") or f"{gua_symbol}{gua_name}"
            gua_ci = entry.get("gua_ci") or ""
            core_gang = entry.get("core_gang") or ""
            yao_ci = entry.get("yao_ci") or ""
            market_judgement = entry.get("market_judgement") or entry.get("market_summary") or ""
            biangua = entry.get("biangua") or entry.get("changed_hexagram_name") or ""
            note = entry.get("note") or ""
            yao_name = entry.get("yao_name") or ""
            state_id = entry.get("state_id") or ""
            if not state_id and gua_order is not None and yao_order:
                state_id = f"{int(gua_order):02d}-{int(yao_order)}"
            action_signal = entry.get("action_signal") or ""
            main_hexagram_id = entry.get("main_hexagram_id") or gua_order
            main_hexagram_name = entry.get("main_hexagram_name") or gua_name
            changed_hexagram_name = entry.get("changed_hexagram_name") or biangua
            changed_hexagram_id = entry.get("changed_hexagram_id")
            market_summary = entry.get("market_summary") or market_judgement
            line_index = entry.get("line_index") or entry.get("yao_order") or yao_order
            line_name = entry.get("line_name") or yao_name
            line_text = entry.get("line_text") or yao_ci
            hexagram_symbol = entry.get("hexagram_symbol") or gua_symbol
        else:
            yao_name = {1: "初", 2: "二", 3: "三", 4: "四", 5: "五", 6: "上"}.get(
                yao_order, str(yao_order)
            )
            full_name = gua_name

        return BaguaResult(
            open_price=o_s,
            high_price=h_s,
            low_price=l_s,
            close_price=c_s,
            open_digit_sum=o_sum,
            high_digit_sum=h_sum,
            low_digit_sum=l_sum,
            close_digit_sum=c_sum,
            upper_id=upper_id,
            lower_id=lower_id,
            upper_name=u["name"],
            lower_name=lo["name"],
            upper_alias=u["alias"],
            lower_alias=lo["alias"],
            upper_symbol=u["symbol"],
            lower_symbol=lo["symbol"],
            yao_order=yao_order,
            yao_name=yao_name,
            gua_order=gua_order,
            gua_symbol=gua_symbol,
            gua_name=gua_name,
            full_name=full_name,
            gua_ci=gua_ci,
            core_gang=core_gang,
            yao_ci=yao_ci,
            market_judgement=market_judgement,
            biangua=biangua,
            note=note,
            state_id=state_id
            or (
                f"{int(gua_order):02d}-{int(yao_order)}"
                if gua_order is not None
                else ""
            ),
            action_signal=action_signal,
            main_hexagram_id=main_hexagram_id if main_hexagram_id is not None else gua_order,
            main_hexagram_name=main_hexagram_name or gua_name,
            changed_hexagram_name=changed_hexagram_name or biangua,
            changed_hexagram_id=changed_hexagram_id,
            market_summary=market_summary or market_judgement,
            line_index=line_index if line_index is not None else yao_order,
            line_name=line_name or yao_name,
            line_text=line_text or yao_ci,
            hexagram_symbol=hexagram_symbol or gua_symbol,
        )


def rebuild_knowledge_from_excel(xlsx_path: Path, out_json: Path, **kwargs) -> dict:
    """Rebuild bagua_384.json from Excel authority (optional action_signal column)."""
    from .rebuild_from_excel import rebuild_knowledge_from_excel as _rebuild

    return _rebuild(xlsx_path, out_json, **kwargs)
