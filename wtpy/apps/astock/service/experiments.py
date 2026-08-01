# -*- coding: utf-8 -*-
"""Experiment center MVP (Stage E): expand param grid, queue variants, aggregate results."""
from __future__ import annotations

import itertools
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Dict, List, Optional, Sequence, Tuple

from ..bagua.filter_rules import BEST3, GuaFilter
from ..config import AStockConfig, get_default_config
from .backtest import BacktestRequest, BacktestService
from . import db as exp_db

# Soft cap — UI must warn; hard refuse above this unless force=True
DEFAULT_MAX_VARIANTS = 50
# Soft default 50; force=true to exceed max_variants; hard refuse above HARD_MAX.
# Env ASTOCK_EXP_HARD_MAX_VARIANTS can raise the hard ceiling for long-running grids
# (clamped to [500, 20000]). Aligns with research.planner spirit (2000).
def _resolve_hard_max_variants() -> int:
    import os
    raw = (os.environ.get("ASTOCK_EXP_HARD_MAX_VARIANTS") or "").strip()
    try:
        n = int(raw) if raw else 2000
    except ValueError:
        n = 2000
    return max(500, min(int(n or 2000), 20000))


HARD_MAX_VARIANTS = _resolve_hard_max_variants()

# Weekday schedule templates (UI labels → engine fields)
WEEKDAY_TEMPLATES = {
    "fri_signal_mon_buy_thu_exit": {
        "label": "仅周五信号·周一买·周四平",
        "signal_weekdays": [5],
        "buy_weekday": 1,
        "exit_weekday": 4,
        "buy_on": "open",
        "sell_on": "open",
        "entry_lag": 1,
        "hold": 1,
    },
    "all_signal_tn12": {
        "label": "不限信号·经典T+1开/T+2收",
        "signal_weekdays": None,
        "buy_weekday": None,
        "exit_weekday": None,
        "buy_on": "open",
        "sell_on": "close",
        "entry_lag": 1,
        "hold": 1,
    },
    "fri_signal_fri_buy_mon_exit": {
        "label": "仅周五信号·周五买·下周一平",
        "signal_weekdays": [5],
        "buy_weekday": 5,
        "exit_weekday": 1,
        "buy_on": "open",
        "sell_on": "open",
        "entry_lag": 1,
        "hold": 1,
    },
}

GUA_PRESETS = {
    "none": {
        "label": "无卦象",
        "gua_filter": {
            "enabled": False,
            "selection_mode": "none",
            "selected_main_hexagram_ids": [],
            "selected_state_ids": [],
            "selected_action_signals": [],
        },
        "with_bagua": False,
    },
    "best3": {
        "label": "最佳3爻",
        "gua_filter": {
            "enabled": True,
            "selection_mode": "exact_line",
            "selected_main_hexagram_ids": [],
            "selected_state_ids": ["24-1", "46-1", "11-1"],
            "selected_action_signals": [],
        },
        "with_bagua": True,
    },
    "bull": {
        "label": "偏多操作信号",
        "gua_filter": {
            "enabled": True,
            "selection_mode": "action_signal",
            "selected_main_hexagram_ids": [],
            "selected_state_ids": [],
            "selected_action_signals": ["新开仓", "加仓"],
        },
        "with_bagua": True,
    },
}



# ---------------------------------------------------------------------------
# Phase2 free axes (signal / buy / sell independent) + legacy template path
# Templates in WEEKDAY_TEMPLATES remain presets only; free axes ignore them.
# ---------------------------------------------------------------------------

_FREE_AXIS_KEYS = (
    "signal_weekdays_options",
    "buy_options",
    "sell_options",
    "take_profit_list",
)


def _has_free_axes(payload_or_kwargs: dict) -> bool:
    """True if any independent-axis field is explicitly provided (non-empty)."""
    for k in _FREE_AXIS_KEYS:
        if k not in payload_or_kwargs:
            continue
        v = payload_or_kwargs.get(k)
        if v is None:
            continue
        if isinstance(v, (list, tuple)) and len(v) == 0:
            continue
        return True
    return False


def _normalize_signal_weekdays(val: Any) -> Optional[List[int]]:
    """None / [] / list of ints — empty list means unrestricted (None)."""
    if val is None:
        return None
    if isinstance(val, (list, tuple)):
        if len(val) == 0:
            return None
        out: List[int] = []
        for x in val:
            if x is None:
                continue
            out.append(int(x))
        return out or None
    return [int(val)]


def _resolve_buy_option(opt: dict) -> dict:
    """Normalize buy axis option.

    Do **not** invent hold=1 on the buy side — hold belongs to sell_options
    (fixed holding period). Returning hold=1 here previously overrode real
    sell hold (e.g. T20/T60) because merge used ``b.hold or s.hold``.
    """
    o = dict(opt or {})
    buy_on = str(o.get("buy_on") or "open").lower()
    if buy_on not in ("open", "close"):
        buy_on = "open"
    entry_lag = o.get("entry_lag")
    if entry_lag is None:
        entry_lag = 1
    out: Dict[str, Any] = {
        "buy_weekday": o.get("buy_weekday"),
        "buy_on": buy_on,
        "entry_lag": int(entry_lag),
    }
    # Only keep buy-side hold if explicitly provided (rare schedule packs).
    if o.get("hold") is not None:
        try:
            out["hold"] = int(o["hold"])
        except (TypeError, ValueError):
            pass
    return out


def _resolve_sell_option(opt: dict) -> dict:
    o = dict(opt or {})
    sell_on = str(o.get("sell_on") or "open").lower()
    if sell_on not in ("open", "close"):
        sell_on = "open"
    hold = o.get("hold")
    return {
        "exit_weekday": o.get("exit_weekday"),
        "sell_on": sell_on,
        "hold": int(hold) if hold is not None else None,
    }


def _merge_hold(buy_resolved: dict, sell_resolved: dict) -> int:
    """Prefer sell hold (fixed holding days); fall back to buy, then 1."""
    sh = sell_resolved.get("hold")
    if sh is not None:
        try:
            h = int(sh)
            if h >= 1:
                return h
        except (TypeError, ValueError):
            pass
    bh = buy_resolved.get("hold")
    if bh is not None:
        try:
            h = int(bh)
            if h >= 1:
                return h
        except (TypeError, ValueError):
            pass
    return 1


def _resolve_gua_option(gkey_or_filter: Any) -> Dict[str, Any]:
    """Resolve gua preset key or inline gua_filter dict → label/with_bagua/filter."""
    if gkey_or_filter is None or gkey_or_filter == "none":
        g = GUA_PRESETS["none"]
        return {
            "gua_key": "none",
            "gua_label": g["label"],
            "with_bagua": False,
            "gua_filter": dict(g["gua_filter"]),
        }
    if isinstance(gkey_or_filter, dict):
        gf = dict(gkey_or_filter)
        # allow wrapping {gua_filter: {...}, label: ...}
        if "gua_filter" in gf and isinstance(gf.get("gua_filter"), dict):
            inner = dict(gf["gua_filter"])
            label = str(gf.get("label") or gf.get("gua_label") or gf.get("manifest_name") or "custom")
            enabled = bool(inner.get("enabled", True))
            return {
                "gua_key": str(gf.get("gua_key") or gf.get("rule_id") or "custom"),
                "gua_label": label,
                "with_bagua": enabled,
                "gua_filter": inner,
            }
        enabled = bool(gf.get("enabled", True))
        mode = str(gf.get("selection_mode") or "none")
        if mode in ("", "none") and not gf.get("selected_state_ids"):
            enabled = False
        label = str(
            gf.get("label")
            or gf.get("gua_label")
            or gf.get("manifest_name")
            or "+".join(str(x) for x in (gf.get("selected_state_ids") or [])[:3])
            or "custom"
        )
        return {
            "gua_key": str(gf.get("rule_id") or gf.get("gua_key") or "custom"),
            "gua_label": label,
            "with_bagua": enabled,
            "gua_filter": {
                "enabled": enabled,
                "selection_mode": mode if enabled else "none",
                "selected_main_hexagram_ids": list(gf.get("selected_main_hexagram_ids") or []),
                "selected_state_ids": [str(x) for x in (gf.get("selected_state_ids") or [])],
                "selected_action_signals": list(gf.get("selected_action_signals") or []),
                "rule_version": gf.get("rule_version") or "",
            },
        }
    key = str(gkey_or_filter).strip()
    if key not in GUA_PRESETS:
        raise ValueError(f"unknown gua preset: {key}")
    g = GUA_PRESETS[key]
    return {
        "gua_key": key,
        "gua_label": g["label"],
        "with_bagua": bool(g.get("with_bagua", False)),
        "gua_filter": dict(g.get("gua_filter") or {}),
    }


def _normalize_gua_options(
    gua_keys: Optional[Sequence[Any]] = None,
    gua_filters: Optional[Sequence[Any]] = None,
) -> List[Any]:
    """Flatten gua_keys (str or dict) + gua_filters into one option list."""
    opts: List[Any] = []
    for g in list(gua_keys or []):
        if g is None or g == "":
            continue
        opts.append(g)
    for gf in list(gua_filters or []):
        if gf is None or gf == "":
            continue
        opts.append(gf)
    if not opts:
        opts = ["none"]
    # validate each resolves
    for o in opts:
        _resolve_gua_option(o)
    return opts


def _payload_to_parameter_space(payload: dict):
    """Build research.ParameterSpace from API/experiment payload. None if research unavailable."""
    try:
        from ..research.models import ParameterSpace
        from ..research.parameter_space import axes_from_legacy_templates
    except Exception:
        return None

    axes = {
        "signal_weekdays_options": payload.get("signal_weekdays_options"),
        "buy_options": payload.get("buy_options"),
        "sell_options": payload.get("sell_options"),
        "take_profit_list": payload.get("take_profit_list"),
    }
    use_free = _has_free_axes(axes)
    rule_ids = list(payload.get("rule_ids") or [])
    # Support preset keys + inline filters (dicts in gua_keys or gua_filters)
    raw_gua_keys = list(payload.get("gua_keys") or [])
    raw_gua_filters = list(payload.get("gua_filters") or [])
    if not raw_gua_keys and not raw_gua_filters:
        raw_gua_keys = ["none"]
    str_keys = [g for g in raw_gua_keys if isinstance(g, str)]
    dict_from_keys = [g for g in raw_gua_keys if isinstance(g, dict)]
    gua_filters_all = dict_from_keys + [g for g in raw_gua_filters if g is not None]
    holiday_policy = payload.get("holiday_policy") or "next_trading_day"
    period = payload.get("period") or "DAY"
    codes = payload.get("codes")
    start = payload.get("start")
    end = payload.get("end")
    account_mode = payload.get("account_mode") or "portfolio"
    research_unadjusted = bool(payload.get("research_unadjusted"))
    stop_loss_list = payload.get("stop_loss_list")
    take_profit_list = payload.get("take_profit_list")

    if use_free:
        sig = payload.get("signal_weekdays_options")
        if sig is None:
            signal_weekdays = [None]
        else:
            signal_weekdays = list(sig) if sig else [None]
        buy_modes = list(payload.get("buy_options") or [{"entry_lag": 1, "buy_on": "open"}])
        sell_modes = list(payload.get("sell_options") or [{"sell_on": "close"}])
        # periods axis via extra_axes when multiple
        periods = payload.get("periods")
        extra: Dict[str, List[Any]] = {}
        if periods and isinstance(periods, (list, tuple)) and len(periods) > 1:
            extra["period"] = [str(p).upper() for p in periods]
            period = str(periods[0]).upper()
        elif periods and isinstance(periods, (list, tuple)) and len(periods) == 1:
            period = str(periods[0]).upper()
        return ParameterSpace(
            rule_ids=rule_ids,
            period=period,
            signal_weekdays=signal_weekdays,
            buy_modes=buy_modes,
            sell_modes=sell_modes,
            gua_keys=str_keys or ([] if gua_filters_all else ["none"]),
            gua_filters=gua_filters_all,
            stop_loss_list=list(stop_loss_list if stop_loss_list is not None else [None]),
            take_profit_list=list(take_profit_list if take_profit_list is not None else [None]),
            holiday_policy=holiday_policy,
            codes=list(codes) if codes else None,
            start=start,
            end=end,
            account_mode=account_mode,
            research_unadjusted=research_unadjusted,
            extra_axes=extra,
        )

    # Legacy weekday_keys templates (preset string keys only)
    wd = payload.get("weekday_keys") or ["all_signal_tn12"]
    legacy_gua = str_keys or ["none"]
    try:
        return axes_from_legacy_templates(
            rule_ids=rule_ids,
            weekday_keys=list(wd),
            gua_keys=legacy_gua,
            stop_loss_list=stop_loss_list,
            take_profit_list=take_profit_list,
            period=period,
            codes=codes,
            start=start,
            end=end,
            account_mode=account_mode,
            research_unadjusted=research_unadjusted,
            holiday_policy=holiday_policy,
        )
    except Exception:
        return None


def _try_research_plan(payload: dict):
    """If research.planner.plan_experiment is available, prefer it.

    Normalizes planner keys to theoretical/rejected/actual/variants/...
    """
    try:
        from ..research.planner import plan_experiment
        from ..research.planner import HARD_MAX_VARIANTS as R_HARD
    except Exception:
        return None
    space = _payload_to_parameter_space(payload)
    if space is None:
        return None
    max_v = int(payload.get("max_variants") or DEFAULT_MAX_VARIANTS)
    force = bool(payload.get("force"))
    # Align hard max with experiments module for API consistency
    hard = int(payload.get("hard_max") or HARD_MAX_VARIANTS)
    try:
        plan = plan_experiment(space, max_variants=max_v, force=force, hard_max=hard)
    except Exception:
        return None
    if not isinstance(plan, dict):
        return None
    # Normalize key names
    theoretical = int(plan.get("theoretical_count") or plan.get("theoretical") or 0)
    rejected = int(plan.get("rejected_count") or plan.get("rejected") or 0)
    actual = int(plan.get("actual_count") or plan.get("actual") or 0)
    variants = list(plan.get("variants") or [])
    preview = list(plan.get("preview") or [])
    reasons = dict(plan.get("rejection_reasons") or {})
    msg = plan.get("error") or plan.get("message")
    return {
        "theoretical": theoretical,
        "rejected": rejected,
        "actual": actual,
        "rejection_reasons": reasons,
        "variants": variants,
        "preview": preview,
        "message": msg,
        "truncated": plan.get("truncated"),
        "hard_max": plan.get("hard_max", hard),
        "max_variants": plan.get("max_variants", max_v),
        "force": force,
        "source": "research.planner",
    }


def expand_param_grid_unified(
    *,
    rule_ids: Sequence[str],
    gua_keys: Optional[Sequence[Any]] = None,
    gua_filters: Optional[Sequence[Any]] = None,
    weekday_keys: Optional[Sequence[str]] = None,
    stop_loss_list: Optional[Sequence[Optional[float]]] = None,
    signal_weekdays_options: Optional[Sequence[Any]] = None,
    buy_options: Optional[Sequence[dict]] = None,
    sell_options: Optional[Sequence[dict]] = None,
    take_profit_list: Optional[Sequence[Optional[float]]] = None,
    holiday_policy: str = "next_trading_day",
    period: str = "DAY",
    periods: Optional[Sequence[str]] = None,
    codes: Optional[Sequence[str]] = None,
    start: Optional[int] = None,
    end: Optional[int] = None,
    account_mode: str = "portfolio",
    research_unadjusted: bool = False,
    collect_rejections: bool = False,
) -> Dict[str, Any]:
    """Expand grid from free axes OR legacy weekday_keys templates.

    When any free axis is provided, weekday_keys templates are ignored.
    ``gua_keys`` may mix preset strings and inline filter dicts.
    ``periods`` expands an extra period axis (DAY/WEEK/...).
    Returns dict with theoretical/rejected/actual/rejection_reasons/variants.
    """
    axes_payload = {
        "signal_weekdays_options": signal_weekdays_options,
        "buy_options": buy_options,
        "sell_options": sell_options,
        "take_profit_list": take_profit_list,
    }
    use_free = _has_free_axes(axes_payload)

    rules = list(rule_ids or [])
    if not rules:
        raise ValueError("rule_ids required")
    guas = _normalize_gua_options(gua_keys, gua_filters)
    sls = list(stop_loss_list if stop_loss_list is not None else [None])
    tps = list(take_profit_list if take_profit_list is not None else [None])
    period_list = [str(p).upper() for p in (periods or [period or "DAY"])]
    if not period_list:
        period_list = ["DAY"]

    rejection_reasons: Dict[str, int] = {}
    variants: List[Dict[str, Any]] = []
    theoretical = 0
    rejected = 0

    def _rej(reason: str) -> None:
        nonlocal rejected
        rejected += 1
        rejection_reasons[reason] = rejection_reasons.get(reason, 0) + 1

    base_codes = list(codes) if codes else ["sh600000", "sz000001"]

    if use_free:
        sig_opts = list(signal_weekdays_options) if signal_weekdays_options is not None else [None]
        if not sig_opts:
            sig_opts = [None]
        buy_opts = list(buy_options) if buy_options is not None else [
            {"entry_lag": 1, "buy_on": "open"}
        ]
        if not buy_opts:
            buy_opts = [{"entry_lag": 1, "buy_on": "open"}]
        sell_opts = list(sell_options) if sell_options is not None else [
            {"sell_on": "close"}
        ]
        if not sell_opts:
            sell_opts = [{"sell_on": "close"}]

        for rule_id, gopt, sig, bopt, sopt, sl, tp, per in itertools.product(
            rules, guas, sig_opts, buy_opts, sell_opts, sls, tps, period_list
        ):
            theoretical += 1
            g = _resolve_gua_option(gopt)
            b = _resolve_buy_option(bopt if isinstance(bopt, dict) else {})
            s = _resolve_sell_option(sopt if isinstance(sopt, dict) else {})

            if isinstance(sig, dict):
                sw = _normalize_signal_weekdays(sig.get("signal_weekdays"))
            else:
                sw = _normalize_signal_weekdays(sig)

            buy_wd = b.get("buy_weekday")
            exit_wd = s.get("exit_weekday")
            buy_on = b.get("buy_on") or "open"
            sell_on = s.get("sell_on") or "open"
            entry_lag = int(b.get("entry_lag") if b.get("entry_lag") is not None else 1)
            hold = _merge_hold(b, s)

            if buy_wd is not None:
                try:
                    bw = int(buy_wd)
                except (TypeError, ValueError):
                    _rej("invalid_buy_weekday")
                    continue
                if bw < 1 or bw > 5:
                    _rej("buy_weekday_out_of_range")
                    continue
                buy_wd = bw
            if exit_wd is not None:
                try:
                    ew = int(exit_wd)
                except (TypeError, ValueError):
                    _rej("invalid_exit_weekday")
                    continue
                if ew < 1 or ew > 5:
                    _rej("exit_weekday_out_of_range")
                    continue
                exit_wd = ew
            if entry_lag < 0:
                _rej("entry_lag_negative")
                continue
            if hold < 1:
                _rej("hold_lt_1")
                continue
            if (
                buy_wd is not None
                and exit_wd is not None
                and int(buy_wd) == int(exit_wd)
                and buy_on == "open"
                and sell_on == "open"
                and entry_lag == 0
            ):
                _rej("same_day_open_open_entry_lag0")
                continue

            meta = {
                "rule_id": rule_id,
                "gua_key": g["gua_key"],
                "gua_label": g["gua_label"],
                "weekday_key": None,
                "weekday_label": "free_axes",
                "stop_loss": sl,
                "take_profit": tp,
                "holiday_policy": holiday_policy,
                "signal_weekdays": sw,
                "buy_option": bopt if isinstance(bopt, dict) else {},
                "sell_option": sopt if isinstance(sopt, dict) else {},
                "period": per,
            }
            params: Dict[str, Any] = {
                "rule_ids": [rule_id],
                "period": per,
                "account_mode": account_mode,
                "codes": list(base_codes),
                "start": start,
                "end": end,
                "research_unadjusted": bool(research_unadjusted),
                "hold": hold,
                "entry_lag": entry_lag,
                "buy_weekday": buy_wd,
                "exit_weekday": exit_wd,
                "signal_weekdays": sw,
                "buy_on": buy_on,
                "sell_on": sell_on,
                "with_bagua": g["with_bagua"],
                "gua_filter": dict(g["gua_filter"] or {}),
                "stop_loss": sl,
                "take_profit": tp,
                "holiday_policy": holiday_policy,
                "_meta": meta,
            }
            variants.append(params)
    else:
        # legacy: only string preset keys
        str_guas = []
        for g in guas:
            if isinstance(g, str):
                str_guas.append(g)
            else:
                raise ValueError(
                    "inline gua_filters require free axes (buy_options/sell_options); "
                    "use free_axes mode for custom state_id filters"
                )
        if not str_guas:
            str_guas = ["none"]
        wds = list(weekday_keys or ["all_signal_tn12"])
        for w in wds:
            if w not in WEEKDAY_TEMPLATES:
                raise ValueError(f"unknown weekday template: {w}")
        for rule_id, gkey, wkey, sl, tp, per in itertools.product(
            rules, str_guas, wds, sls, tps, period_list
        ):
            theoretical += 1
            g = _resolve_gua_option(gkey)
            w = WEEKDAY_TEMPLATES[wkey]
            params = {
                "rule_ids": [rule_id],
                "period": per,
                "account_mode": account_mode,
                "codes": list(base_codes),
                "start": start,
                "end": end,
                "research_unadjusted": bool(research_unadjusted),
                "hold": w.get("hold", 1),
                "entry_lag": w.get("entry_lag", 1),
                "buy_weekday": w.get("buy_weekday"),
                "exit_weekday": w.get("exit_weekday"),
                "signal_weekdays": w.get("signal_weekdays"),
                "buy_on": w.get("buy_on", "open"),
                "sell_on": w.get("sell_on", "open"),
                "with_bagua": g["with_bagua"],
                "gua_filter": dict(g["gua_filter"] or {}),
                "stop_loss": sl,
                "take_profit": tp,
                "holiday_policy": holiday_policy,
                "_meta": {
                    "rule_id": rule_id,
                    "gua_key": g["gua_key"],
                    "gua_label": g["gua_label"],
                    "weekday_key": wkey,
                    "weekday_label": w.get("label"),
                    "stop_loss": sl,
                    "take_profit": tp,
                    "holiday_policy": holiday_policy,
                    "period": per,
                },
            }
            variants.append(params)

    actual = len(variants)
    return {
        "theoretical": theoretical,
        "rejected": rejected,
        "actual": actual,
        "rejection_reasons": rejection_reasons,
        "variants": variants,
    }


def estimate_grid_from_payload(payload: dict) -> dict:
    """Unified estimate / preview for legacy templates OR free axes.

    Prefers research.planner.plan_experiment when available; otherwise local
    expand_param_grid_unified. Returns theoretical/rejected/actual plus UI
    aliases estimated_variants/n/count.
    """
    max_v = int(payload.get("max_variants") or DEFAULT_MAX_VARIANTS)
    force = bool(payload.get("force"))
    hard = HARD_MAX_VARIANTS

    research_plan = _try_research_plan(payload)
    if isinstance(research_plan, dict) and research_plan.get("source") == "research.planner":
        theoretical = int(research_plan.get("theoretical") or 0)
        rejected = int(research_plan.get("rejected") or 0)
        variants = list(research_plan.get("variants") or [])
        actual = int(research_plan.get("actual") or len(variants))
        reasons = dict(research_plan.get("rejection_reasons") or {})
        preview = list(research_plan.get("preview") or variants[:50])
        msg = research_plan.get("message")
        hard = int(research_plan.get("hard_max") or hard)
    else:
        plan = expand_param_grid_unified(
            rule_ids=payload.get("rule_ids") or [],
            gua_keys=payload.get("gua_keys") or ["none"],
            gua_filters=payload.get("gua_filters"),
            weekday_keys=payload.get("weekday_keys"),
            stop_loss_list=payload.get("stop_loss_list"),
            signal_weekdays_options=payload.get("signal_weekdays_options"),
            buy_options=payload.get("buy_options"),
            sell_options=payload.get("sell_options"),
            take_profit_list=payload.get("take_profit_list"),
            holiday_policy=payload.get("holiday_policy") or "next_trading_day",
            period=payload.get("period") or "DAY",
            periods=payload.get("periods"),
            codes=payload.get("codes"),
            start=payload.get("start"),
            end=payload.get("end"),
            account_mode=payload.get("account_mode") or "portfolio",
            research_unadjusted=bool(payload.get("research_unadjusted")),
            collect_rejections=True,
        )
        theoretical = int(plan["theoretical"])
        rejected = int(plan["rejected"])
        actual = int(plan["actual"])
        reasons = dict(plan.get("rejection_reasons") or {})
        preview = list(plan.get("variants") or [])[:50]
        msg = None

    ok = True
    if actual > hard:
        ok = False
        msg = msg or ("组合数 %s 超过硬顶 %s" % (actual, hard))
    elif actual > max_v and not force:
        ok = False
        msg = msg or (
            "组合数 %s 超过上限 max_variants=%s；请缩小空间或提高 max_variants，或 force=true（硬顶 %s）"
            % (actual, max_v, hard)
        )
    elif actual > max_v and force:
        msg = msg or ("force=true：允许 %s 组（上限 %s，硬顶 %s）" % (actual, max_v, hard))
    elif actual > 20:
        msg = msg or ("组合数 %s 较大，建议先用演示池试跑" % actual)

    # Multi signal sources + bagua planes expand like create_experiment.
    n_src = 1
    try:
        sv = payload.get("signal_variants") or []
        if isinstance(sv, (list, tuple)) and len(sv) > 0:
            n_src = len(sv)
        elif bool(payload.get("dual_source_compare")):
            n_src = 2
    except Exception:
        n_src = 1

    n_plane = 1
    try:
        planes = payload.get("bagua_price_planes") or []
        if isinstance(planes, (list, tuple)) and len(planes) > 0:
            n_plane = len(planes)
        # only multiply when bagua filter is enabled
        gua_on = False
        for gf in list(payload.get("gua_filters") or []):
            if isinstance(gf, dict) and gf.get("enabled"):
                gua_on = True
                break
        if not gua_on:
            # single gua_filter object or with_bagua flag
            gf1 = payload.get("gua_filter")
            if isinstance(gf1, dict) and gf1.get("enabled"):
                gua_on = True
            elif bool(payload.get("with_bagua")):
                gua_on = True
        if not gua_on:
            n_plane = 1
    except Exception:
        n_plane = 1

    if n_src > 1:
        theoretical *= n_src
        actual *= n_src
    if n_plane > 1:
        theoretical *= n_plane
        actual *= n_plane

    if rejected and theoretical:
        soft = "理论 %s，过滤 %s，实际 %s" % (theoretical, rejected, actual)
        msg = ("%s；%s" % (msg, soft)) if msg else soft
    extra_notes = []
    if n_src > 1:
        extra_notes.append("信号源 ×%s" % n_src)
    if n_plane > 1:
        extra_notes.append("周卦口径 ×%s" % n_plane)
    if extra_notes:
        note = "、".join(extra_notes)
        msg = ("%s；%s" % (msg, note)) if msg else note

    # re-check caps after multiplies
    ok = True
    if actual > hard:
        ok = False
        msg = msg or ("组合数 %s 超过硬顶 %s" % (actual, hard))
    elif actual > max_v and not force:
        ok = False
        msg = msg or (
            "组合数 %s 超过上限 max_variants=%s；请缩小空间或提高 max_variants，或 force=true（硬顶 %s）"
            % (actual, max_v, hard)
        )

    return {
        "theoretical": theoretical,
        "rejected": rejected,
        "actual": actual,
        "rejection_reasons": reasons,
        "preview": preview,
        "max_variants": max_v,
        "hard_max": hard,
        "ok": ok,
        "message": msg,
        "estimated_variants": actual,
        "n": actual,
        "count": actual,
        "exceeds_soft_cap": actual > max_v,
        "warning": msg if actual > max_v or (actual > 20) else None,
        "force": force,
        "mode": "free_axes" if _has_free_axes(payload) else "legacy_templates",
        "signal_source_count": n_src,
        "bagua_plane_count": n_plane,
    }


def estimate_grid_size(
    rule_ids: Sequence[str],
    gua_keys: Sequence[str],
    weekday_keys: Sequence[str],
    stop_loss_list: Optional[Sequence[Optional[float]]] = None,
    *,
    take_profit_list: Optional[Sequence[Optional[float]]] = None,
    signal_weekdays_options: Optional[Sequence[Any]] = None,
    buy_options: Optional[Sequence[dict]] = None,
    sell_options: Optional[Sequence[dict]] = None,
) -> int:
    """Legacy-compatible count; free axes use product of independent options.

    Does not apply constraint filtering (use estimate_grid_from_payload).
    Caps: DEFAULT_MAX_VARIANTS=50 soft; HARD_MAX_VARIANTS=500 hard.
    """
    axes = {
        "signal_weekdays_options": signal_weekdays_options,
        "buy_options": buy_options,
        "sell_options": sell_options,
        "take_profit_list": take_profit_list,
    }
    if _has_free_axes(axes):
        n_rules = max(1, len(list(rule_ids or [])))
        n_gua = max(1, len(list(gua_keys or ["none"])))
        n_sig = max(1, len(list(signal_weekdays_options if signal_weekdays_options is not None else [None])))
        n_buy = max(1, len(list(buy_options if buy_options is not None else [{}])))
        n_sell = max(1, len(list(sell_options if sell_options is not None else [{}])))
        n_sl = max(1, len(list(stop_loss_list if stop_loss_list is not None else [None])))
        n_tp = max(1, len(list(take_profit_list if take_profit_list is not None else [None])))
        return n_rules * n_gua * n_sig * n_buy * n_sell * n_sl * n_tp
    n_rules = max(1, len(list(rule_ids or [])))
    n_gua = max(1, len(list(gua_keys or ["none"])))
    n_wd = max(1, len(list(weekday_keys or ["all_signal_tn12"])))
    n_sl = max(1, len(list(stop_loss_list if stop_loss_list is not None else [None])))
    n_tp = max(1, len(list(take_profit_list if take_profit_list is not None else [None])))
    return n_rules * n_gua * n_wd * n_sl * n_tp


def expand_param_grid(
    *,
    rule_ids: Sequence[str],
    gua_keys: Optional[Sequence[Any]] = None,
    gua_filters: Optional[Sequence[Any]] = None,
    weekday_keys: Sequence[str] = ("all_signal_tn12",),
    stop_loss_list: Optional[Sequence[Optional[float]]] = None,
    period: str = "DAY",
    periods: Optional[Sequence[str]] = None,
    codes: Optional[Sequence[str]] = None,
    start: Optional[int] = None,
    end: Optional[int] = None,
    account_mode: str = "portfolio",
    research_unadjusted: bool = False,
    signal_weekdays_options: Optional[Sequence[Any]] = None,
    buy_options: Optional[Sequence[dict]] = None,
    sell_options: Optional[Sequence[dict]] = None,
    take_profit_list: Optional[Sequence[Optional[float]]] = None,
    holiday_policy: str = "next_trading_day",
) -> List[Dict[str, Any]]:
    """Cartesian product of research axes -> list of BacktestRequest-like dicts.

    Free axes take precedence over weekday_keys templates when provided.
    """
    plan = expand_param_grid_unified(
        rule_ids=rule_ids,
        gua_keys=gua_keys,
        gua_filters=gua_filters,
        weekday_keys=weekday_keys,
        stop_loss_list=stop_loss_list,
        signal_weekdays_options=signal_weekdays_options,
        buy_options=buy_options,
        sell_options=sell_options,
        take_profit_list=take_profit_list,
        holiday_policy=holiday_policy,
        period=period,
        periods=periods,
        codes=codes,
        start=start,
        end=end,
        account_mode=account_mode,
        research_unadjusted=research_unadjusted,
        collect_rejections=False,
    )
    return list(plan["variants"])




# Formal dual-source comparison template (Gate C D2). This is a TEMPLATE,
# not a hardcoded special case: the create path treats it exactly like a
# user-supplied signal_variants list, and explicit signal_variants always
# override it. Legacy is never auto-added as a fallback variant.
DUAL_SOURCE_COMPARE_TEMPLATE: List[Dict[str, Any]] = [
    {"signal_data_source": "tdxquant", "signal_adjustment": "front"},
    {
        "signal_data_source": "internal",
        "signal_adjustment": "tushare_factor_qfq",
    },
]

_REPO_SIGNAL_SOURCES = ("tdxquant", "tushare", "internal", "raw")


def _signal_resolve_candidates(src: str, adj: Optional[str]) -> List[tuple]:
    """Ordered (source, adjustment) pairs for product signal sources."""
    if src == "raw":
        return [
            ("local_vendor", "none"),
            ("tushare", "none"),
            ("tdxquant", "none"),
            ("tdx_local", "none"),
            ("internal", "composite_none"),
        ]
    if src == "tushare":
        return [
            ("tushare", "qfq"),
            ("internal", "tushare_factor_qfq"),
            ("internal", "composite_tushare_factor_qfq"),
        ]
    if src == "tdxquant":
        return [("tdxquant", adj or "front")]
    if src == "internal":
        return [
            ("internal", adj or "tushare_factor_qfq"),
            ("internal", "composite_tushare_factor_qfq"),
        ]
    return [(src, adj or "")]


def _normalize_signal_variants(
    signal_variants: Optional[Sequence[dict]],
) -> List[Dict[str, Any]]:
    """Validate the user-supplied signal_variants list (Gate C D2)."""
    if not signal_variants:
        return []
    out: List[Dict[str, Any]] = []
    seen = set()
    for i, sv in enumerate(signal_variants):
        if not isinstance(sv, dict):
            raise ValueError(f"signal_variants[{i}] must be an object")
        src = str(sv.get("signal_data_source") or "").strip()
        adj = str(sv.get("signal_adjustment") or "").strip() or None
        ds = sv.get("dataset_id") or sv.get("signal_dataset_id") or None
        if src not in _REPO_SIGNAL_SOURCES:
            raise ValueError(
                f"signal_variants[{i}].signal_data_source must be one of "
                f"{_REPO_SIGNAL_SOURCES}, got {src!r} (legacy is not allowed "
                f"as an experiment variant)"
            )
        if not adj:
            if src == "raw":
                adj = "none"
            elif src == "tushare":
                adj = "qfq"
            elif src == "internal":
                adj = "tushare_factor_qfq"
            elif src == "tdxquant":
                adj = "front"
            else:
                from ..data.providers.base import DataSource, SIGNAL_SOURCE_ADJUSTMENT

                _enum = {
                    "tdxquant": DataSource.TDXQUANT,
                    "tushare": DataSource.TUSHARE,
                }.get(src)
                if _enum and _enum in SIGNAL_SOURCE_ADJUSTMENT:
                    adj = SIGNAL_SOURCE_ADJUSTMENT[_enum].value
        if not adj:
            raise ValueError(
                f"signal_variants[{i}].signal_adjustment is required for source {src}"
            )
        key = (src, adj, ds or "")
        if key in seen:
            raise ValueError(f"signal_variants[{i}] duplicates {key}")
        seen.add(key)
        out.append(
            {"signal_data_source": src, "signal_adjustment": adj, "dataset_id": ds}
        )
    if len(out) > 4:
        raise ValueError("signal_variants supports at most 4 signal sources")
    return out


def _resolve_variant_datasets_and_common_universe(
    cfg: AStockConfig,
    descriptors: List[Dict[str, Any]],
    *,
    requested_codes: Sequence[str],
    execution_data_source: str,
    execution_dataset_id: Optional[str],
    requested_end: Optional[int],
    universe_dataset_id: Optional[str] = None,
):
    """Resolve + bind every signal variant dataset and the shared execution
    dataset; compute the dynamic common universe and common cutoff (Gate C
    D2 §八). Nothing here is hardcoded to specific counts or dates."""
    from ..data.dataset_store import DatasetStore
    from ..data.repository import MarketDataRepository, DatasetNotFoundError
    from ..data.dataset_binding import (
        DatasetBindingError,
        classify_symbol_coverage,
        manifest_symbol_index,
        validate_execution_dataset_binding,
        validate_signal_dataset_binding,
    )

    repo = MarketDataRepository(DatasetStore(cfg.market_data_root))

    resolved: List[Dict[str, Any]] = []
    for desc in descriptors:
        src = desc["signal_data_source"]
        adj = desc.get("signal_adjustment")
        ds_id = desc.get("dataset_id")
        if ds_id:
            manifest = validate_signal_dataset_binding(
                repo,
                ds_id,
                source=src if src != "raw" else None,
                adjustment=adj,
                period="1d",
                allow_raw_signal=(src == "raw") or (adj in ("none", "composite_none")),
            )
        else:
            manifest = None
            _last = None
            for _cs, _ca in _signal_resolve_candidates(src, adj):
                try:
                    manifest = repo.resolve_latest_ready(
                        source=_cs, adjustment=_ca, period="1d"
                    )
                    break
                except DatasetNotFoundError as _e:
                    _last = _e
            if manifest is None:
                raise DatasetBindingError(
                    "DATASET_NOT_FOUND",
                    f"No ready dataset for signal variant source={src} "
                    f"adjustment={adj}. Run sync first.",
                    http_status=404,
                    requested_source=src,
                    requested_adjustment=adj,
                    remediation="先同步该信号源的 ready 数据集，或从 variant 列表移除",
                ) from _last
            ds_id = manifest.dataset_id
        label = f"{src}/{manifest.adjustment}"
        resolved.append(
            {
                "signal_data_source": src,
                "signal_adjustment": manifest.adjustment,
                "dataset_id": ds_id,
                "manifest": manifest,
                "label": label,
            }
        )
    if len({r["dataset_id"] for r in resolved}) != len(resolved):
        raise ValueError("signal variants must use distinct signal datasets")

    exec_src = (execution_data_source or "local_vendor").strip()
    if execution_dataset_id:
        exec_manifest = validate_execution_dataset_binding(
            repo, execution_dataset_id, source=exec_src, period="1d"
        )
        exec_id = execution_dataset_id
    else:
        try:
            exec_manifest = repo.resolve_latest_ready(
                source=exec_src, adjustment="none", period="1d"
            )
            exec_id = exec_manifest.dataset_id
        except DatasetNotFoundError:
            raise DatasetBindingError(
                "DATASET_NOT_FOUND",
                f"No ready {exec_src}/none execution dataset for experiment.",
                http_status=404,
                requested_source=exec_src,
                requested_adjustment="none",
                remediation="先同步执行数据集（如 local_vendor/none），或显式指定 execution_dataset_id",
            ) from None

    # ---- dynamic common universe (requested ∩ every signal set ∩ execution) ----
    sig_indexes = [(r["label"], manifest_symbol_index(r["manifest"])) for r in resolved]
    exec_index = manifest_symbol_index(exec_manifest)
    # Gate C D7: corporate-action factor coverage is part of eligibility
    # (per-symbol explicit exclusion — never whole-board silent exclusion).
    # Gate B8: coverage mirrors factor_resolution_v1 — a symbol is covered if
    # the main factor set, the supplement factor set (delisted stocks), or
    # the PIT-universe alias canonical in either of them has it.
    _factor_cache: Dict[str, Any] = {}
    _pit_for_alias = None
    if universe_dataset_id:
        from ..data.pit_universe import PointInTimeUniverse

        try:
            _pit_for_alias = PointInTimeUniverse.from_root(
                cfg.market_data_root, universe_dataset_id
            )
        except FileNotFoundError:
            raise DatasetBindingError(
                "UNIVERSE_NOT_FOUND",
                f"point-in-time universe not found: {universe_dataset_id}",
                http_status=404,
                dataset_id=universe_dataset_id,
                remediation="检查 universe_dataset_id 拼写，或先构建点时宇宙",
            ) from None
        except ValueError as _ue:
            raise DatasetBindingError(
                "UNIVERSE_CORRUPT",
                f"point-in-time universe unreadable: {universe_dataset_id}: {_ue}",
                http_status=422,
                dataset_id=universe_dataset_id,
                remediation="universe 文件内容哈希不一致（损坏）；禁止继续，请检查数据根",
            ) from None

    def _index_for_id(fid: str) -> Optional[Dict[str, Any]]:
        if not fid:
            return None
        if fid not in _factor_cache:
            try:
                fm = repo.get_dataset(fid)
            except DatasetNotFoundError:
                _factor_cache[fid] = None
                return None
            _factor_cache[fid] = manifest_symbol_index(fm)
        return _factor_cache[fid]

    def _factor_indexes_for(r) -> List[Dict[str, Any]]:
        """Ordered coverage tiers: main factor set, then supplement.

        CA-factor eligibility is only required for **derived** QFQ signal
        sets that pin a factor parent (or internal/*qfq). Native vendor
        signals (tdxquant/front, tushare/qfq, raw/none) do not gate the
        common pool on adj_factor coverage — otherwise a tiny incomplete
        factor snapshot can empty the entire experiment pool.
        """
        m = r["manifest"]
        fid = getattr(m, "factor_dataset_id", "") or ""
        adj = (getattr(m, "adjustment", None) or "").strip()
        src = (getattr(m, "source", None) or "").strip()
        derived = bool(fid) or (
            src == "internal"
            and adj in ("tushare_factor_qfq", "composite_tushare_factor_qfq", "asof_qfq")
        )
        if not derived:
            return []
        if not fid:
            # Prefer fullest ready adj_factor (symbol_count), then cutoff/created.
            cands = repo.list_datasets(
                source="tushare", adjustment="adj_factor", period="1d", status="ready",
            )
            if cands:
                cands.sort(
                    key=lambda x: (
                        int(x.data_cutoff_date or 0),
                        int(x.symbol_count or 0),
                        x.created_at or "",
                    ),
                    reverse=True,
                )
                fid = cands[0].dataset_id
            else:
                fid = ""
        _prov = getattr(m, "provenance", None) or {}
        sup_fid = str(_prov.get("supplement_factor_dataset_id") or "")
        return [i for i in (_index_for_id(fid), _index_for_id(sup_fid))
                if i is not None]

    def _ca_factor_covered(fidxs: List[Dict[str, Any]], code: str) -> Optional[str]:
        """None when covered; else the coverage failure class of tier 1."""
        if not fidxs:
            return None  # no factor dataset at all -> legacy behavior
        lookups = [code]
        if _pit_for_alias is not None:
            _w = _pit_for_alias.resolve(code)
            _canon = getattr(_w, "canonical_symbol", None) if _w else None
            if _canon and _canon != code:
                lookups.append(_canon)
        first_fail = None
        for fidx in fidxs:
            for lk in lookups:
                fcov = classify_symbol_coverage(fidx, lk)
                if fcov == "ok":
                    return None
                if first_fail is None:
                    first_fail = fcov
        return first_fail or "missing"

    factor_indexes = [(r["label"], _factor_indexes_for(r)) for r in resolved]
    eligible: List[str] = []
    exclusions: List[Dict[str, Any]] = []
    excluded_by_signal: Dict[str, int] = {r["label"]: 0 for r in resolved}
    excluded_by_execution = 0
    for code in requested_codes:
        reasons = []
        for (label, idx), (_, fidxs) in zip(sig_indexes, factor_indexes):
            cov = classify_symbol_coverage(idx, code)
            if cov != "ok":
                reasons.append(f"signal[{label}]:{cov}")
                excluded_by_signal[label] += 1
                continue
            _ffail = _ca_factor_covered(fidxs, code)
            if _ffail is not None:
                reasons.append(f"signal[{label}]:ca_factor_{_ffail}")
                excluded_by_signal[label] += 1
        cov_e = classify_symbol_coverage(exec_index, code)
        if cov_e != "ok":
            reasons.append(f"execution:{cov_e}")
            excluded_by_execution += 1
        if reasons:
            exclusions.append({"symbol": code, "reason": ";".join(reasons)})
        else:
            eligible.append(code)

    # ---- dynamic common cutoff ----
    def _max_last_date(m) -> Optional[int]:
        lasts = [s.last_date for s in m.symbols if s.last_date]
        return max(lasts) if lasts else None

    cutoffs = [c for c in (
        [_max_last_date(r["manifest"]) for r in resolved]
        + [_max_last_date(exec_manifest)]
    ) if c]
    dataset_cutoff = min(cutoffs) if cutoffs else None
    effective_end = dataset_cutoff
    if requested_end and dataset_cutoff:
        effective_end = min(int(requested_end), int(dataset_cutoff))
    elif requested_end:
        effective_end = int(requested_end)

    info = {
        "signal_variants": [
            {
                "signal_data_source": r["signal_data_source"],
                "signal_adjustment": r["signal_adjustment"],
                "dataset_id": r["dataset_id"],
                "label": r["label"],
                "signal_raw_dataset_id": getattr(r["manifest"], "raw_dataset_id", "") or "",
                "signal_factor_dataset_id": getattr(r["manifest"], "factor_dataset_id", "") or "",
                "formula_version": getattr(r["manifest"], "formula_version", "") or "",
                "data_cutoff": _max_last_date(r["manifest"]),
            }
            for r in resolved
        ],
        "execution": {
            "source": exec_manifest.source,
            "adjustment": exec_manifest.adjustment,
            "dataset_id": exec_id,
            "data_cutoff": _max_last_date(exec_manifest),
        },
        "requested_universe_count": len(requested_codes),
        "common_universe_count": len(eligible),
        "excluded_by_signal_counts": excluded_by_signal,
        "excluded_by_execution_count": excluded_by_execution,
        "excluded_total": len(exclusions),
        "exclusions": exclusions[:500],
        "requested_end_date": requested_end,
        "dataset_common_cutoff": dataset_cutoff,
        "effective_end_date": effective_end,
        "eligible_codes": eligible,
    }
    exec_resolved = {"source": exec_manifest.source, "dataset_id": exec_id}
    return resolved, exec_resolved, info


def create_experiment_from_grid(
    cfg: Optional[AStockConfig],
    *,
    name: str,
    rule_ids: Sequence[str],
    gua_keys: Optional[Sequence[Any]] = None,
    gua_filters: Optional[Sequence[Any]] = None,
    weekday_keys: Optional[Sequence[str]] = None,
    stop_loss_list: Optional[Sequence[Optional[float]]] = None,
    period: str = "DAY",
    periods: Optional[Sequence[str]] = None,
    codes: Optional[Sequence[str]] = None,
    universe: Optional[str] = None,
    start: Optional[int] = None,
    end: Optional[int] = None,
    account_mode: str = "portfolio",
    research_unadjusted: bool = False,
    max_variants: int = DEFAULT_MAX_VARIANTS,
    concurrency: int = 1,
    force: bool = False,
    note: str = "",
    signal_weekdays_options: Optional[Sequence[Any]] = None,
    buy_options: Optional[Sequence[dict]] = None,
    sell_options: Optional[Sequence[dict]] = None,
    take_profit_list: Optional[Sequence[Optional[float]]] = None,
    holiday_policy: str = "next_trading_day",
    engine: str = "fast",
    artifact_level: str = "summary",
    use_signal_cache: bool = True,
    promote_top_n: int = 3,
    promote_metric: str = "total_return",
    signal_data_source: Optional[str] = None,
    signal_adjustment: Optional[str] = None,
    dataset_id: Optional[str] = None,
    weekly_bar_mode: str = "local_aggregate",
    execution_data_source: str = "local_vendor",
    execution_dataset_id: Optional[str] = None,
    dual_source_compare: bool = False,
    signal_variants: Optional[Sequence[dict]] = None,
    bagua_period: str = "WEEK",
    bagua_price_plane: Optional[str] = None,
    bagua_price_planes: Optional[Sequence[str]] = None,
    baseline: Optional[str] = None,
    universe_dataset_id: Optional[str] = None,
    delist_exit_scenario: Optional[str] = None,
    delist_recovery_discount: Optional[float] = None,
) -> Dict[str, Any]:
    """Create experiment from legacy weekday_keys templates OR free axes.

    Phase-3 defaults: engine=fast, artifact_level=summary, use_signal_cache=True
    so large grids screen quickly; promote winners with engine=full later.

    Templates are presets only: when any free axis is provided, weekday_keys
    are ignored.     Caps: DEFAULT_MAX_VARIANTS=50 soft (UI often raises); HARD_MAX_VARIANTS
    default 2000 (env ASTOCK_EXP_HARD_MAX_VARIANTS). force=True required to
    exceed max_variants; hard max always refuses.

    Extra (爻辞回测):
      - gua_filters: inline exact_line filters (or mix dicts into gua_keys)
      - periods: e.g. [\"DAY\", \"WEEK\"] expands period axis
      - universe: demo | full | custom (with codes)
    """
    cfg = cfg or get_default_config()
    from .yao_rules import resolve_universe_codes, normalize_periods

    # Gate B7: survivorship-safe baseline selector. Explicit user fields win;
    # the baseline fills whatever is unset. Fail-closed: missing baseline
    # datasets raise BaselineUnavailableError (mapped to 4xx) — never a
    # silent fallback to legacy datasets.
    if baseline:
        if baseline != "survivorship_safe":
            raise ValueError(
                f"unknown baseline: {baseline!r}; expected 'survivorship_safe' "
                f"(omit for legacy explicit dataset selection)"
            )
        from .baseline import resolve_survivorship_safe_baseline

        _bl = resolve_survivorship_safe_baseline(
            cfg,
            delist_exit_scenario=delist_exit_scenario,
            delist_recovery_discount=delist_recovery_discount,
        )
        signal_data_source = signal_data_source or _bl["signal_data_source"]
        signal_adjustment = signal_adjustment or _bl["signal_adjustment"]
        dataset_id = dataset_id or _bl["dataset_id"]
        # Unset / product-default exec → baseline pins; explicit source+id kept.
        if not execution_dataset_id and execution_data_source in (
            None, "", "tdx_local", "local_vendor"
        ):
            execution_data_source = _bl["execution_data_source"]
        execution_dataset_id = execution_dataset_id or _bl["execution_dataset_id"]
        universe_dataset_id = universe_dataset_id or _bl["universe_dataset_id"]
        delist_exit_scenario = delist_exit_scenario or _bl["delist_exit_scenario"]

    resolved_codes = resolve_universe_codes(
        cfg, universe=universe, codes=codes
    )
    period_list = normalize_periods(period, periods)
    period0 = period_list[0] if period_list else (period or "DAY")

    axes = {
        "signal_weekdays_options": signal_weekdays_options,
        "buy_options": buy_options,
        "sell_options": sell_options,
        "take_profit_list": take_profit_list,
    }
    use_free = _has_free_axes(axes)
    wd_keys = list(weekday_keys) if weekday_keys is not None else ["all_signal_tn12"]
    gkeys = list(gua_keys) if gua_keys is not None else ["none"]
    gfilters = list(gua_filters) if gua_filters is not None else None

    payload = {
        "rule_ids": list(rule_ids),
        "gua_keys": gkeys,
        "gua_filters": gfilters,
        "weekday_keys": None if use_free else wd_keys,
        "stop_loss_list": stop_loss_list,
        "signal_weekdays_options": signal_weekdays_options,
        "buy_options": buy_options,
        "sell_options": sell_options,
        "take_profit_list": take_profit_list,
        "holiday_policy": holiday_policy,
        "period": period0,
        "periods": period_list,
        "codes": resolved_codes,
        "universe": universe or "demo",
        "start": start,
        "end": end,
        "account_mode": account_mode,
        "research_unadjusted": research_unadjusted,
        "max_variants": max_variants,
        "force": force,
        "hard_max": HARD_MAX_VARIANTS,
    }

    research_plan = _try_research_plan(payload)
    if research_plan and research_plan.get("source") == "research.planner":
        variants = list(research_plan.get("variants") or [])
        n = int(research_plan.get("actual") or len(variants))
        theoretical = research_plan.get("theoretical")
        rejected = research_plan.get("rejected")
        rejection_reasons = research_plan.get("rejection_reasons")
        # When planner truncates (error set) without force, refuse create
        if research_plan.get("message") and not variants and n > max_variants and not force:
            raise ValueError(
                "组合数 %s 超过上限 max_variants=%s；请缩小空间或提高 max_variants，或 force=true（硬顶 %s）"
                % (n, max_variants, HARD_MAX_VARIANTS)
            )
        if n > max_variants and not force:
            raise ValueError(
                "组合数 %s 超过上限 max_variants=%s；请缩小空间或提高 max_variants，或 force=true（硬顶 %s）"
                % (n, max_variants, HARD_MAX_VARIANTS)
            )
        if n > HARD_MAX_VARIANTS:
            raise ValueError("组合数 %s 超过硬顶 %s" % (n, HARD_MAX_VARIANTS))
        # If truncated due to cap, variants may be empty — expand via force path already handled
        if not variants and n > 0:
            # re-plan with force to materialize for storage only if user forced
            if force:
                payload["force"] = True
                research_plan = _try_research_plan(payload) or research_plan
                variants = list(research_plan.get("variants") or [])
                n = int(research_plan.get("actual") or len(variants))
            else:
                raise ValueError(
                    "组合数 %s 超过上限 max_variants=%s；请缩小空间或 force=true（硬顶 %s）"
                    % (n, max_variants, HARD_MAX_VARIANTS)
                )
    else:
        plan = expand_param_grid_unified(
            rule_ids=rule_ids,
            gua_keys=gkeys,
            gua_filters=gfilters,
            weekday_keys=None if use_free else wd_keys,
            stop_loss_list=stop_loss_list,
            signal_weekdays_options=signal_weekdays_options,
            buy_options=buy_options,
            sell_options=sell_options,
            take_profit_list=take_profit_list,
            holiday_policy=holiday_policy,
            period=period0,
            periods=period_list,
            codes=resolved_codes,
            start=start,
            end=end,
            account_mode=account_mode,
            research_unadjusted=research_unadjusted,
            collect_rejections=True,
        )
        variants = list(plan["variants"])
        n = int(plan["actual"])
        theoretical = plan.get("theoretical")
        rejected = plan.get("rejected")
        rejection_reasons = plan.get("rejection_reasons")
        if n > max_variants and not force:
            raise ValueError(
                "组合数 %s 超过上限 max_variants=%s；请缩小空间或提高 max_variants，或 force=true（硬顶 %s）"
                % (n, max_variants, HARD_MAX_VARIANTS)
            )
        if n > HARD_MAX_VARIANTS:
            raise ValueError("组合数 %s 超过硬顶 %s" % (n, HARD_MAX_VARIANTS))

    _ms_fields = {
        "signal_data_source": signal_data_source,
        "signal_adjustment": signal_adjustment,
        "dataset_id": dataset_id,
        "weekly_bar_mode": weekly_bar_mode,
        "execution_data_source": execution_data_source,
        "execution_dataset_id": execution_dataset_id,
        # Gate B7: survivorship-safe chain fields on every variant
        "universe_dataset_id": universe_dataset_id,
        "delist_exit_scenario": delist_exit_scenario,
        "delist_recovery_discount": delist_recovery_discount,
    }
    for v in variants:
        v.update({k: val for k, val in _ms_fields.items() if val is not None})

    # ---- Gate C D2: configurable signal-variant list (no hardcoded pairs) ----
    # dual_source_compare is now just an alias for the formal template below;
    # explicit signal_variants always wins. Legacy is NEVER auto-added.
    _descriptors = _normalize_signal_variants(signal_variants)
    if not _descriptors and dual_source_compare:
        _descriptors = [dict(d) for d in DUAL_SOURCE_COMPARE_TEMPLATE]
    _repo_mode_single = (
        not _descriptors
        and signal_data_source in ("tdxquant", "tushare", "internal", "raw")
    )
    if _repo_mode_single:
        _descriptors = [
            {
                "signal_data_source": signal_data_source,
                "signal_adjustment": signal_adjustment,
                "dataset_id": dataset_id,
            }
        ]
    common_universe_info = None
    if _descriptors:
        (
            _resolved_descriptors,
            _exec_resolved,
            common_universe_info,
        ) = _resolve_variant_datasets_and_common_universe(
            cfg,
            _descriptors,
            requested_codes=resolved_codes,
            execution_data_source=execution_data_source,
            execution_dataset_id=execution_dataset_id,
            requested_end=end,
            universe_dataset_id=universe_dataset_id,
        )
        eligible_codes = common_universe_info["eligible_codes"]
        if not eligible_codes:
            raise ValueError(
                "共同股票池为空：请求的股票均不被所选信号/执行数据集覆盖"
            )
        effective_end = common_universe_info["effective_end_date"]
        multi = len(_resolved_descriptors) > 1
        new_variants = []
        for di, desc in enumerate(_resolved_descriptors):
            for v in variants:
                dv = dict(v)
                dv["signal_data_source"] = desc["signal_data_source"]
                dv["signal_adjustment"] = desc["signal_adjustment"]
                dv["dataset_id"] = desc["dataset_id"]
                dv["execution_data_source"] = _exec_resolved["source"]
                dv["execution_dataset_id"] = _exec_resolved["dataset_id"]
                # raw signal plane = research_unadjusted for that variant only
                if desc["signal_data_source"] == "raw" or (
                    desc.get("signal_adjustment") in ("none", "composite_none")
                ):
                    dv["research_unadjusted"] = True
                dv["codes"] = list(eligible_codes)
                if effective_end:
                    dv["end"] = effective_end
                dv["_meta"] = dict(dv.get("_meta") or {})
                dv["_meta"]["signal_data_source"] = desc["signal_data_source"]
                dv["_meta"]["signal_dataset_id"] = desc["dataset_id"]
                dv["_meta"]["signal_variant_label"] = desc["label"]
                new_variants.append(dv)
        variants = new_variants
        n = len(variants)
        if multi and n > HARD_MAX_VARIANTS:
            raise ValueError("组合数 %s 超过硬顶 %s" % (n, HARD_MAX_VARIANTS))
        # experiment-level fields now describe the shared execution config
        execution_data_source = _exec_resolved["source"]
        execution_dataset_id = _exec_resolved["dataset_id"]
        resolved_codes = list(eligible_codes)
        if effective_end:
            end = effective_end
        if _repo_mode_single:
            signal_data_source = _resolved_descriptors[0]["signal_data_source"]
            signal_adjustment = _resolved_descriptors[0]["signal_adjustment"]
            dataset_id = _resolved_descriptors[0]["dataset_id"]

    # ---- Weekly bagua price-plane axis (raw / tdx_front / tushare_qfq) ----
    # Multi-select expands only variants that enable bagua filter; others keep
    # a single default plane for repro without multiplying the grid.
    _planes: List[str] = []
    _seen_pl = set()
    for p in list(bagua_price_planes or []):
        s = str(p or "").strip().lower()
        if not s or s in _seen_pl:
            continue
        if s in ("tdxquant_front", "front", "通达信前复权"):
            s = "tdx_front"
        elif s in ("ts_qfq", "qfq", "tushare前复权"):
            s = "tushare_qfq"
        elif s in ("none", "unadjusted", "未复权"):
            s = "raw"
        if s not in ("raw", "tdx_front", "tushare_qfq"):
            continue
        _seen_pl.add(s)
        _planes.append(s)
    if not _planes:
        s0 = str(bagua_price_plane or "raw").strip().lower() or "raw"
        if s0 in ("tdxquant_front", "front"):
            s0 = "tdx_front"
        elif s0 in ("ts_qfq", "qfq"):
            s0 = "tushare_qfq"
        elif s0 in ("none", "unadjusted"):
            s0 = "raw"
        if s0 not in ("raw", "tdx_front", "tushare_qfq"):
            s0 = "raw"
        _planes = [s0]
    _bp = (bagua_period or "WEEK").strip().upper() or "WEEK"
    if _bp not in ("WEEK", "DAY", "MONTH"):
        _bp = "WEEK"
    _plane_expanded: List[Dict[str, Any]] = []
    for v in variants:
        _gf = v.get("gua_filter") if isinstance(v.get("gua_filter"), dict) else {}
        _gua_on = bool(v.get("with_bagua")) or bool(_gf.get("enabled"))
        _use_planes = _planes if _gua_on else [_planes[0]]
        for plane in _use_planes:
            dv = dict(v)
            dv["bagua_period"] = _bp
            dv["bagua_price_plane"] = plane
            dv["_meta"] = dict(dv.get("_meta") or {})
            dv["_meta"]["bagua_period"] = _bp
            dv["_meta"]["bagua_price_plane"] = plane
            _plane_expanded.append(dv)
    variants = _plane_expanded
    n = len(variants)
    if n > HARD_MAX_VARIANTS:
        raise ValueError("组合数 %s 超过硬顶 %s（含周卦口径展开）" % (n, HARD_MAX_VARIANTS))
    if n > max_variants and not force:
        raise ValueError(
            "组合数 %s 超过上限 max_variants=%s；请少选周卦口径/信号源，或 force=true（硬顶 %s）"
            % (n, max_variants, HARD_MAX_VARIANTS)
        )

    warning = None
    if n > 20:
        warning = "组合数 %s 较大，建议先用演示池试跑" % n
    uni_label = (universe or "demo").strip().lower() or "demo"
    if uni_label in ("full", "all", "market") and len(resolved_codes) > 100:
        warning = ((warning + "；") if warning else "") + (
            "全市场 %s 只，单组耗时可能很长" % len(resolved_codes)
        )

    # Serialize gua options for config (dicts kept; large lists ok)
    def _ser_gua(items: Optional[Sequence[Any]]) -> List[Any]:
        out: List[Any] = []
        for x in list(items or []):
            if isinstance(x, dict):
                out.append(dict(x))
            else:
                out.append(x)
        return out

    config = {
        "rule_ids": list(rule_ids),
        "gua_keys": _ser_gua(gkeys),
        "gua_filters": _ser_gua(gfilters) if gfilters else [],
        "weekday_keys": [] if use_free else wd_keys,
        "weekday_keys_note": "templates are presets only; ignored when free axes provided",
        "stop_loss_list": list(stop_loss_list) if stop_loss_list is not None else [None],
        "take_profit_list": list(take_profit_list) if take_profit_list is not None else [None],
        "signal_weekdays_options": list(signal_weekdays_options)
        if signal_weekdays_options is not None
        else None,
        "buy_options": list(buy_options) if buy_options is not None else None,
        "sell_options": list(sell_options) if sell_options is not None else None,
        "holiday_policy": holiday_policy,
        "engine": (engine or "fast").strip().lower(),
        "artifact_level": (artifact_level or "summary").strip().lower(),
        "use_signal_cache": bool(use_signal_cache),
        "promote_top_n": int(promote_top_n or 0),
        "promote_metric": str(promote_metric or "total_return"),
        "mode": "free_axes" if use_free else "legacy_templates",
        "period": period0,
        "periods": list(period_list),
        "universe": uni_label,
        "codes": list(resolved_codes),
        "n_codes": len(resolved_codes),
        "start": start,
        "end": end,
        "account_mode": account_mode,
        "research_unadjusted": research_unadjusted,
        "signal_data_source": signal_data_source,
        "signal_adjustment": signal_adjustment,
        "dataset_id": dataset_id,
        "weekly_bar_mode": weekly_bar_mode,
        "execution_data_source": execution_data_source,
        "execution_dataset_id": execution_dataset_id,
        # Gate B7: survivorship-safe chain configuration
        "baseline": baseline,
        "universe_dataset_id": universe_dataset_id,
        "delist_exit_scenario": delist_exit_scenario,
        "delist_recovery_discount": delist_recovery_discount,
        "dual_source_compare": dual_source_compare,
        "signal_variants": (
            [
                {k: v for k, v in d.items() if k != "manifest"}
                for d in (common_universe_info or {}).get("signal_variants", [])
            ]
            if common_universe_info
            else None
        ),
        "bagua_period": _bp,
        "bagua_price_plane": _planes[0],
        "bagua_price_planes": list(_planes),
        "common_universe": (
            {k: v for k, v in common_universe_info.items() if k != "eligible_codes"}
            if common_universe_info
            else None
        ),
        "estimated": n,
        "theoretical": theoretical,
        "rejected": rejected,
        "rejection_reasons": rejection_reasons,
        "warning": warning,
    }
    note_parts = [note or ""]
    if use_free:
        note_parts.append("[free_axes; weekday templates ignored]")
    exp = exp_db.create_experiment(
        cfg,
        name=name or ("实验·%s组合" % n),
        config=config,
        variants=variants,
        max_variants=max_variants,
        concurrency=concurrency,
        note=" ".join(x for x in note_parts if x).strip(),
    )
    if warning:
        exp["warning"] = warning
    exp["theoretical"] = theoretical
    exp["rejected"] = rejected
    exp["actual"] = n
    return exp


class ExperimentRunner:
    """In-process limited-concurrency runner for experiment variants."""

    def __init__(self, cfg: Optional[AStockConfig] = None):
        self.cfg = cfg or get_default_config()
        self._lock = threading.Lock()
        self._cancel: Dict[str, bool] = {}
        self._threads: Dict[str, threading.Thread] = {}

    def cancel(self, experiment_id: str) -> None:
        with self._lock:
            self._cancel[experiment_id] = True

    def is_cancelled(self, experiment_id: str) -> bool:
        with self._lock:
            return bool(self._cancel.get(experiment_id))

    def start(self, experiment_id: str) -> Dict[str, Any]:
        exp = exp_db.get_experiment(self.cfg, experiment_id)
        if exp.get("status") in ("running",):
            return exp
        with self._lock:
            self._cancel[experiment_id] = False
            if experiment_id in self._threads and self._threads[experiment_id].is_alive():
                return exp
            t = threading.Thread(
                target=self._run_experiment,
                args=(experiment_id,),
                daemon=True,
                name=f"exp-{experiment_id}",
            )
            self._threads[experiment_id] = t
            t.start()
        exp_db.update_experiment_status(self.cfg, experiment_id, "running")
        return exp_db.get_experiment(self.cfg, experiment_id)

    def _run_one(self, experiment_id: str, variant: dict) -> Tuple[str, str, Optional[str], Optional[str]]:
        """Returns (variant_id, status, run_id, error)."""
        vid = variant["variant_id"]
        if self.is_cancelled(experiment_id):
            return vid, "cancelled", None, "cancelled"
        params = dict(variant.get("params") or {})
        meta = params.pop("_meta", None)
        ph = variant.get("param_hash") or exp_db.param_hash(params)
        # research fingerprint is metadata only; param_hash dedup stays param-only
        research_fp = None
        try:
            from ..research.fingerprint import research_fingerprint_from_params
            research_fp = research_fingerprint_from_params(params).full_hex(16)
        except Exception:
            research_fp = None
        # de-dup: only skip if prior run has real activity (signals/trades).
        # Empty historical runs (0 trades / 0 signals) must be re-executed.
        existing = exp_db.find_run_id_by_param_hash(self.cfg, ph)
        if existing:
            reuse = True
            try:
                from .runs import load_run_summary
                prev = load_run_summary(self.cfg, existing) or {}
                pm = prev.get("metrics") if isinstance(prev.get("metrics"), dict) else {}
                if not pm and isinstance(prev.get("summary"), dict):
                    pm = (prev.get("summary") or {}).get("metrics") or {}
                n_tr = pm.get("n_trades")
                if n_tr is None:
                    n_tr = pm.get("n_round_trips")
                if n_tr is None:
                    n_tr = pm.get("n_buys")
                n_sig = pm.get("n_signals_fast")
                if n_sig is None:
                    n_sig = pm.get("n_events")
                if n_sig is None:
                    n_sig = pm.get("n_signals_after_bagua")
                # reuse only when we have evidence of a non-empty or intentional run
                empty = (n_tr is None or int(n_tr or 0) == 0) and (n_sig is None or int(n_sig or 0) == 0)
                # if metrics missing entirely, also re-run
                if not pm or empty:
                    reuse = False
            except Exception:
                reuse = True  # keep old skip behavior if summary unreadable
            if reuse:
                exp_db.update_variant(
                    self.cfg, vid, status="skipped", run_id=existing, error="param_hash dedup"
                )
                return vid, "skipped", existing, None
            # fall through to re-run with a fresh run_id

        exp_db.update_variant(self.cfg, vid, status="running")
        try:
            # Experiment-level defaults (Phase-3): fast screen + summary artifacts + cache
            exp_row = exp_db.get_experiment(self.cfg, experiment_id)
            exp_cfg = dict(exp_row.get("config") or {})
            eng = (
                params.get("engine")
                or exp_cfg.get("engine")
                or "fast"
            )
            art = (
                params.get("artifact_level")
                or exp_cfg.get("artifact_level")
                or "summary"
            )
            use_cache = params.get("use_signal_cache")
            if use_cache is None:
                use_cache = exp_cfg.get("use_signal_cache", True)
            req = BacktestRequest(
                rule_ids=list(params.get("rule_ids") or []),
                period=params.get("period") or "DAY",
                hold=int(params.get("hold") or 1),
                entry_lag=int(params.get("entry_lag") or 1),
                signal_weekdays=params.get("signal_weekdays"),
                buy_on=params.get("buy_on") or "open",
                sell_on=params.get("sell_on") or "open",
                buy_weekday=params.get("buy_weekday"),
                exit_weekday=params.get("exit_weekday"),
                codes=params.get("codes"),
                start=params.get("start"),
                end=params.get("end"),
                with_bagua=bool(params.get("with_bagua")),
                gua_filter=params.get("gua_filter"),
                bagua_period=params.get("bagua_period")
                or exp_cfg.get("bagua_period")
                or "WEEK",
                bagua_price_plane=params.get("bagua_price_plane")
                or exp_cfg.get("bagua_price_plane")
                or "raw",
                research_unadjusted=bool(params.get("research_unadjusted")),
                stop_loss=params.get("stop_loss"),
                take_profit=params.get("take_profit"),
                account_mode=params.get("account_mode") or "portfolio",
                engine=str(eng or "fast"),
                artifact_level=str(art or "summary"),
                use_signal_cache=bool(use_cache),
                holiday_policy=params.get("holiday_policy")
                or exp_cfg.get("holiday_policy")
                or "next_trading_day",
                signal_data_source=params.get("signal_data_source")
                or exp_cfg.get("signal_data_source"),
                signal_adjustment=params.get("signal_adjustment")
                or exp_cfg.get("signal_adjustment"),
                dataset_id=params.get("dataset_id")
                or exp_cfg.get("dataset_id"),
                weekly_bar_mode=params.get("weekly_bar_mode")
                or exp_cfg.get("weekly_bar_mode")
                or "local_aggregate",
                execution_data_source=params.get("execution_data_source")
                or exp_cfg.get("execution_data_source")
                or "local_vendor",
                execution_dataset_id=params.get("execution_dataset_id")
                or exp_cfg.get("execution_dataset_id"),
                universe_dataset_id=params.get("universe_dataset_id")
                or exp_cfg.get("universe_dataset_id"),
                delist_exit_scenario=params.get("delist_exit_scenario")
                or exp_cfg.get("delist_exit_scenario"),
                delist_recovery_discount=(
                    params.get("delist_recovery_discount")
                    if params.get("delist_recovery_discount") is not None
                    else exp_cfg.get("delist_recovery_discount")
                ),
            )
            svc = BacktestService(self.cfg)
            summary = svc.run(req)
            rid = summary.get("run_id")
            # link experiment on run row (Gate C D5: upsert failures are NOT
            # swallowed — the variant is marked failed by the outer handler)
            if rid:
                    exp_db.upsert_run_from_index_row(
                        self.cfg,
                        {
                            "run_id": rid,
                            "title": summary.get("title"),
                            "status": summary.get("status"),
                            "created_at": int(time.time()),
                            "indicator_ids": (summary.get("repro") or {}).get("indicator_ids")
                            or params.get("rule_ids"),
                            "period": params.get("period"),
                            "hold": params.get("hold"),
                            "entry_lag": params.get("entry_lag"),
                            "buy_weekday": params.get("buy_weekday"),
                            "exit_weekday": params.get("exit_weekday"),
                            "buy_on": params.get("buy_on"),
                            "sell_on": params.get("sell_on"),
                            "signal_weekdays": params.get("signal_weekdays"),
                            "account_mode": params.get("account_mode"),
                            "start": params.get("start"),
                            "end": params.get("end"),
                            "with_bagua": params.get("with_bagua"),
                            "gua_filter": summary.get("gua_filter") or params.get("gua_filter"),
                            "metrics": summary.get("metrics"),
                            "param_hash": ph,
                            "research_fingerprint": research_fp
                            or (summary.get("research_fingerprint") if isinstance(summary, dict) else None),
                            "experiment_id": experiment_id,
                            "variant_id": vid,
                            "signal_data_source": getattr(req, "signal_data_source", None),
                            "signal_adjustment": getattr(req, "signal_adjustment", None),
                            "dataset_id": getattr(req, "dataset_id", None),
                            "weekly_bar_mode": getattr(req, "weekly_bar_mode", None) or "local_aggregate",
                            "execution_data_source": getattr(req, "execution_data_source", None) or "local_vendor",
                            "execution_dataset_id": getattr(req, "execution_dataset_id", None),
                            "execution_adjustment": getattr(req, "execution_adjustment", None),
                            "raw_dataset_id": getattr(req, "signal_raw_parent_dataset_id", None),
                            "factor_dataset_id": getattr(req, "signal_factor_parent_dataset_id", None),
                            "signal_formula_version": getattr(req, "signal_formula_version", None),
                            "universe_dataset_id": getattr(req, "universe_dataset_id", None),
                            "universe_rule_version": getattr(req, "universe_rule_version", None),
                            "delist_exit_rule_version": getattr(req, "delist_exit_rule_version", None),
                            "delist_exit_scenario": getattr(req, "delist_exit_scenario", None),
                            "delist_recovery_discount": getattr(req, "delist_recovery_discount", None),
                            "signal_supplement_factor_dataset_id": getattr(
                                req, "signal_supplement_factor_dataset_id", None),
                            "baseline_generation": getattr(req, "baseline_generation", None),
                            "data_cutoff_date": getattr(req, "data_cutoff_date", None),
                        },
                    )
            st = "succeeded"
            if (summary.get("status") or "").startswith("no_go") or summary.get("status") == "failed":
                st = "failed"
            exp_db.update_variant(self.cfg, vid, status=st, run_id=rid, error=summary.get("error"))
            return vid, st, rid, summary.get("error")
        except Exception as e:
            err = f"{e}\n{traceback.format_exc()[:500]}"
            exp_db.update_variant(self.cfg, vid, status="failed", error=str(e)[:500])
            return vid, "failed", None, str(e)

    def _run_experiment(self, experiment_id: str) -> None:
        try:
            exp = exp_db.get_experiment(self.cfg, experiment_id)
            concurrency = max(1, int(exp.get("concurrency") or 1))
            # Re-run pending and failed when experiment is (re)started.
            pending = [
                v
                for v in (exp.get("variants") or [])
                if v.get("status") in ("pending", "failed")
            ]
            completed = int(exp.get("completed_variants") or 0)
            failed = int(exp.get("failed_variants") or 0)
            skipped = int(exp.get("skipped_variants") or 0)

            with ThreadPoolExecutor(max_workers=concurrency) as pool:
                futs = {
                    pool.submit(self._run_one, experiment_id, v): v for v in pending
                }
                for fut in as_completed(futs):
                    if self.is_cancelled(experiment_id):
                        break
                    try:
                        _vid, st, _rid, _err = fut.result()
                    except Exception:
                        st = "failed"
                    if st == "succeeded":
                        completed += 1
                    elif st == "skipped":
                        skipped += 1
                    elif st == "cancelled":
                        pass
                    else:
                        failed += 1
                    exp_db.update_experiment_status(
                        self.cfg,
                        experiment_id,
                        "running",
                        completed_variants=completed,
                        failed_variants=failed,
                        skipped_variants=skipped,
                    )

            # Gate C D2 §7: an experiment with ANY failed variant must not be
            # reported as completed/succeeded.
            if self.is_cancelled(experiment_id):
                final = "cancelled"
            elif failed > 0:
                final = "failed"
            else:
                final = "completed"
            exp_db.update_experiment_status(
                self.cfg,
                experiment_id,
                final,
                completed_variants=completed,
                failed_variants=failed,
                skipped_variants=skipped,
            )
            if final == "completed" and not self.is_cancelled(experiment_id):
                try:
                    self._promote_top_n_full(experiment_id)
                except Exception:
                    pass
        except Exception as e:
            exp_db.update_experiment_status(
                self.cfg, experiment_id, "failed"
            )
            # best-effort
            _ = e

    def _promote_top_n_full(self, experiment_id: str) -> None:
        """Re-run top-N variants with engine=full + artifact_level=full (Phase-3)."""
        exp = exp_db.get_experiment(self.cfg, experiment_id)
        conf = dict(exp.get("config") or {})
        top_n = int(conf.get("promote_top_n") or 0)
        if top_n <= 0:
            return
        metric = str(conf.get("promote_metric") or "total_return")
        table = exp_db.experiment_results_table(self.cfg, experiment_id)
        rows = list(table.get("rows") or [])
        scored = []
        for row in rows:
            if row.get("status") != "succeeded":
                continue
            m = row.get("metrics") or {}
            val = m.get(metric)
            if val is None:
                continue
            try:
                scored.append((float(val), row))
            except (TypeError, ValueError):
                continue
        if not scored:
            return
        scored.sort(key=lambda x: x[0], reverse=True)
        winners = scored[:top_n]
        for rank, (_score, row) in enumerate(winners, 1):
            if self.is_cancelled(experiment_id):
                break
            params = dict(row.get("params") or {})
            params.pop("_meta", None)
            params["engine"] = "full"
            params["artifact_level"] = "full"
            params["use_signal_cache"] = conf.get("use_signal_cache", True)
            vid = str(row.get("variant_id") or ("promote_%s" % rank))
            variant = {
                "variant_id": "%s__full" % vid,
                "params": params,
                "param_hash": None,
                "status": "pending",
            }
            try:
                _vid, st, rid, err = self._run_one(experiment_id, variant)
                exp_db.update_variant(
                    self.cfg,
                    _vid,
                    status=st,
                    run_id=rid,
                    error=(err[:500] if err else None),
                )
            except Exception:
                continue




_RUNNER: Optional[ExperimentRunner] = None
_RUNNER_LOCK = threading.Lock()


def get_runner(cfg: Optional[AStockConfig] = None) -> ExperimentRunner:
    global _RUNNER
    with _RUNNER_LOCK:
        if _RUNNER is None:
            _RUNNER = ExperimentRunner(cfg)
        return _RUNNER


def build_experiment_matrix_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Normalize experiment result rows for ``build_result_matrix``.

    Extracts ``exit_weekday``, ``sell_on``, ``gua_key`` from params/_meta and
    flattens metrics (``total_return``, ``max_drawdown``, ...) onto each row.
    Rows missing required axes are skipped.
    """
    out: List[Dict[str, Any]] = []
    for row in rows or []:
        p = dict(row.get("params") or {})
        meta = dict(p.get("_meta") or {})
        m = dict(row.get("metrics") or {})

        exit_wd = p.get("exit_weekday")
        if exit_wd is None:
            exit_wd = meta.get("exit_weekday")
        sell_on = p.get("sell_on")
        if sell_on is None:
            sell_on = meta.get("sell_on")
        gua_key = meta.get("gua_key")
        if gua_key is None:
            gua_key = p.get("gua_key")
        if gua_key is None:
            gua_key = "none"

        if exit_wd is None or sell_on is None:
            continue
        try:
            exit_wd_i = int(exit_wd)
        except (TypeError, ValueError):
            continue

        flat: Dict[str, Any] = {
            "exit_weekday": exit_wd_i,
            "sell_on": str(sell_on),
            "gua_key": str(gua_key),
            "variant_id": row.get("variant_id"),
            "status": row.get("status"),
            "run_id": row.get("run_id"),
            "param_hash": row.get("param_hash"),
            "id": row.get("variant_id") or row.get("param_hash"),
        }
        for k, v in m.items():
            flat[k] = v
        # also expose nested metrics for evaluate_trials
        flat["metrics"] = m
        out.append(flat)
    return out


def _append_matrix_sheet(wb, sheet_title: str, matrix: Dict[str, Any]) -> None:
    """Write one matrix table sheet: exit_weekday, sell_on, then gua columns."""
    ws = wb.create_sheet(sheet_title)
    cols = list(matrix.get("columns") or [])
    headers = ["exit_weekday", "sell_on"] + cols
    ws.append(headers)
    for entry in matrix.get("table") or []:
        ws.append([entry.get(h) for h in headers])


def write_experiment_excel(cfg: AStockConfig, experiment_id: str, path=None):
    """Write a results workbook for the experiment (summary + optional matrix/evaluate)."""
    from pathlib import Path

    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError("openpyxl required for experiment excel") from e

    from ..research.matrix import build_result_matrix

    table = exp_db.experiment_results_table(cfg, experiment_id)
    raw_rows = list(table.get("rows") or [])
    out = Path(path) if path else Path(cfg.output_root) / experiment_id / "experiment_summary.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "实验结果"
    headers = [
        "variant_id",
        "status",
        "run_id",
        "rule",
        "gua",
        "weekday",
        "stop_loss",
        "total_return",
        "annual_return",
        "max_drawdown",
        "win_rate",
        "n_round_trips",
        "payoff_ratio",
        "param_hash",
        "error",
    ]
    ws.append(headers)
    for row in raw_rows:
        p = row.get("params") or {}
        meta = p.get("_meta") or {}
        m = row.get("metrics") or {}
        ws.append(
            [
                row.get("variant_id"),
                row.get("status"),
                row.get("run_id"),
                meta.get("rule_id") or (p.get("rule_ids") or [None])[0],
                meta.get("gua_label") or meta.get("gua_key"),
                meta.get("weekday_label") or meta.get("weekday_key"),
                meta.get("stop_loss"),
                m.get("total_return"),
                m.get("annual_return"),
                m.get("max_drawdown"),
                m.get("win_rate"),
                m.get("n_round_trips"),
                m.get("payoff_ratio") or m.get("profit_loss_ratio"),
                row.get("param_hash"),
                row.get("error"),
            ]
        )

    # Matrix sheet(s) when rows carry exit_weekday + sell_on + gua
    matrix_rows = build_experiment_matrix_rows(raw_rows)
    if matrix_rows:
        m_ret = build_result_matrix(matrix_rows, metric_key="total_return")
        if m_ret.get("table"):
            _append_matrix_sheet(wb, "matrix", m_ret)
        # Second sheet when max_drawdown is present on any row
        if any(r.get("max_drawdown") is not None for r in matrix_rows):
            m_dd = build_result_matrix(matrix_rows, metric_key="max_drawdown")
            if m_dd.get("table"):
                _append_matrix_sheet(wb, "matrix_max_drawdown", m_dd)

    # Optional evaluate sheet from succeeded trials
    succeeded = [r for r in matrix_rows if r.get("total_return") is not None]
    status_ok = [
        r
        for r in matrix_rows
        if str(r.get("status") or "").lower()
        in ("succeeded", "success", "done", "ok", "completed")
    ]
    if len(status_ok) >= 2:
        trial_src = status_ok
    elif len(succeeded) >= 2:
        trial_src = succeeded
    else:
        trial_src = []

    if trial_src:
        try:
            from ..research.evaluation import evaluate_trials

            ev = evaluate_trials(trial_src)
            ranking = list(ev.get("ranking") or [])
            if ranking:
                ws_ev = wb.create_sheet("evaluate")
                rank_headers = [
                    "rank",
                    "id",
                    "composite",
                    "total_return",
                    "max_drawdown",
                    "win_rate",
                    "n_round_trips",
                    "exit_weekday",
                    "sell_on",
                    "gua_key",
                    "hard_ok",
                ]
                ws_ev.append(rank_headers)
                for i, r in enumerate(ranking[:50], 1):
                    tid = (
                        r.get("id")
                        or r.get("variant_id")
                        or r.get("trial_id")
                        or r.get("param_hash")
                    )
                    ws_ev.append(
                        [
                            i,
                            tid,
                            r.get("_composite") or r.get("composite"),
                            r.get("total_return"),
                            r.get("max_drawdown"),
                            r.get("win_rate"),
                            r.get("n_round_trips") or r.get("n_trades"),
                            r.get("exit_weekday"),
                            r.get("sell_on"),
                            r.get("gua_key"),
                            r.get("_hard_ok"),
                        ]
                    )
        except Exception:
            # evaluate is best-effort; summary/matrix still useful
            pass

    wb.save(out)
    return out
