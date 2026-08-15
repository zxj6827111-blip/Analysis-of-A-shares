# -*- coding: utf-8 -*-
"""Single-stock hexagram (卦象) query by code + date + period.

Uses the same OHLC digit-sum algorithm as backtest bagua attach:
  upper = open digit sum mod 8
  lower = close digit sum mod 8
  yao   = (high + low) digit sum mod 6

Price plane selectable:
  - raw / unadjusted: L2 unadjusted OHLC (historical default for this query tool)
  - standard_qfq / qfq: ordinary forward-adjust to snapshot end
  - asof_forward_qfq / asof: 时点动态前复权 anchored at query date (L1 formal)
"""

from __future__ import annotations

import re
import threading
import time as _bq_time
from datetime import date as _ymd_date
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

import numpy as np

from ..bagua.calculator import BaguaCalculator
from ..config import AStockConfig
from ..data.adjustments import build_factor_series
from ..data.data_store import DataStore
from ..data.tdx_reader import DayBar, TdxDayReader
from ..data.universe import to_std_code
from ..study import (
    build_period_bars,
    day_bars_for_signals,
    day_bars_for_signals_affine,
    day_bars_to_standard_qfq,
)
from ..data.affine_adjust import build_affine_series
from .index_etf import (
    classify_symbol,
    display_code as display_index_etf_code,
    list_etf_std_codes,
    load_index_etf_day_bars,
    resolve_index_etf_name,
    to_index_etf_std_code,
)
from .stock_names import display_code_with_name, resolve_stock_name


def _parse_ymd(value: Union[str, int, None]) -> int:
    if value is None or value == "":
        raise ValueError("date is required (YYYY-MM-DD or YYYYMMDD)")
    if isinstance(value, int):
        d = value
        if d < 19900101 or d > 21001231:
            raise ValueError(f"invalid date: {value}")
        return d
    s = str(value).strip().replace("/", "-")
    if "-" in s:
        parts = s.split("-")
        if len(parts) != 3:
            raise ValueError(f"invalid date: {value}")
        y, m, day = int(parts[0]), int(parts[1]), int(parts[2])
        d = y * 10000 + m * 100 + day
    else:
        digits = "".join(ch for ch in s if ch.isdigit())
        if len(digits) != 8:
            raise ValueError(f"invalid date: {value}")
        d = int(digits)
    if d < 19900101 or d > 21001231:
        raise ValueError(f"invalid date: {value}")
    return d


def _prev_month_end(ymd: int) -> int:
    """上一个月最后一天 (YYYYMMDD int)。用于月卦默认取查询月份的上一个月。"""
    y, m = ymd // 10000, (ymd // 100) % 100
    if m == 1:
        y, m = y - 1, 12
    else:
        m -= 1
    if m in (1, 3, 5, 7, 8, 10, 12):
        d = 31
    elif m in (4, 6, 9, 11):
        d = 30
    else:
        d = 29 if (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)) else 28
    return y * 10000 + m * 100 + d


def normalize_period(period: Optional[str]) -> str:
    p = (period or "DAY").strip().upper()
    if p in ("DAY", "D", "1D", "日", "日线", "按日"):
        return "DAY"
    if p in ("WEEK", "W", "1W", "周", "周线", "按周"):
        return "WEEK"
    if p in ("MONTH", "M", "1M", "月", "月线", "按月"):
        return "MONTH"
    raise ValueError("period must be DAY, WEEK, or MONTH")


def normalize_adjust_mode(mode: Optional[str]) -> str:
    """Return canonical: raw | tdx_front | tushare_qfq | standard_qfq | asof_forward_qfq."""
    m = (mode or "tushare_qfq").strip().lower()
    if m in ("raw", "unadjusted", "none", "未复权", "不复权"):
        return "raw"
    if m in ("tdx_front", "tdxquant_front", "通达信前复权"):
        return "tdx_front"
    if m in ("tushare_qfq", "ts_qfq", "tushare前复权"):
        return "tushare_qfq"
    if m in ("standard_qfq", "qfq", "ordinary_qfq", "forward", "前复权", "普通前复权"):
        return "standard_qfq"
    if m in (
        "asof_forward_qfq",
        "asof",
        "asof_qfq",
        "dynamic_qfq",
        "pit_forward",
        "时点前复权",
        "动态前复权",
    ):
        return "asof_forward_qfq"
    raise ValueError(
        "adjust must be raw | tdx_front (通达信前复权) | tushare_qfq (Tushare前复权) "
        "| standard_qfq (普通前复权) | asof_forward_qfq (时点前复权)"
    )


def normalize_query_code(raw: str) -> str:
    """Accept 600000 / sh600000 / SSE.STK.600000 -> WonderTrader std code.

    Index/ETF codes (sh000001 上证指数, sh510300 ETF, sz399001 深证成指,
    sz159915 ETF) map to SSE.IDX.* / SZSE.IDX.* / SSE.ETF.* / SZSE.ETF.*.
    """
    t = (raw or "").strip()
    if not t:
        raise ValueError("code is required")
    t = t.split()[0].split("　")[0]
    if t.startswith("SSE.") or t.startswith("SZSE."):
        return t
    idx_etf = to_index_etf_std_code(t)
    if idx_etf:
        return idx_etf
    # Stock ts_code form: 600000.SH / 000001.SZ / 920001.BJ (exchange is
    # explicit, so 000001.SZ resolves to the stock, not the index).
    m = re.match(r"^(\d{6})\.(SH|SZ|BJ)$", t, re.IGNORECASE)
    if m:
        return {"SH": "SSE", "SZ": "SZSE", "BJ": "BSE"}[m.group(2).upper()] + f".STK.{m.group(1)}"
    return to_std_code(t)


def display_code(std_code: str) -> str:
    if std_code.startswith("SSE.STK."):
        return "sh" + std_code.split(".")[-1]
    if std_code.startswith("SZSE.STK."):
        return "sz" + std_code.split(".")[-1]
    return display_index_etf_code(std_code)


def load_day_bars(cfg: AStockConfig, std_code: str) -> List[DayBar]:
    store = DataStore(cfg.storage_root)
    try:
        return store.load_symbol(std_code)
    except FileNotFoundError:
        pass
    reader = TdxDayReader(cfg.tdx_root)
    raw = ("sh" if std_code.startswith("SSE") else "sz") + std_code.split(".")[-1]
    bars, _ = reader.read(raw)
    if not bars:
        raise FileNotFoundError(f"no bars for {std_code}")
    return list(bars)


def _symbol_variants(symbol: str) -> List[str]:
    """SSE.STK.600000 / 600000.SH / sh600000 / 600000 等格式互转。"""
    from ..data.repository import MarketDataRepository

    return MarketDataRepository._symbol_variants(symbol)


class SourceDisabledError(ValueError):
    """A price plane was disabled by the Tushare-only data policy."""


def _source_match_pairs(source_key: str) -> List[Tuple[str, str]]:
    """Map UI adjust key to warehouse (source, adjustment) pairs, priority order.

    Tushare-only policy (tushare_only_v1):
      raw        -> formal L2 (internal/composite_none) first; the latest
                    complete tushare/none is the bootstrap fallback
      tushare_qfq -> formal L1 (internal/composite_tushare_factor_qfq) first;
                     without a formal pair the bootstrap fallback is the
                     read-only native tushare/qfq set; the legacy internal
                     tushare_factor_qfq derived set is no longer readable
      tdx_front  -> DISABLED: explicit requests raise SourceDisabledError
    """
    if source_key == "tdx_front":
        raise SourceDisabledError(
            "tdx_front 已停用：系统已切换为 Tushare-only 数据策略，不再提供"
            "通达信前复权数据面。请使用 tushare_qfq（正式 L1 前复权）或 raw"
            "（正式 L2 未复权）。"
        )
    if source_key == "tushare_qfq":
        # 正式 L1 优先；无正式面对时 bootstrap 只读后备为原生 tushare/qfq
        # （旧 internal/tushare_factor_qfq 派生集已不可读）
        return [
            ("internal", "composite_tushare_factor_qfq"),
            ("tushare", "qfq"),
        ]
    if source_key == "raw":
        # 未复权：与回测 L2 一致，正式 L2 优先，bootstrap 后备为完整 tushare/none
        return [
            ("internal", "composite_none"),
            ("tushare", "none"),
        ]
    raise ValueError(f"unsupported dataset source_key: {source_key}")


def _find_symbol_rec(manifest, std_code: str):
    """Locate symbol record with a readable blob (ok preferred)."""
    variants = set(_symbol_variants(std_code))
    ok_rec = None
    any_rec = None
    for r in manifest.symbols:
        if r.symbol not in variants or not r.blob_sha256:
            continue
        if r.quality == "ok":
            ok_rec = r
            break
        if any_rec is None:
            any_rec = r
    return ok_rec or any_rec


def _bars_from_blob(store, blob_sha256: str) -> List[DayBar]:
    arr = store.load_bars(blob_sha256)
    dates = arr["trade_date"]
    bars: List[DayBar] = []
    for i in range(len(dates)):
        bars.append(
            DayBar(
                date=int(dates[i]),
                open=float(arr["open"][i]),
                high=float(arr["high"][i]),
                low=float(arr["low"][i]),
                close=float(arr["close"][i]),
                amount=float(arr["amount"][i]),
                volume=float(arr["volume"][i]),
            )
        )
    if len(bars) > 1 and bars[0].date > bars[-1].date:
        bars.reverse()
    return bars


def load_day_bars_for_plane(
    cfg: AStockConfig,
    std_code: str,
    plane: str,
    *,
    asof: Optional[int] = None,
    start: Optional[int] = None,
    end: Optional[int] = None,
    session: Optional["BaguaPlaneSession"] = None,
) -> Tuple[List[DayBar], Dict[str, Any]]:
    """Load day bars for bagua price plane (shared by query + backtest).

    plane: raw | tdx_front | tushare_qfq
    Returns (bars, meta). For raw, prefers warehouse none sets then legacy day files.
    Optional start/end trim the series after load.

    Pass a shared ``BaguaPlaneSession`` from full-universe backtests so
    manifests are listed once (avoids per-symbol full-warehouse scans).
    """
    key = normalize_adjust_mode(plane)
    if key == "tdx_front":
        raise SourceDisabledError(
            "tdx_front 已停用：系统已切换为 Tushare-only 数据策略。"
            "请使用 tushare_qfq 或 raw。"
        )
    if key not in ("raw", "tushare_qfq"):
        raise ValueError(f"unsupported bagua price plane: {plane}")
    try:
        if session is not None:
            bars, meta = session.load_symbol(std_code, asof=asof)
        else:
            bars, meta = _load_dataset_bars(cfg, std_code, key, asof=asof)
    except FileNotFoundError:
        if key != "raw":
            raise
        bars = load_day_bars(cfg, std_code)
        meta = {
            "dataset_id": None,
            "dataset_source": "legacy_tdx_day",
            "dataset_adjustment": "none",
            "dataset_status": "legacy",
            "legacy_fallback": True,
        }
    if start is not None or end is not None:
        lo = int(start) if start is not None else 0
        hi = int(end) if end is not None else 10**9
        bars = [b for b in bars if lo <= int(b.date) <= hi]
    return bars, meta


class BaguaPlaneSession:
    """One-shot warehouse index for multi-symbol bagua plane loads.

    Built once per backtest bagua phase: lists matching manifests, builds
    per-manifest symbol indexes, and serves per-code blob loads without
    re-scanning the entire data root for every stock (full-A freeze fix).

    Tushare-only policy: candidates are restricted to the product role's
    dataset families, short-window orphan surfaces are ineligible, and the
    formal L1/L2 product pair is preferred. ``bootstrap_fallback`` in the
    returned meta marks datasets that are NOT the current formal surface
    (quick/read-only fallback only).
    """

    def __init__(self, cfg: AStockConfig, source_key: str):
        from ..data.dataset_store import DatasetStore

        self.source_key = normalize_adjust_mode(source_key)
        if self.source_key == "tdx_front":
            raise SourceDisabledError(
                "tdx_front 已停用：系统已切换为 Tushare-only 数据策略。"
                "请使用 tushare_qfq 或 raw。"
            )
        if self.source_key not in ("raw", "tushare_qfq"):
            raise ValueError(f"unsupported bagua price plane: {source_key}")
        md_root = getattr(cfg, "market_data_root", None)
        if not md_root or not Path(md_root).exists():
            raise FileNotFoundError(f"market data root not found: {md_root}")
        self.store = DatasetStore(md_root)
        self.match_pairs = _source_match_pairs(self.source_key)
        self.pair_rank = {pair: idx for idx, pair in enumerate(self.match_pairs)}
        self.status_rank = {"ready": 3, "partial": 2, "building": 1, "failed": 0}
        # list of (manifest, symbol_index dict, pair_priority)
        self._indexed: List[Tuple[Any, Dict[str, Any], int]] = []
        # per-manifest history signals (computed once in _build_index, reused
        # by every _score call — avoids an O(N) scan + np.median per symbol)
        self._manifest_sig: Dict[str, Any] = {}
        self._saw_any_pair = False
        # current formal product surface (Tushare-only), if any
        self.formal_l1_id: Optional[str] = None
        self.formal_l2_id: Optional[str] = None
        try:
            from ..data.tushare_product import resolve_active_tushare_product_pair

            # Read-only resolution: skip manifest deepcopy (100k+ symbol
            # records each) — this session never mutates the manifests.
            pair = resolve_active_tushare_product_pair(self.store, deep_copy=False)
            if pair is not None:
                self.formal_l1_id = pair.l1_dataset_id
                self.formal_l2_id = pair.l2_dataset_id
        except Exception:
            self.formal_l1_id = None
            self.formal_l2_id = None
        self._build_index()

    def _is_index_etf_surface(self, m: Any) -> bool:
        """True when a manifest's blob-bearing symbols are mainly index/ETF.

        Index/ETF symbols are synced into tushare/none surfaces of their own
        (separate from the stock full-market base). Newly listed ETFs
        legitimately start with a short window, so those fallback surfaces
        must not be dropped wholesale by the orphan-window gate.
        """
        syms = [r.symbol for r in m.symbols if r.blob_sha256]
        if not syms:
            return False
        idx_etf = sum(1 for s in syms if ".IDX." in s or ".ETF." in s)
        return idx_etf > 0 and idx_etf * 2 >= len(syms)

    def _build_index(self) -> None:
        from ..data.tushare_product import manifest_history_signals

        for mid in self.store.list_manifests():
            # Read-only indexing: reuse cached manifests (no deepcopy).
            m = self.store.load_manifest(mid, deep_copy=False)
            if m is None:
                continue
            if (m.period or "1d") not in ("1d", "", None):
                continue
            # eligibility: product candidates must be ready (partial/building/
            # failed never enter the selection chain)
            if m.status != "ready":
                continue
            pair = (m.source, m.adjustment)
            if pair not in self.pair_rank:
                continue
            formal_id = (
                self.formal_l2_id
                if self.source_key == "raw"
                else self.formal_l1_id
            )
            if formal_id:
                # Once the atomic product pair exists, only that exact
                # manifest is eligible. Same-date legacy composites or native
                # QFQ datasets must not win through row-count tie breakers.
                if m.dataset_id != formal_id:
                    # Index/ETF symbols are synced into tushare/none, never
                    # into the stock composite product pair; keep the latest
                    # ready tushare/none surface indexed on the raw plane so
                    # index/ETF queries still resolve from the warehouse.
                    if not (
                        self.source_key == "raw"
                        and pair == ("tushare", "none")
                    ):
                        continue
            else:
                # Bootstrap remains Tushare-only. Old internal composites may
                # carry local_vendor/TDX lineage and are therefore ineligible.
                bootstrap_pair = (
                    ("tushare", "none")
                    if self.source_key == "raw"
                    else ("tushare", "qfq")
                )
                if pair != bootstrap_pair:
                    continue
            # quality gate: quality_status must not be explicitly failed
            prov = getattr(m, "provenance", None) or {}
            if prov.get("quality_status") == "failed":
                continue
            sig = manifest_history_signals(m)
            self._manifest_sig[m.dataset_id] = sig
            # short-window orphan surfaces are ineligible as product
            # candidates — except tushare/none index/ETF fallback surfaces
            # (newly listed ETFs legitimately start with a short window and
            # would otherwise drop the whole fallback surface). Ready and
            # quality filters stay in force.
            is_idx_etf_fallback = (
                pair == ("tushare", "none") and self._is_index_etf_surface(m)
            )
            if sig.is_short_window and not is_idx_etf_fallback:
                continue
            self._saw_any_pair = True
            idx: Dict[str, Any] = {}
            for r in m.symbols:
                if not getattr(r, "blob_sha256", None):
                    continue
                # index all variants so lookup is O(1)
                for v in _symbol_variants(r.symbol):
                    prev = idx.get(v)
                    if prev is None:
                        idx[v] = r
                    elif getattr(r, "quality", None) == "ok" and getattr(prev, "quality", None) != "ok":
                        idx[v] = r
            self._indexed.append((m, idx, self.pair_rank[pair]))
        # Scan order does NOT rank by vendor priority: freshness and history
        # depth decide within the product role (plan 8.2).
        self._indexed.sort(
            key=lambda t: (
                self.status_rank.get(t[0].status or "", -1),
                int(t[0].data_cutoff_date or 0),
                int(t[0].symbol_count or 0),
                int(t[0].row_count or 0),
                t[0].created_at or "",
            ),
            reverse=True,
        )

    def load_symbol(
        self,
        std_code: str,
        *,
        asof: Optional[int] = None,
    ) -> Tuple[List[DayBar], Dict[str, Any]]:
        hits: List[Tuple[Any, Any, int, bool, int, int]] = []
        variants = _symbol_variants(std_code)
        # Variant priority matters: a bare 6-digit code collides across kinds
        # (e.g. SSE.IDX.000001 上证指数 vs SZSE.STK.000001 平安银行 both map to
        # "000001"). Qualified variants (canonical / ts_code / sh-prefix forms)
        # are tried first across every dataset; only when they match nothing do
        # we fall back to the bare code, so sh000001 / 000001.SH /
        # SSE.IDX.000001 resolve to the index and never to the same-coded stock.
        qualified = [v for v in variants if not (len(v) == 6 and v.isdigit())]
        bare = [v for v in variants if len(v) == 6 and v.isdigit()]

        def _scan(variant_list) -> Tuple[List[Tuple[Any, Any, int, bool, int, int]], bool]:
            """Collect (manifest, rec, pair_rank, covers_asof, first, last) hits
            for one variant tier; return (hits, matched_any_record)."""
            found: List[Tuple[Any, Any, int, bool, int, int]] = []
            matched_any = False
            for v in variant_list:
                for m, idx, pr in self._indexed:
                    rec = idx.get(v)
                    if rec is None:
                        continue
                    matched_any = True
                    d0 = int(rec.first_date) if rec.first_date else None
                    d1 = int(rec.last_date) if rec.last_date else None
                    if d0 is not None and d1 is not None:
                        first, last = (d0, d1) if d0 <= d1 else (d1, d0)
                    else:
                        first, last = d0, d1
                    covers_asof = False
                    has_on_or_before = True
                    if asof is not None:
                        if first is not None and last is not None:
                            covers_asof = first <= asof <= last
                            has_on_or_before = first <= asof
                        elif last is not None:
                            covers_asof = last >= asof
                        elif first is not None:
                            has_on_or_before = first <= asof
                    if asof is not None and not has_on_or_before:
                        continue
                    found.append(
                        (m, rec, pr, covers_asof, int(first or 0), int(last or 0))
                    )
                if found:
                    break
            return found, matched_any

        hits, qualified_matched = _scan(qualified)
        if not hits:
            if qualified_matched:
                # The requested symbol exists in its own kind but no bar covers
                # the query date — do NOT fall back to the bare code, which
                # would resolve to a same-coded stock (e.g. 上证指数 queried
                # before its inception would return 平安银行).
                raise FileNotFoundError(
                    f"{std_code} 无 {asof} 当日或之前K线"
                    if asof is not None
                    else f"{std_code} 在 {self.source_key} 全部数据集中均无可用K线"
                )
            hits, _ = _scan(bare)

        if not hits:
            if not self._saw_any_pair:
                raise FileNotFoundError(
                    f"未找到 {self.source_key} 数据集，请先同步 Tushare 前复权数据"
                )
            raise FileNotFoundError(
                f"{std_code} 在 {self.source_key} 全部数据集中均无可用K线"
                + (f"（查询日 {asof}）" if asof else "")
            )

        # Formal product lock for stocks: once the atomic L1/L2 pair exists,
        # stock queries must resolve only from the formal composite. The
        # extra tushare/none surfaces indexed in _build_index exist for
        # index/ETF symbols, which are outside the composite pair.
        formal_id = (
            self.formal_l2_id if self.source_key == "raw" else self.formal_l1_id
        )
        is_index_etf = any(".IDX." in v or ".ETF." in v for v in variants)
        if formal_id and not is_index_etf:
            hits = [h for h in hits if h[0].dataset_id == formal_id]
            if not hits:
                raise FileNotFoundError(
                    f"{std_code} 在 {self.source_key} 全部数据集中均无可用K线"
                    + (f"（查询日 {asof}）" if asof else "")
                )

        def _score(item):
            m, rec, pr, covers, _first, latest = item
            cutoff = int(m.data_cutoff_date or latest or 0)
            rows = int(rec.row_count or 0)
            # history depth: median per-symbol rows (1.0 = ~1000+ rows).
            # Per-manifest signals were computed once in _build_index, so
            # scoring never re-scans the manifest symbol records.
            sig = self._manifest_sig.get(m.dataset_id)
            med = float(sig.median_rows) if sig is not None else float(rows)
            completeness = min(1.0, med / 1000.0)
            if asof is None:
                # "Latest available" queries: freshest real data wins within the
                # product role — no vendor/source priority overrides freshness.
                return (
                    latest,
                    cutoff,
                    completeness,
                    rows,
                    m.created_at or "",
                )
            # Historical asof: the candidate must span the query date; the
            # anchored (cutoff >= asof, nearest) version wins, which keeps the
            # point-in-time semantics of derived QFQ versions. Distances are
            # negated so the NEAREST version sorts first under reverse=True.
            near = -(
                abs(cutoff - asof)
                if cutoff >= asof
                else (asof - cutoff) + 10**9
            )
            return (
                1 if covers else 0,
                1 if cutoff >= asof else 0,
                near,
                latest,
                completeness,
                rows,
                m.created_at or "",
            )

        hits.sort(key=_score, reverse=True)
        manifest, rec, _, covers_asof, effective_first, effective_last = hits[0]
        bars = _bars_from_blob(self.store, rec.blob_sha256)
        if not bars:
            raise FileNotFoundError(
                f"{std_code} 数据集 {manifest.dataset_id} blob 为空"
            )
        formal_id = (
            self.formal_l2_id if self.source_key == "raw" else self.formal_l1_id
        )
        prov = getattr(manifest, "provenance", None) or {}
        meta: Dict[str, Any] = {
            "dataset_id": manifest.dataset_id,
            "dataset_source": manifest.source,
            "dataset_adjustment": manifest.adjustment,
            "dataset_status": manifest.status,
            "dataset_cutoff": manifest.data_cutoff_date,
            "symbol_first_date": rec.first_date,
            "symbol_last_date": rec.last_date,
            "symbol_row_count": rec.row_count,
            "covers_asof": bool(covers_asof),
            "symbol_effective_first_date": effective_first,
            "symbol_effective_last_date": effective_last,
            "candidate_datasets": len(hits),
            "session_indexed": True,
            "data_policy": prov.get("data_policy"),
            "expected_formal_l1_id": self.formal_l1_id,
            "expected_formal_l2_id": self.formal_l2_id,
            "bootstrap_fallback": bool(
                formal_id is None or manifest.dataset_id != formal_id
            ),
        }
        return bars, meta


# Process-wide plane-session reuse. Building the warehouse index is the
# dominant cost of single-shot queries (scan all manifests + per-manifest
# symbol indexes, ~1-2s with 50+ datasets). Sessions are read-only after
# construction, so concurrent reads are safe under a plain TTL cache.
_SESSION_CACHE_TTL = 300.0
_session_cache_lock = threading.Lock()
_session_cache: Dict[Tuple[str, str], Tuple[float, "BaguaPlaneSession"]] = {}


def _get_plane_session(cfg: AStockConfig, source_key: str) -> "BaguaPlaneSession":
    """Return a cached (TTL) BaguaPlaneSession for the given price plane."""
    md_root = getattr(cfg, "market_data_root", None)
    key = (str(md_root), normalize_adjust_mode(source_key))
    now = _bq_time.time()
    with _session_cache_lock:
        hit = _session_cache.get(key)
        if hit is not None and now - hit[0] < _SESSION_CACHE_TTL:
            return hit[1]
    session = BaguaPlaneSession(cfg, source_key)
    with _session_cache_lock:
        _session_cache[key] = (now, session)
    return session


def _load_dataset_bars(
    cfg: AStockConfig,
    std_code: str,
    source_key: str,
    asof: Optional[int] = None,
) -> Tuple[List[DayBar], Dict[str, Any]]:
    """Load day bars from warehouse for bagua query (cached shared session)."""
    session = _get_plane_session(cfg, source_key)
    return session.load_symbol(std_code, asof=asof)


def _find_day_bar(bars: Sequence[DayBar], asof: int) -> Tuple[DayBar, bool]:
    if not bars:
        raise FileNotFoundError("empty bar series")
    by_date = {int(b.date): b for b in bars}
    if asof in by_date:
        return by_date[asof], True
    candidates = [b for b in bars if int(b.date) <= asof]
    if not candidates:
        first = bars[0]
        raise FileNotFoundError(
            f"no bar on/before {asof}; first available {first.date}"
        )
    return candidates[-1], False


def _find_period_bar(
    day_bars: Sequence[DayBar],
    period: str,
    asof: int,
) -> Tuple[Any, bool]:
    p_bars = build_period_bars(day_bars, period, asof=asof, include_open=True)
    if not p_bars:
        raise FileNotFoundError(f"no {period} bar for {asof}")
    for pb in reversed(p_bars):
        start = int(getattr(pb, "start_date", pb.date))
        end = int(getattr(pb, "end_date", pb.date))
        if start <= asof <= end:
            exact = end == asof or bool(getattr(pb, "closed", True))
            return pb, exact
    before = [pb for pb in p_bars if int(getattr(pb, "end_date", pb.date)) <= asof]
    if not before:
        return p_bars[0], False
    return before[-1], False


def _adjust_day_bars(
    cfg: AStockConfig,
    std_code: str,
    day_raw: Sequence[DayBar],
    adjust: str,
    asof: int,
) -> Tuple[List[DayBar], Dict[str, Any]]:
    """Return (day bars in chosen price plane, meta)."""
    meta: Dict[str, Any] = {
        "adjust": adjust,
        "price_plane": "L2_trade_price" if adjust == "raw" else "L1_signal_price",
    }
    if adjust == "raw" or not day_raw:
        meta["price_format"] = "unadjusted, 2 decimal places"
        return list(day_raw), meta

    dates = [int(b.date) for b in day_raw]

    affine = build_affine_series(std_code, dates, adj_root=cfg.adj_root)
    if affine.quality == "complete" and not affine.is_identity:
        out = day_bars_for_signals_affine(
            day_raw,
            affine,
            research_unadjusted=False,
            signal_adjust=adjust,
            asof_date=asof,
        )
        meta["factor_source"] = affine.source
        meta["factor_quality"] = affine.quality
        meta["factor_manifest_sha"] = affine.sha256
        meta["model"] = "affine"
        if adjust == "standard_qfq":
            meta["price_format"] = "standard_qfq affine (a*raw+b), 2 decimal places"
            meta["signal_adjust"] = "standard_qfq"
        else:
            meta["price_format"] = "asof_forward_qfq affine (a*raw+b), 2 decimal places"
            meta["signal_adjust"] = "asof_forward_qfq"
            meta["asof_date"] = asof
        return out, meta

    series = build_factor_series(
        std_code, dates, adj_root=cfg.adj_root, prefer_baostock=True
    )
    fac = np.array(series.factors, dtype=float)
    meta["factor_source"] = series.source
    meta["factor_quality"] = series.quality
    meta["factor_manifest_sha"] = series.sha256
    meta["model"] = "multiplicative_fallback"

    if adjust == "standard_qfq":
        out = day_bars_to_standard_qfq(day_raw, fac)
        meta["price_format"] = "standard_qfq (factor_t/snapshot_end), 2 decimal places"
        meta["signal_adjust"] = "standard_qfq"
    else:
        out = day_bars_for_signals(
            day_raw,
            fac,
            research_unadjusted=False,
            signal_adjust="asof_forward_qfq",
            asof_date=asof,
            dates=dates,
        )
        meta["price_format"] = (
            "asof_forward_qfq (factor_t/factor_asof at query date), 2 decimal places"
        )
        meta["signal_adjust"] = "asof_forward_qfq"
        meta["asof_date"] = asof
    return out, meta


def _query_bagua_index_etf(
    cfg: AStockConfig,
    *,
    std: str,
    symbol_type: str,
    asof: int,
    per: str,
    requested_adjust: str,
    calc: BaguaCalculator,
) -> Dict[str, Any]:
    """Hexagram for index / ETF symbols (always unadjusted raw prices).

    Indices have no 复权 concept and ETF factor data is not stored in the
    warehouse, so the price plane is always raw regardless of the requested
    adjust mode (a note records the original request).

    Data sources, mirroring stocks: warehouse datasets first (tushare/none
    etc.), falling back to TDX local day files.
    """
    std_id = to_index_etf_std_code(std) or std
    try:
        day_bars, wmeta = _load_dataset_bars(cfg, std_id, "raw", asof=asof)
        ds_source = wmeta.get("dataset_source")
        ds_adjust = wmeta.get("dataset_adjustment")
        ds_status = wmeta.get("dataset_status")
        ds_id = wmeta.get("dataset_id")
        model = "warehouse"
        src_desc = "Tushare 数据仓库" if ds_source == "tushare" else f"数据仓库({ds_source}/{ds_adjust})"
    except FileNotFoundError:
        try:
            day_bars = load_index_etf_day_bars(cfg, std)
        except FileNotFoundError:
            raise FileNotFoundError(f"no market data for {display_code(std)}")
        ds_source, ds_adjust, ds_status, ds_id = "legacy_tdx_day", "none", "legacy", None
        model = "legacy_day_file"
        src_desc = "通达信本地 day 文件"

    adj = "raw"
    adj_meta: Dict[str, Any] = {
        "adjust": adj,
        "price_plane": "L2_trade_price",
        "price_format": f"未复权（指数/ETF 无复权口径，{src_desc}，两位小数）",
        "signal_adjust": adj,
        "model": model,
        "dataset_id": ds_id,
        "dataset_source": ds_source,
        "dataset_adjustment": ds_adjust,
        "dataset_status": ds_status,
        "legacy_fallback": model == "legacy_day_file",
        "requested_adjust": requested_adjust,
    }

    if per == "DAY":
        bar, exact = _find_day_bar(day_bars, asof)
        bar_meta = {
            "date": int(bar.date),
            "start_date": int(bar.date),
            "end_date": int(bar.date),
            "n_days": 1,
            "closed": True,
            "open": float(bar.open),
            "high": float(bar.high),
            "low": float(bar.low),
            "close": float(bar.close),
        }
        o, h, l, c = bar.open, bar.high, bar.low, bar.close
    else:
        pb, exact = _find_period_bar(day_bars, per, asof)
        bar_meta = {
            "date": int(pb.date),
            "start_date": int(getattr(pb, "start_date", pb.date)),
            "end_date": int(getattr(pb, "end_date", pb.date)),
            "n_days": int(getattr(pb, "n_days", 1)),
            "closed": bool(getattr(pb, "closed", True)),
            "open": float(pb.open),
            "high": float(pb.high),
            "low": float(pb.low),
            "close": float(pb.close),
        }
        o, h, l, c = pb.open, pb.high, pb.low, pb.close

    result = calc.calculate(open_price=o, high_price=h, low_price=l, close_price=c)
    bagua = result.to_dict()

    notes: List[str] = []
    notes.append(
        "算法：开盘定上卦(mod8)、收盘定下卦(mod8)、最高+最低定动爻(mod6)；"
        f"指数/ETF 无复权概念，价格直接读取{src_desc}（未复权、两位小数）。"
    )
    if requested_adjust != "raw":
        notes.append(
            f"请求的复权口径 {requested_adjust} 对指数/ETF 不适用，已按未复权(raw)计算。"
        )
    if not exact and per == "DAY":
        notes.append(f"请求日期 {asof} 非交易日或无日线，已使用最近交易日 {bar_meta['date']}。")
    if per != "DAY" and not bar_meta.get("closed", True):
        notes.append(f"该{('周' if per == 'WEEK' else '月')}K 尚未收官，卦象可能随后续交易日变化。")
    if per == "DAY" and int(bar_meta["date"]) != asof:
        notes.append(f"实际使用日线日期：{bar_meta['date']}。")

    name = resolve_index_etf_name(std)
    disp = display_code(std)

    return {
        "ok": True,
        "code": disp,
        "name": name,
        "display": display_code_with_name(disp, name),
        "std_code": std,
        "symbol_type": symbol_type,
        "query_date": asof,
        "period": per,
        "adjust": adj,
        "price_plane": adj_meta.get("price_plane"),
        "bar_date_exact": exact if per == "DAY" else (int(bar_meta["end_date"]) == asof),
        "bar": bar_meta,
        "bagua": bagua,
        "algorithm": {
            "open_to_upper": "digit_sum(open) mod 8 (0→8)",
            "close_to_lower": "digit_sum(close) mod 8 (0→8)",
            "hl_to_yao": "digit_sum(high)+digit_sum(low) mod 6 (0→6)",
            "price_format": adj_meta.get("price_format"),
            "adjust": adj,
        },
        "adjust_meta": adj_meta,
        "notes": notes,
        "summary": {
            "full_name": bagua.get("full_name") or bagua.get("gua_name") or "",
            "yao_name": bagua.get("yao_name") or bagua.get("line_name") or "",
            "state_id": bagua.get("state_id") or "",
            "action_signal": bagua.get("action_signal") or "",
            "market_judgement": bagua.get("market_judgement")
            or bagua.get("market_summary")
            or "",
            "upper": f"{bagua.get('upper_alias') or bagua.get('upper_name') or ''}"
            f"({bagua.get('upper_id')})",
            "lower": f"{bagua.get('lower_alias') or bagua.get('lower_name') or ''}"
            f"({bagua.get('lower_id')})",
            "yao_order": bagua.get("yao_order"),
        },
    }


def query_bagua(
    cfg: AStockConfig,
    *,
    code: str,
    date: Union[str, int],
    period: str = "DAY",
    adjust: str = "raw",
    session: Optional["BaguaPlaneSession"] = None,
    calc: Optional[BaguaCalculator] = None,
) -> Dict[str, Any]:
    """Query hexagram for one stock at a given date and period.

    adjust: raw | tdx_front | tushare_qfq | standard_qfq | asof_forward_qfq

    Optional ``session`` / ``calc`` reuse the warehouse index and knowledge
    base across multi-stock batch / export (avoids per-symbol full scans).
    """
    std = normalize_query_code(code)
    asof = _parse_ymd(date)
    per = normalize_period(period)
    adj = normalize_adjust_mode(adjust)

    if adj == "tdx_front":
        raise SourceDisabledError(
            "tdx_front 已停用：系统已切换为 Tushare-only 数据策略，不再提供"
            "通达信前复权数据面。请使用 tushare_qfq（正式 L1）或 raw（正式 L2）。"
        )

    if calc is None:
        if not cfg.bagua_json:
            raise FileNotFoundError("bagua knowledge json not configured")
        calc = BaguaCalculator.from_json(cfg.bagua_json)

    symbol_type = classify_symbol(code)
    if symbol_type in ("index", "etf"):
        return _query_bagua_index_etf(
            cfg,
            std=std,
            symbol_type=symbol_type,
            asof=asof,
            per=per,
            requested_adjust=adj,
            calc=calc,
        )

    if adj in ("tdx_front", "tushare_qfq", "raw"):
        try:
            if session is not None:
                day_bars, ds_meta = session.load_symbol(std, asof=asof)
            else:
                day_bars, ds_meta = _load_dataset_bars(cfg, std, adj, asof=asof)
        except FileNotFoundError:
            if adj != "raw":
                raise
            # 仓库无未复权集时回退旧 DataStore / 通达信 day 文件
            day_bars = load_day_bars(cfg, std)
            ds_meta = {
                "dataset_id": None,
                "dataset_source": "legacy_tdx_day",
                "dataset_adjustment": "none",
                "dataset_status": "legacy",
                "covers_asof": None,
                "candidate_datasets": 0,
                "legacy_fallback": True,
            }
        if not day_bars:
            raise FileNotFoundError(f"no market data for {display_code(std)}")
        if adj == "tdx_front":
            src_label = "通达信前复权数据集"
            price_plane = "L1_signal_price"
        elif adj == "tushare_qfq":
            ds_src = ds_meta.get("dataset_source")
            ds_adj = ds_meta.get("dataset_adjustment")
            if ds_src == "tushare" and ds_adj == "qfq":
                src_label = "Tushare官方前复权数据集"
            elif ds_adj in ("tushare_factor_qfq", "composite_tushare_factor_qfq"):
                src_label = "派生Tushare因子前复权数据集"
            else:
                src_label = "Tushare前复权数据集"
            price_plane = "L1_signal_price"
        else:
            ds_src = ds_meta.get("dataset_source")
            if ds_src == "internal":
                src_label = "正式L2复合数据集" if ds_meta.get("dataset_adjustment") == "composite_none" else "内部未复权数据集"
            elif ds_src == "tushare":
                src_label = "Tushare未复权日线"
            elif ds_src == "legacy_tdx_day":
                src_label = "本地通达信day文件(未复权后备)"
            else:
                src_label = "未复权数据集"
            price_plane = "L2_trade_price"
        adj_meta: Dict[str, Any] = {
            "adjust": adj,
            "price_plane": price_plane,
            "price_format": f"{src_label}（仓库直接读取，两位小数）"
            if not ds_meta.get("legacy_fallback")
            else f"{src_label}（两位小数）",
            "signal_adjust": adj,
            "model": "dataset_precomputed" if not ds_meta.get("legacy_fallback") else "legacy_day_file",
            **ds_meta,
        }
    else:
        day_raw = load_day_bars(cfg, std)
        if not day_raw:
            raise FileNotFoundError(f"no market data for {display_code(std)}")
        day_bars, adj_meta = _adjust_day_bars(cfg, std, day_raw, adj, asof)

    if per == "DAY":
        bar, exact = _find_day_bar(day_bars, asof)
        bar_meta = {
            "date": int(bar.date),
            "start_date": int(bar.date),
            "end_date": int(bar.date),
            "n_days": 1,
            "closed": True,
            "open": float(bar.open),
            "high": float(bar.high),
            "low": float(bar.low),
            "close": float(bar.close),
        }
        o, h, l, c = bar.open, bar.high, bar.low, bar.close
    else:
        pb, exact = _find_period_bar(day_bars, per, asof)
        bar_meta = {
            "date": int(pb.date),
            "start_date": int(getattr(pb, "start_date", pb.date)),
            "end_date": int(getattr(pb, "end_date", pb.date)),
            "n_days": int(getattr(pb, "n_days", 1)),
            "closed": bool(getattr(pb, "closed", True)),
            "open": float(pb.open),
            "high": float(pb.high),
            "low": float(pb.low),
            "close": float(pb.close),
        }
        o, h, l, c = pb.open, pb.high, pb.low, pb.close

    result = calc.calculate(open_price=o, high_price=h, low_price=l, close_price=c)
    bagua = result.to_dict()

    notes: List[str] = []
    if adj == "raw":
        ds_id = adj_meta.get("dataset_id") or ""
        ds_pair = f"{adj_meta.get('dataset_source')}/{adj_meta.get('dataset_adjustment')}"
        if adj_meta.get("legacy_fallback"):
            notes.append(
                "算法：开盘定上卦(mod8)、收盘定下卦(mod8)、最高+最低定动爻(mod6)；"
                "价格未复权，来自本地通达信day文件后备路径。"
            )
        else:
            notes.append(
                "算法：开盘定上卦(mod8)、收盘定下卦(mod8)、最高+最低定动爻(mod6)；"
                "价格未复权，直接读取仓库原始数据集"
                + (f"：{ds_pair} @ {ds_id}" if ds_id else "。")
            )
    elif adj == "tdx_front":
        ds_id = adj_meta.get("dataset_id") or ""
        notes.append(
            "算法同未复权；价格直接读取仓库通达信前复权数据集（与通达信软件一致，不做二次因子计算）"
            + (f"：{ds_id}" if ds_id else "。")
        )
    elif adj == "tushare_qfq":
        ds_id = adj_meta.get("dataset_id") or ""
        ds_pair = f"{adj_meta.get('dataset_source')}/{adj_meta.get('dataset_adjustment')}"
        notes.append(
            "算法同未复权；价格直接读取仓库 Tushare 前复权数据（不做二次因子计算）"
            + (f"：{ds_pair} @ {ds_id}" if ds_id else "。")
        )
    elif adj == "standard_qfq":
        notes.append(
            "算法同未复权；价格为普通前复权(standard_qfq，锚点=因子快照末端)，与盘面通达信风格接近。"
        )
    else:
        notes.append(
            "算法同未复权；价格为时点动态前复权(asof_forward_qfq)，锚点=查询日及以前可知因子，与回测 L1 信号默认一致。"
        )
    if not exact and per == "DAY":
        notes.append(f"请求日期 {asof} 非交易日或无日线，已使用最近交易日 {bar_meta['date']}。")
    if per != "DAY" and not bar_meta.get("closed", True):
        notes.append(f"该{('周' if per == 'WEEK' else '月')}K 尚未收官，卦象可能随后续交易日变化。")
    if per == "DAY" and int(bar_meta["date"]) != asof:
        notes.append(f"实际使用日线日期：{bar_meta['date']}。")

    stock_name = resolve_stock_name(cfg, display_code(std), std_code=std)

    return {
        "ok": True,
        "code": display_code(std),
        "name": stock_name,
        "display": display_code_with_name(display_code(std), stock_name),
        "std_code": std,
        "symbol_type": "stock",
        "query_date": asof,
        "period": per,
        "adjust": adj,
        "price_plane": adj_meta.get("price_plane"),
        "bar_date_exact": exact if per == "DAY" else (int(bar_meta["end_date"]) == asof),
        "bar": bar_meta,
        "bagua": bagua,
        "algorithm": {
            "open_to_upper": "digit_sum(open) mod 8 (0→8)",
            "close_to_lower": "digit_sum(close) mod 8 (0→8)",
            "hl_to_yao": "digit_sum(high)+digit_sum(low) mod 6 (0→6)",
            "price_format": adj_meta.get("price_format"),
            "adjust": adj,
        },
        "adjust_meta": adj_meta,
        "notes": notes,
        "summary": {
            "full_name": bagua.get("full_name") or bagua.get("gua_name") or "",
            "yao_name": bagua.get("yao_name") or bagua.get("line_name") or "",
            "state_id": bagua.get("state_id") or "",
            "action_signal": bagua.get("action_signal") or "",
            "market_judgement": bagua.get("market_judgement")
            or bagua.get("market_summary")
            or "",
            "upper": f"{bagua.get('upper_alias') or bagua.get('upper_name') or ''}"
            f"({bagua.get('upper_id')})",
            "lower": f"{bagua.get('lower_alias') or bagua.get('lower_name') or ''}"
            f"({bagua.get('lower_id')})",
            "yao_order": bagua.get("yao_order"),
        },
    }


def _resolve_batch_codes(
    cfg: AStockConfig,
    codes: Optional[Sequence[str]] = None,
    *,
    all_stocks: bool = False,
) -> List[str]:
    """Resolve display/std codes for batch query or full-market export."""
    from .backtest_universe import select_universe

    if all_stocks:
        return select_universe(cfg, None)
    raw = [str(c).strip() for c in (codes or []) if str(c).strip()]
    if not raw:
        raise ValueError("codes or all_stocks required")
    # De-dupe while preserving order (by std code)
    seen: set = set()
    out: List[str] = []
    for c in raw:
        try:
            std = normalize_query_code(c)
        except ValueError:
            continue
        if std in seen:
            continue
        seen.add(std)
        out.append(c)
    if not out:
        raise ValueError("no valid stock codes")
    return out


def _enumerate_export_etf_pool(cfg: AStockConfig) -> List[str]:
    """ETF pool for full-market exports: warehouse first, TDX fallback.

    The Tushare-only deployment has NO local TDX day files, so enumerating
    ``vipdoc/.../lday`` (list_etf_std_codes) yields an empty (or near-empty)
    ``etf-all`` export sheet. Prefer the ETF symbols present in the ready
    tushare/none warehouse datasets (blob-backed), falling back to TDX day
    files only when the warehouse has none.
    """
    try:
        from ..data.repository import MarketDataRepository

        md_root = getattr(cfg, "market_data_root", None)
        if md_root and Path(md_root).exists():
            repo = MarketDataRepository.from_root(md_root)
            etfs: set = set()
            for m in repo.list_datasets(
                source="tushare", adjustment="none", period="1d",
                deep_copy=False,
            ):
                if m.status != "ready":
                    continue
                for r in m.symbols:
                    if r.blob_sha256 and ".ETF." in r.symbol:
                        etfs.add(r.symbol)
            if etfs:
                return sorted(etfs)
    except Exception:
        pass
    # TDX 兜底(本地有通达信盘后数据时)
    return list_etf_std_codes(cfg)


def batch_query_bagua(
    cfg: AStockConfig,
    *,
    codes: Optional[Sequence[str]] = None,
    all_stocks: bool = False,
    date: Union[str, int],
    period: str = "DAY",
    adjust: str = "tushare_qfq",
    limit: Optional[int] = None,
    on_progress: Optional[Any] = None,
    session: Optional["BaguaPlaneSession"] = None,
) -> Dict[str, Any]:
    """Multi-stock hexagram query (same algorithm as single-stock).

    Uses a shared ``BaguaPlaneSession`` for warehouse planes so full-market
    scans do not re-list manifests per symbol.

    Optional ``on_progress(done, total, ok_count, err_count, code)`` callback.
    Optional ``session`` reuses a caller-built plane session (avoids building
    a second warehouse index when the target query already built one).
    """
    asof = _parse_ymd(date)
    per = normalize_period(period)
    adj = normalize_adjust_mode(adjust)
    code_list = _resolve_batch_codes(cfg, codes, all_stocks=all_stocks)
    if limit is not None:
        code_list = code_list[: int(limit)]

    if not cfg.bagua_json:
        raise FileNotFoundError("bagua knowledge json not configured")
    calc = BaguaCalculator.from_json(cfg.bagua_json)

    if adj in ("tdx_front", "tushare_qfq", "raw"):
        if adj == "tdx_front":
            raise SourceDisabledError(
                "tdx_front 已停用：系统已切换为 Tushare-only 数据策略。"
                "请使用 tushare_qfq 或 raw。"
            )
        if session is None:
            try:
                session = BaguaPlaneSession(cfg, adj)
            except FileNotFoundError:
                session = None

    results: List[Dict[str, Any]] = []
    ok_count = 0
    err_count = 0
    total = len(code_list)
    for idx, raw_code in enumerate(code_list, 1):
        try:
            row = query_bagua(
                cfg,
                code=raw_code,
                date=asof,
                period=per,
                adjust=adj,
                session=session,
                calc=calc,
            )
            results.append(row)
            ok_count += 1
        except Exception as e:
            err_count += 1
            try:
                std = normalize_query_code(raw_code)
                disp = display_code(std)
            except Exception:
                std = ""
                disp = str(raw_code)
            name = ""
            try:
                name = resolve_stock_name(cfg, disp, std_code=std or None) or ""
            except Exception:
                pass
            results.append(
                {
                    "ok": False,
                    "code": disp,
                    "name": name,
                    "display": display_code_with_name(disp, name) if name else disp,
                    "std_code": std,
                    "symbol_type": classify_symbol(raw_code),
                    "query_date": asof,
                    "period": per,
                    "adjust": adj,
                    "error": str(e),
                }
            )
        if on_progress is not None and (idx == total or idx % 25 == 0 or idx == 1):
            try:
                on_progress(idx, total, ok_count, err_count, raw_code)
            except Exception:
                pass

    return {
        "ok": True,
        "query_date": asof,
        "period": per,
        "adjust": adj,
        "all_stocks": bool(all_stocks),
        "requested": len(code_list),
        "count": len(results),
        "ok_count": ok_count,
        "error_count": err_count,
        "results": results,
    }


def _bagua_match_target(row: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """Compact projection of the target stock for same-hexagram matching."""
    if not row or row.get("error"):
        return {}
    b = row.get("bagua") or {}
    s = row.get("summary") or {}
    return {
        "code": row.get("code"),
        "name": row.get("name"),
        "display": row.get("display") or row.get("code") or "",
        "std_code": row.get("std_code"),
        "state_id": s.get("state_id") or b.get("state_id") or "",
        "full_name": s.get("full_name") or b.get("full_name") or "",
        "yao_name": s.get("yao_name") or b.get("yao_name") or "",
        "yao_order": s.get("yao_order")
        if s.get("yao_order") is not None
        else b.get("yao_order"),
    }


def _bagua_match_row(row: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """Compact projection of one matched stock for same-hexagram matching."""
    if not row or row.get("error"):
        return {"ok": False}
    b = row.get("bagua") or {}
    s = row.get("summary") or {}
    bar = row.get("bar") or {}
    return {
        "ok": True,
        "code": row.get("code"),
        "name": row.get("name"),
        "display": row.get("display") or row.get("code") or "",
        "std_code": row.get("std_code"),
        "state_id": s.get("state_id") or b.get("state_id") or "",
        "full_name": s.get("full_name") or b.get("full_name") or "",
        "gua_name": b.get("gua_name") or "",
        "yao_name": s.get("yao_name") or b.get("yao_name") or "",
        "yao_order": s.get("yao_order")
        if s.get("yao_order") is not None
        else b.get("yao_order"),
        "action_signal": s.get("action_signal") or b.get("action_signal") or "",
        "market_judgement": s.get("market_judgement")
        or b.get("market_judgement")
        or "",
        "bar": {
            "date": bar.get("date"),
            "start_date": bar.get("start_date"),
            "end_date": bar.get("end_date"),
            "open": bar.get("open"),
            "high": bar.get("high"),
            "low": bar.get("low"),
            "close": bar.get("close"),
        },
    }


def find_same_bagua(
    cfg: AStockConfig,
    *,
    code: str,
    date: Union[str, int],
    period: str = "DAY",
    adjust: str = "raw",
    scope: Optional[Sequence[str]] = None,
    limit: Optional[int] = None,
    on_progress: Optional[Any] = None,
) -> Dict[str, Any]:
    """Find other stocks with the exact same hexagram state (主卦+动爻, 384态).

    Computes the target stock's ``state_id`` (``gua_order-yao_order``) at the
    given date/period/adjust, scans the whole market (or ``scope`` codes) with
    the same parameters, and returns every stock whose ``state_id`` matches.

    A shared ``BaguaPlaneSession`` covers both the target query and the batch
    scan so the warehouse index is built only once.
    """
    std = normalize_query_code(code)
    asof = _parse_ymd(date)
    per = normalize_period(period)
    adj = normalize_adjust_mode(adjust)
    if adj == "tdx_front":
        raise SourceDisabledError(
            "tdx_front 已停用：系统已切换为 Tushare-only 数据策略。"
            "请使用 tushare_qfq 或 raw。"
        )

    session: Optional[BaguaPlaneSession] = None
    if adj in ("raw", "tushare_qfq"):
        try:
            session = BaguaPlaneSession(cfg, adj)
        except FileNotFoundError:
            session = None

    target = query_bagua(
        cfg,
        code=code,
        date=asof,
        period=per,
        adjust=adj,
        session=session,
    )
    if target.get("error") or not target.get("ok", True):
        raise ValueError(f"目标股票卦象查询失败: {target.get('error')}")
    target_state = _bagua_match_target(target).get("state_id") or ""
    if not target_state:
        raise ValueError("目标股票未计算出卦象 state_id")

    use_all = scope is None
    batch = batch_query_bagua(
        cfg,
        codes=scope,
        all_stocks=use_all,
        date=asof,
        period=per,
        adjust=adj,
        limit=None,
        on_progress=on_progress,
        session=session,
    )

    results: List[Dict[str, Any]] = []
    for row in batch.get("results") or []:
        if not row or row.get("error") or row.get("ok") is False:
            continue
        if row.get("std_code") == std:
            continue
        proj = _bagua_match_row(row)
        if proj.get("state_id") != target_state:
            continue
        results.append(proj)
        if limit is not None and len(results) >= int(limit):
            break

    return {
        "ok": True,
        "mode": "bagua",
        "target": _bagua_match_target(target),
        "match_key": target_state,
        "query_date": asof,
        "period": per,
        "adjust": adj,
        "all_stocks": use_all,
        "scanned": batch.get("requested")
        or batch.get("count")
        or len(batch.get("results") or []),
        "count": len(results),
        "results": results,
    }


def find_same_rizhu(
    cfg: AStockConfig,
    *,
    code: str,
    rizhu_path: Optional[Path] = None,
    scope: Optional[Sequence[str]] = None,
    limit: Optional[int] = None,
) -> Dict[str, Any]:
    """Find other stocks sharing the target's static 日柱 (code6 -> 日柱 Excel).

    日柱 is a per-stock static attribute loaded from the Desktop
    ``股票+卦象/日柱(1).xlsx`` map (``load_rizhu_map``). No market data or date
    is involved — the whole universe (or ``scope``) is grouped by that value.
    """
    std = normalize_query_code(code)
    rizhu_map = load_rizhu_map(rizhu_path)
    rizhu_src = _RIZHU_CACHE.get("path") or ""
    target_c6 = _code6_from_any(std)
    target_rizhu = (rizhu_map.get(target_c6) or "").strip()

    disp = display_code(std)
    target_name = resolve_stock_name(cfg, disp, std_code=std)

    if not target_rizhu:
        return {
            "ok": False,
            "mode": "rizhu",
            "error": (
                f"未在日柱表中找到 {disp}（code6={target_c6}）的日柱；"
                "请确认桌面「股票+卦象/日柱(1).xlsx」存在且包含该股票。"
            ),
            "target": {
                "code": disp,
                "name": target_name,
                "display": display_code_with_name(disp, target_name),
                "std_code": std,
                "code6": target_c6,
                "rizhu": "",
            },
            "match_key": "",
            "rizhu_source": rizhu_src,
            "scanned": 0,
            "count": 0,
            "results": [],
        }

    universe = _resolve_batch_codes(cfg, scope, all_stocks=scope is None)

    results: List[Dict[str, Any]] = []
    for raw_code in universe:
        try:
            cur_std = normalize_query_code(raw_code)
        except ValueError:
            continue
        if cur_std == std:
            continue
        cur_c6 = _code6_from_any(cur_std)
        if not cur_c6:
            continue
        if (rizhu_map.get(cur_c6) or "").strip() != target_rizhu:
            continue
        cur_disp = display_code(cur_std)
        name = resolve_stock_name(cfg, cur_disp, std_code=cur_std)
        results.append(
            {
                "ok": True,
                "code": cur_disp,
                "name": name,
                "display": display_code_with_name(cur_disp, name),
                "std_code": cur_std,
                "code6": cur_c6,
                "rizhu": target_rizhu,
            }
        )
        if limit is not None and len(results) >= int(limit):
            break

    return {
        "ok": True,
        "mode": "rizhu",
        "target": {
            "code": disp,
            "name": target_name,
            "display": display_code_with_name(disp, target_name),
            "std_code": std,
            "code6": target_c6,
            "rizhu": target_rizhu,
        },
        "match_key": target_rizhu,
        "rizhu_source": rizhu_src,
        "scanned": len(universe),
        "count": len(results),
        "results": results,
    }


def _yao_prefix_digit(yao_order: Any) -> str:
    """Weekly 变卦 prefix: 1..5 = 初..五, 0 = 上爻 (yao_order 6)."""
    try:
        n = int(yao_order)
    except (TypeError, ValueError):
        return ""
    if n == 6:
        return "0"
    if 1 <= n <= 5:
        return str(n)
    return str(n)


_BIANGUA_FULLNAME_CACHE: Dict[str, Any] = {}


def _load_biangua_fullname_map() -> Dict[str, str]:
    """Build short-name → full_name (with 卦符) lookup from bagua_384.json."""
    global _BIANGUA_FULLNAME_CACHE
    if _BIANGUA_FULLNAME_CACHE.get("map"):
        return _BIANGUA_FULLNAME_CACHE["map"]
    import json

    from ..config import get_default_config

    cfg = get_default_config()
    path = cfg.bagua_json
    out: Dict[str, str] = {}
    if path and Path(path).exists():
        data = json.loads(Path(path).read_text(encoding="utf-8"))
        for e in data.get("entries") or []:
            go = e.get("gua_order")
            fn = str(e.get("full_name") or "").strip()
            gn = str(e.get("gua_name") or "").strip()
            if not fn:
                continue
            if go is not None:
                out[f"__order_{go}"] = fn
            if gn:
                out[gn] = fn
                # short name: 乾为天→乾, 天火同人→同人
                if "为" in gn:
                    out[gn.split("为")[0]] = fn
                else:
                    elements = "天地水火雷风山泽"
                    if len(gn) >= 3 and gn[0] in elements and gn[1] in elements:
                        out[gn[2:]] = fn
    _BIANGUA_FULLNAME_CACHE["map"] = out
    return out


def _resolve_biangua_fullname(b: Dict[str, Any]) -> str:
    """Resolve biangua short name to full name with 卦符 (e.g. 坤→䷁坤为地)."""
    # Prefer changed_hexagram_id → full name
    cid = b.get("changed_hexagram_id")
    if cid is not None:
        m = _load_biangua_fullname_map()
        fn = m.get(f"__order_{cid}")
        if fn:
            return fn
    # Fallback: biangua_full_name field (set by fill_missing_biangua)
    fn = b.get("biangua_full_name") or ""
    if fn:
        return str(fn).strip()
    # Fallback: resolve short name via map
    short = str(b.get("biangua") or b.get("changed_hexagram_name") or "").strip()
    if not short:
        return ""
    m = _load_biangua_fullname_map()
    return m.get(short, short)


def _strip_gua_symbols(s: Any) -> str:
    """去掉卦象符号字符（U+4DC0–U+4DFF），导出表不再包含卦符。"""
    if not s:
        return ""
    return re.sub(r"[\u4dc0-\u4dff]", "", str(s)).strip()


def _bagua_combo(row: Optional[Dict[str, Any]]) -> str:
    """Build ``本卦|N-变卦`` like weekly_analysis 组合 column (no 卦符)."""
    if not row or row.get("error") or not row.get("ok", True):
        return ""
    s = row.get("summary") or {}
    b = row.get("bagua") or {}
    ben = (
        s.get("full_name")
        or b.get("full_name")
        or b.get("gua_name")
        or b.get("main_hexagram_name")
        or ""
    )
    ben = str(ben).strip()
    bian = _resolve_biangua_fullname(b)
    yo = s.get("yao_order")
    if yo is None:
        yo = b.get("yao_order")
    prefix = _yao_prefix_digit(yo)
    if not ben and not bian:
        return ""
    if bian and prefix != "":
        right = f"{prefix}-{bian}"
    elif bian:
        right = bian
    else:
        right = ""
    if ben and right:
        out = f"{ben}|{right}"
    else:
        out = ben or right
    return _strip_gua_symbols(out)


def _bagua_yao_explain(row: Optional[Dict[str, Any]]) -> str:
    if not row or row.get("error") or not row.get("ok", True):
        return ""
    s = row.get("summary") or {}
    b = row.get("bagua") or {}
    return _strip_gua_symbols(
        str(
            s.get("market_judgement")
            or b.get("market_judgement")
            or b.get("market_summary")
            or b.get("yao_ci")
            or b.get("line_text")
            or ""
        )
    )


def _fmt_ymd_dash(ymd: Any) -> str:
    try:
        n = int(ymd)
    except (TypeError, ValueError):
        s = str(ymd or "").strip()
        if len(s) == 8 and s.isdigit():
            n = int(s)
        else:
            return s
    return f"{n // 10000:04d}-{(n // 100) % 100:02d}-{n % 100:02d}"


def _week_iso_label(ymd: Any) -> str:
    """bar 结束日 -> ISO 周标签，如 2026-08-14 -> '2026-W33'；解析失败返回 ''。"""
    s = _fmt_ymd_dash(ymd)
    try:
        y, m, d = (int(p) for p in s.split("-"))
        iso = _ymd_date(y, m, d).isocalendar()
    except Exception:
        return ""
    return f"{iso.year}-W{iso.week:02d}"


def _month_label(ymd: Any) -> str:
    """bar 结束日 -> 月份标签，如 2026-07-31 -> '2026-07'；解析失败返回 ''。"""
    s = _fmt_ymd_dash(ymd)
    try:
        y, m, _d = (int(p) for p in s.split("-"))
    except Exception:
        return ""
    return f"{y}-{m:02d}"


def _code6_from_any(code: Any) -> str:
    s = str(code or "").strip().upper()
    if not s:
        return ""
    # 000001.SZ / SH600000 / SSE.STK.600000 / sh600000 / 600000
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) >= 6:
        return digits[-6:]
    return digits


_RIZHU_CACHE: Dict[str, Any] = {"path": None, "mtime": None, "map": {}}


def _rizhu_disk_cache_path() -> Path:
    """Persistent json cache for the 28MB 日柱 Excel (cold parse ~20s)."""
    return Path(__file__).resolve().parents[4] / "storage" / "astock" / "rizhu_cache.json"


def _default_rizhu_path() -> Optional[Path]:
    """Locate 日柱(1).xlsx under Desktop/股票+卦象 (or common fallbacks)."""
    home = Path.home()
    candidates = [
        home / "Desktop" / "股票+卦象",
        home / "桌面" / "股票+卦象",
        Path(r"C:\Users") / Path.home().name / "Desktop" / "股票+卦象",
    ]
    for folder in candidates:
        if not folder.exists():
            continue
        # Prefer name containing 日柱
        named = sorted(
            [p for p in folder.glob("*.xlsx") if "日柱" in p.name],
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if named:
            return named[0]
        # Fallback: largest xlsx in folder (日柱 table is ~30MB)
        big = sorted(
            folder.glob("*.xlsx"),
            key=lambda p: p.stat().st_size,
            reverse=True,
        )
        if big and big[0].stat().st_size > 5_000_000:
            return big[0]
    return None


def _read_rizhu_excel(p: Path) -> Dict[str, str]:
    """Parse code6 -> 日柱 from the Excel workbook (总表/SH/SZ/BJ)."""
    import openpyxl

    out: Dict[str, str] = {}
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            it = ws.iter_rows(values_only=True)
            headers = next(it, None)
            if not headers:
                continue
            hlist = [str(h).strip() if h is not None else "" for h in headers]
            # Find code + 日柱 columns
            code_i = None
            rizhu_i = None
            for i, h in enumerate(hlist):
                if not h:
                    continue
                if code_i is None and (
                    h in ("股票代码", "code", "证券代码", "代码") or "代码" in h
                ):
                    code_i = i
                if rizhu_i is None and (h == "日柱" or h.endswith("日柱")):
                    rizhu_i = i
            if code_i is None:
                continue
            # If no explicit 日柱 header, try common position (col index 8 in sample)
            if rizhu_i is None:
                for i, h in enumerate(hlist):
                    if h and ("日柱" in h or h in ("日干支", "干支")):
                        rizhu_i = i
                        break
            if rizhu_i is None and len(hlist) > 8:
                rizhu_i = 8
            if rizhu_i is None:
                continue
            for row in it:
                if not row or code_i >= len(row):
                    continue
                c6 = _code6_from_any(row[code_i])
                if not c6:
                    continue
                val = row[rizhu_i] if rizhu_i < len(row) else None
                if val is None or str(val).strip() == "":
                    continue
                # Prefer first non-empty; 总表 first usually
                if c6 not in out:
                    out[c6] = str(val).strip()
    finally:
        wb.close()
    return out


def load_rizhu_map(path: Optional[Path] = None) -> Dict[str, str]:
    """Load code6 -> 日柱 string from 日柱 Excel (总表/SH/SZ/BJ).

    The workbook is ~28MB and cold parses in ~20s, so the result is cached
    in memory (per process) and persisted to a json file under storage/ so a
    backend restart does not re-parse it.
    """
    import json

    p = Path(path) if path else _default_rizhu_path()
    if p is None or not p.exists():
        return {}
    try:
        mtime = p.stat().st_mtime
    except OSError:
        return {}
    if (
        _RIZHU_CACHE.get("path") == str(p)
        and _RIZHU_CACHE.get("mtime") == mtime
        and _RIZHU_CACHE.get("map")
    ):
        return dict(_RIZHU_CACHE["map"])

    disk = _rizhu_disk_cache_path()
    if str(p) != _RIZHU_CACHE.get("path") or _RIZHU_CACHE.get("mtime") != mtime:
        # try disk cache first
        try:
            if disk.exists():
                j = json.loads(disk.read_text(encoding="utf-8"))
                if (
                    isinstance(j, dict)
                    and j.get("path") == str(p)
                    and j.get("mtime") == mtime
                    and isinstance(j.get("map"), dict)
                ):
                    _RIZHU_CACHE["path"] = str(p)
                    _RIZHU_CACHE["mtime"] = mtime
                    _RIZHU_CACHE["map"] = j["map"]
                    return dict(j["map"])
        except Exception:
            pass

    out = _read_rizhu_excel(p)

    _RIZHU_CACHE["path"] = str(p)
    _RIZHU_CACHE["mtime"] = mtime
    _RIZHU_CACHE["map"] = out
    # persist for future processes
    try:
        disk.parent.mkdir(parents=True, exist_ok=True)
        disk.write_text(
            json.dumps(
                {"path": str(p), "mtime": mtime, "map": out},
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
    except Exception:
        pass
    return dict(out)


# ---- 日柱补齐：Excel 表外的次新股 / ETF 按上市日期推算 60 甲子 ----

_GAN = "甲乙丙丁戊己庚辛壬癸"
_ZHI = "子丑寅卯辰巳午未申酉戌亥"
_RIZHU_ANCHOR = _ymd_date(1899, 12, 22)  # 甲子日


def _rizhu_from_list_date(ymd: int) -> str:
    """上市日期 → 日干支（60 甲子），口径与 日柱(1).xlsx 一致（上市日当天干支）。"""
    d = _ymd_date(ymd // 10000, (ymd // 100) % 100, ymd % 100)
    n = (d - _RIZHU_ANCHOR).days % 60
    return _GAN[n % 10] + _ZHI[n % 12]


def _rizhu_list_dates_cache_path() -> Path:
    return Path(__file__).resolve().parents[4] / "storage" / "astock" / "rizhu_list_dates.json"


_LIST_DATES_CACHE: Dict[str, Any] = {}


def _load_list_date_cache() -> Tuple[Dict[str, int], Dict[str, int]]:
    """读 code6 -> list_date 的股票/ETF 缓存（rizhu_list_dates.json）。"""
    cached = _LIST_DATES_CACHE.get("data")
    if cached is not None:
        return cached[0], cached[1]
    import json as _json

    try:
        j = _json.loads(_rizhu_list_dates_cache_path().read_text(encoding="utf-8"))
        stocks = {str(k): int(v) for k, v in j.get("stocks", {}).items()}
        etfs = {str(k): int(v) for k, v in j.get("etfs", {}).items()}
        _LIST_DATES_CACHE["data"] = (stocks, etfs)
        return stocks, etfs
    except Exception:
        return {}, {}


def _save_list_date_cache(stocks: Dict[str, int], etfs: Dict[str, int]) -> None:
    import json as _json

    try:
        p = _rizhu_list_dates_cache_path()
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(
            _json.dumps(
                {
                    "schema_version": 1,
                    "fetched_at": _bq_time.strftime("%Y-%m-%d"),
                    "stocks": stocks,
                    "etfs": etfs,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        _LIST_DATES_CACHE["data"] = (stocks, etfs)
    except Exception:
        pass


def _code6_from_entry(entry: Any, kind: Optional[str] = None) -> Optional[str]:
    """从 UniverseEntry.symbol（SSE.STK.600000 / SSE.ETF.510300）取 6 位数字码。

    股票/ETF 与指数的代码段存在重叠（000001 同时是上证指数与平安银行），
    因此按 kind 过滤，绝不把指数当作股票/ETF。
    """
    sym = str(getattr(entry, "symbol", "") or "")
    parts = sym.split(".")
    if len(parts) != 3:
        return None
    if kind and parts[1] != kind:
        return None
    code = parts[2]
    return code if code.isdigit() and len(code) == 6 else None


def _fetch_list_dates_from_tushare(
    cfg: AStockConfig,
) -> Tuple[Dict[str, int], Dict[str, int]]:
    """全量拉取 A 股与 ETF 的上市日期（stock_basic / fund_basic）。

    复用 TushareProvider 的限流/重试/类型化异常；token 取环境变量，
    未配置时由 provider 回退 ts.get_token()。指数被 _code6_from_entry 过滤。
    """
    import os as _os

    from ..data.providers.base import ProviderUnavailable
    from ..data.providers.tushare import TushareProvider

    token = _os.environ.get("TUSHARE_TOKEN") or _os.environ.get("TS_TOKEN") or None
    try:
        provider = TushareProvider(token=token)
        stocks: Dict[str, int] = {}
        etfs: Dict[str, int] = {}
        for e in provider.fetch_universe():
            c6 = _code6_from_entry(e, "STK")
            if c6 and getattr(e, "list_date", None):
                stocks[c6] = int(e.list_date)
        for e in provider.fetch_index_etf_universe():
            c6 = _code6_from_entry(e, "ETF")
            if c6 and getattr(e, "list_date", None):
                etfs[c6] = int(e.list_date)
        return stocks, etfs
    except Exception as exc:
        raise ProviderUnavailable(f"tushare list-date fetch failed: {exc}") from exc


def ensure_rizhu_coverage(
    cfg: AStockConfig,
    needed_codes: Sequence[str],
    existing: Optional[Dict[str, str]] = None,
) -> Dict[str, str]:
    """用上市日期推算补齐 code6 -> 日柱（Excel 表外的次新股 / ETF）。

    只处理 needed_codes 中尚未有日柱的代码（existing 里的 code6 跳过）；
    优先命中本地上市日期缓存，未命中才调 Tushare 全量刷新并落盘；
    Tushare 不可用时静默降级返回 {}，导出结果与不补齐时一致，
    不因网络问题中断导出。
    """
    base = existing or {}
    missing: set = set()
    for c in needed_codes:
        c6 = _code6_from_any(c)
        if c6 and c6 not in base:
            missing.add(c6)
    if not missing:
        return {}

    cached_stocks, cached_etfs = _load_list_date_cache()
    known = dict(cached_stocks)
    known.update(cached_etfs)

    out: Dict[str, str] = {}
    still_missing: set = set()
    for c6 in sorted(missing):
        ld = known.get(c6)
        if ld:
            out[c6] = _rizhu_from_list_date(ld)
        else:
            still_missing.add(c6)
    if not still_missing:
        return out

    try:
        stocks, etfs = _fetch_list_dates_from_tushare(cfg)
    except Exception:
        return out
    merged_stocks = {**cached_stocks, **stocks}
    merged_etfs = {**cached_etfs, **etfs}
    _save_list_date_cache(merged_stocks, merged_etfs)
    for c6 in still_missing:
        ld = merged_stocks.get(c6) or merged_etfs.get(c6)
        if ld:
            out[c6] = _rizhu_from_list_date(ld)
    return out


def _weekly_style_row(
    *,
    week_row: Optional[Dict[str, Any]],
    month_row: Optional[Dict[str, Any]],
    rizhu: str = "",
    fallback_code: str = "",
) -> List[Any]:
    """One stock row in weekly_analysis stock-all layout (周卦列在前、月卦列在后)."""
    base = week_row if (week_row and not week_row.get("error")) else month_row
    bar = (week_row or {}).get("bar") or (base or {}).get("bar") or {}
    code = (
        (week_row or {}).get("code")
        or (month_row or {}).get("code")
        or fallback_code
        or ""
    )
    # Prefer 6-digit code like weekly file
    c6 = _code6_from_any(code) or str(code).replace("sh", "").replace("sz", "").replace("bj", "")
    name = (
        (week_row or {}).get("name")
        or (month_row or {}).get("name")
        or ""
    )
    week_end = bar.get("end_date") or bar.get("date") or (week_row or {}).get("query_date")
    return [
        c6,
        name,
        _fmt_ymd_dash(week_end) if week_end else "",
        bar.get("open") if bar else "",
        bar.get("high") if bar else "",
        bar.get("low") if bar else "",
        bar.get("close") if bar else "",
        rizhu or "",
        _bagua_combo(week_row),
        _bagua_yao_explain(week_row),
        _bagua_combo(month_row),
        _bagua_yao_explain(month_row),
    ]


def export_bagua_xlsx(
    cfg: AStockConfig,
    *,
    date: Union[str, int],
    period: str = "DAY",
    adjust: str = "tushare_qfq",
    codes: Optional[Sequence[str]] = None,
    all_stocks: bool = True,
    limit: Optional[int] = None,
    path: Optional[Path] = None,
) -> Path:
    """Export multi-stock / full-market hexagrams to xlsx (single period).

    Default is full-market (``all_stocks=True``). Returns the written path.
    """
    return export_bagua_multi_period_xlsx(
        cfg,
        date=date,
        periods=[period],
        adjust=adjust,
        codes=codes,
        all_stocks=all_stocks,
        limit=limit,
        path=path,
    )


def _query_bagua_periods_for_code(
    cfg: AStockConfig,
    *,
    code: str,
    asof: int,
    periods: Sequence[str],
    adjust: str,
    session: Optional["BaguaPlaneSession"] = None,
    calc: Optional[BaguaCalculator] = None,
    asof_map: Optional[Dict[str, int]] = None,
) -> Dict[str, Dict[str, Any]]:
    """Compute bagua for one stock across multiple periods (one bar load)."""
    out: Dict[str, Dict[str, Any]] = {}
    # First period loads bars via query_bagua; subsequent reuse same session/calc.
    # Still one warehouse load per period internally — but shared session avoids
    # re-indexing. For true single-load, call query once per period with session.
    for per in periods:
        per_asof = asof_map.get(per, asof) if asof_map else asof
        try:
            out[per] = query_bagua(
                cfg,
                code=code,
                date=per_asof,
                period=per,
                adjust=adjust,
                session=session,
                calc=calc,
            )
        except Exception as e:
            try:
                std = normalize_query_code(code)
                disp = display_code(std)
            except Exception:
                std = ""
                disp = str(code)
            name = ""
            try:
                name = resolve_stock_name(cfg, disp, std_code=std or None) or ""
            except Exception:
                pass
            out[per] = {
                "ok": False,
                "code": disp,
                "name": name,
                "display": display_code_with_name(disp, name) if name else disp,
                "std_code": std,
                "symbol_type": classify_symbol(code),
                "query_date": per_asof,
                "period": per,
                "adjust": adjust,
                "error": str(e),
            }
    return out


def _display_width(value: Any) -> float:
    """Approx cell width: CJK / full-width chars count as 2 units."""
    s = "" if value is None else str(value)
    w = 0.0
    for ch in s:
        code = ord(ch)
        if (
            0x4E00 <= code <= 0x9FFF
            or 0x3000 <= code <= 0x303F
            or 0xFF00 <= code <= 0xFFEF
            or 0x2E80 <= code <= 0x4DBF
        ):
            w += 2.0
        else:
            w += 1.0
    return w


def _autofit_columns(ws: Any, *, max_width: float = 90.0, min_width: float = 8.0) -> None:
    """Set column widths to fit content (combo columns get generous room)."""
    from openpyxl.utils import get_column_letter

    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        best = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0] if row else None
            w = _display_width(val) + 2
            if w > best:
                best = w
        ws.column_dimensions[letter].width = min(max(best, min_width), max_width)


def _export_sheet_rows(
    cfg: AStockConfig,
    *,
    pool: Sequence[str],
    asof: int,
    query_pers: Sequence[str],
    adjust: str,
    session: Optional["BaguaPlaneSession"],
    calc: BaguaCalculator,
    asof_map: Dict[str, int],
    rizhu_map: Dict[str, str],
    on_progress: Optional[Any],
    base_idx: int,
    total: int,
    totals: Dict[str, int],
    resolve_names: bool = False,
) -> Tuple[List[List[Any]], Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    """Query bagua for one code pool and build weekly-style sheet rows.

    Returns (rows, first_week_row, first_month_row). ``base_idx`` offsets this
    pool inside the combined progress bar and ``total`` is the combined pool
    size. ``resolve_names`` fills empty name cells via ``resolve_stock_name``
    (used for ETF rows, where display names are sparse).
    """
    sheet_rows: List[List[Any]] = []
    first_week_row: Optional[Dict[str, Any]] = None
    first_month_row: Optional[Dict[str, Any]] = None
    for k, raw_code in enumerate(pool, 1):
        idx = base_idx + k
        per_rows = _query_bagua_periods_for_code(
            cfg,
            code=raw_code,
            asof=asof,
            periods=query_pers,
            adjust=adjust,
            session=session,
            calc=calc,
            asof_map=asof_map,
        )
        week_row = per_rows.get("WEEK")
        month_row = per_rows.get("MONTH")
        if first_week_row is None and week_row and not week_row.get("error"):
            first_week_row = week_row
        if first_month_row is None and month_row and not month_row.get("error"):
            first_month_row = month_row
        c6 = _code6_from_any(
            (week_row or {}).get("code")
            or (month_row or {}).get("code")
            or raw_code
        )
        rizhu = rizhu_map.get(c6, "") if c6 else ""
        if rizhu:
            totals["rizhu_hit"] += 1
        ok_w = bool(week_row and week_row.get("ok") and not week_row.get("error"))
        ok_m = bool(month_row and month_row.get("ok") and not month_row.get("error"))
        if ok_w or ok_m:
            totals["ok"] += 1
        else:
            totals["error"] += 1
        row = _weekly_style_row(
            week_row=week_row,
            month_row=month_row,
            rizhu=rizhu,
            fallback_code=raw_code,
        )
        if resolve_names and c6 and not row[1] and ".ETF." in str(raw_code):
            try:
                row[1] = resolve_stock_name(cfg, c6) or ""
            except Exception:
                pass
        sheet_rows.append(row)
        if on_progress is not None and (idx == total or idx % 25 == 0 or idx == 1):
            try:
                on_progress(
                    {
                        "phase": "query",
                        "period": "WEEK+MONTH",
                        "period_index": 1,
                        "period_count": 1,
                        "done": idx,
                        "total": total,
                        "ok_count": totals["ok"],
                        "error_count": totals["error"],
                        "code": raw_code,
                    }
                )
            except Exception:
                pass
    return sheet_rows, first_week_row, first_month_row


def export_bagua_multi_period_xlsx(
    cfg: AStockConfig,
    *,
    date: Union[str, int],
    periods: Optional[Sequence[str]] = None,
    adjust: str = "tushare_qfq",
    codes: Optional[Sequence[str]] = None,
    all_stocks: bool = True,
    limit: Optional[int] = None,
    path: Optional[Path] = None,
    on_progress: Optional[Any] = None,
    rizhu_path: Optional[Path] = None,
) -> Path:
    """Export bagua in weekly_analysis stock-all layout, one sheet per pool.

    Full-market export (``all_stocks=True``) writes two sheets into the same
    workbook:
      - ``stock-all``: A-share universe rows
      - ``etf-all``   : every ETF enumerated from the TDX local day files
    Manual ``codes`` are split by symbol type — stocks stay in ``stock-all``,
    index/ETF codes go to ``etf-all``.

    ``limit`` is a total row cap applied as stock-first: stocks fill the cap
    before ETFs are included (a cap ≤ stock count yields no etf-all sheet).
    Manual codes that cannot be recognized are silently dropped.

    Columns:
      code, name, week_end, open, high, low, close, 日柱,
      周卦周线-组合(周标签), 爻辞解释, 月卦月线-组合(月标签), 爻辞解释

    周卦在前、月卦在后；表头标注周卦所在周（ISO 周）与月卦所在月份。
    月卦默认取查询月份的上一个月（如8月查询导出7月月卦，避免未收官月卦），
    周卦取查询日期所在周。Always computes WEEK + MONTH (DAY is ignored for
    this layout). 日柱 is joined from Desktop ``股票+卦象/日柱(1).xlsx`` when
    available. 导出卦象组合已去除卦符字符（U+4DC0–U+4DFF）。
    """
    import time

    import openpyxl
    from openpyxl.styles import Font

    # Layout always needs week + month combos; keep periods only for meta.
    raw_pers = list(periods or ["WEEK", "MONTH"])
    pers: List[str] = []
    seen_p: set = set()
    for p in raw_pers:
        try:
            np = normalize_period(p)
        except ValueError:
            continue
        if np not in seen_p:
            seen_p.add(np)
            pers.append(np)
    # Force week+month for weekly-style sheet
    need = ["WEEK", "MONTH"]
    for n in need:
        if n not in seen_p:
            pers.append(n)
            seen_p.add(n)

    asof = _parse_ymd(date)
    month_asof = _prev_month_end(asof)
    adj = normalize_adjust_mode(adjust)
    use_all = bool(all_stocks)
    stock_pool: List[str] = []
    etf_pool: List[str] = []
    if use_all:
        stock_pool = _resolve_batch_codes(cfg, None, all_stocks=True)
        etf_pool = _enumerate_export_etf_pool(cfg)
    else:
        stock_raw: List[str] = []
        etf_raw: List[str] = []
        for c in (codes or []):
            c = str(c).strip()
            if not c:
                continue
            if classify_symbol(c) in ("index", "etf"):
                etf_raw.append(c)
            else:
                try:
                    normalize_query_code(c)
                except ValueError:
                    continue  # 无法识别的代码静默丢弃（与全量列表行为一致）
                stock_raw.append(c)
        if stock_raw:
            stock_pool = _resolve_batch_codes(cfg, stock_raw, all_stocks=False)
        seen_etf: set = set()
        for c in etf_raw:
            std = to_index_etf_std_code(c)
            if std and std not in seen_etf:
                seen_etf.add(std)
                etf_pool.append(std)
        if not stock_pool and not etf_pool:
            raise ValueError("codes or all_stocks required")
    if limit is not None:
        lim = int(limit)
        stock_pool = stock_pool[:lim]
        etf_pool = etf_pool[: max(0, lim - len(stock_pool))]
    if not stock_pool and not etf_pool:
        raise ValueError("codes or all_stocks required")

    if not cfg.bagua_json:
        raise FileNotFoundError("bagua knowledge json not configured")
    calc = BaguaCalculator.from_json(cfg.bagua_json)
    session: Optional[BaguaPlaneSession] = None
    if adj in ("tdx_front", "tushare_qfq", "raw"):
        if adj == "tdx_front":
            raise SourceDisabledError(
                "tdx_front 已停用：系统已切换为 Tushare-only 数据策略。"
                "请使用 tushare_qfq 或 raw。"
            )
        try:
            session = BaguaPlaneSession(cfg, adj)
        except FileNotFoundError:
            session = None

    rizhu_map = load_rizhu_map(rizhu_path)
    rizhu_src = _RIZHU_CACHE.get("path") or ""
    # Excel 表外的次新股 / ETF 按上市日期推算补齐日柱，保证导出列完整
    # （ensure 在前，Excel 表值优先，不被推算结果覆盖）
    rizhu_map = {
        **ensure_rizhu_coverage(cfg, [*stock_pool, *etf_pool], rizhu_map),
        **rizhu_map,
    }

    export_root = Path(cfg.storage_root) / "bagua_exports"
    export_root.mkdir(parents=True, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    scope = "all" if use_all else "batch"
    out = Path(
        path
        or (export_root / f"bagua_weekly_{scope}_{asof}_{adj}_{stamp}.xlsx")
    )
    out.parent.mkdir(parents=True, exist_ok=True)

    pools: List[Tuple[str, List[str], bool]] = []
    if stock_pool:
        pools.append(("stock-all", stock_pool, False))
    if etf_pool:
        pools.append(("etf-all", etf_pool, True))

    totals = {
        "requested": len(stock_pool) + len(etf_pool),
        "ok": 0,
        "error": 0,
        "rizhu_hit": 0,
    }
    total = totals["requested"]
    query_pers = ["WEEK", "MONTH"]
    asof_map = {"WEEK": asof, "MONTH": month_asof}
    sheet_rows_by_name: Dict[str, List[List[Any]]] = {}
    first_rows: Dict[str, Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]] = {}
    base_idx = 0
    for sheet_name, pool, resolve_names in pools:
        rows, fw, fm = _export_sheet_rows(
            cfg,
            pool=pool,
            asof=asof,
            query_pers=query_pers,
            adjust=adj,
            session=session,
            calc=calc,
            asof_map=asof_map,
            rizhu_map=rizhu_map,
            on_progress=on_progress,
            base_idx=base_idx,
            total=total,
            totals=totals,
            resolve_names=resolve_names,
        )
        base_idx += len(pool)
        sheet_rows_by_name[sheet_name] = rows
        first_rows[sheet_name] = (fw, fm)

    wb = openpyxl.Workbook()
    from openpyxl.utils import get_column_letter

    for si, (sheet_name, _pool, _rn) in enumerate(pools):
        rows = sheet_rows_by_name[sheet_name]
        fw, fm = first_rows[sheet_name]
        ws = wb.active if si == 0 else wb.create_sheet(sheet_name)
        ws.title = sheet_name
        week_label = _week_iso_label(
            ((fw or {}).get("bar") or {}).get("end_date") or asof
        )
        month_label = _month_label(
            ((fm or {}).get("bar") or {}).get("end_date") or month_asof
        )
        headers = [
            "code",
            "name",
            "week_end",
            "open",
            "high",
            "low",
            "close",
            "日柱",
            f"周卦周线-组合({week_label})",
            "爻辞解释",
            f"月卦月线-组合({month_label})",
            "爻辞解释",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append(row)
        _autofit_columns(ws)
        # combo columns (周卦周线-组合 / 月卦月线-组合) need extra room
        for i, hdr in enumerate(headers, 1):
            if "组合" in hdr:
                letter = get_column_letter(i)
                cur = ws.column_dimensions[letter].width or 0
                ws.column_dimensions[letter].width = max(cur, 32)

    meta = wb.create_sheet("meta", 0)
    meta.append(["key", "value"])
    for cell in meta[1]:
        cell.font = Font(bold=True)
    for k, v in [
        ("layout", "weekly_analysis stock-all"),
        ("query_date", asof),
        ("month_asof", month_asof),
        ("note", "MONTH 列取查询月份的上一个月（如8月查询导出7月月卦）；WEEK 列取查询日期所在周"),
        ("periods", "WEEK,MONTH"),
        ("adjust", adj),
        ("all_stocks", use_all),
        ("requested", totals["requested"]),
        ("stock_count", len(stock_pool)),
        ("etf_count", len(etf_pool)),
        ("sheets", ",".join(name for name, _p, _r in pools)),
        ("ok_total", totals["ok"]),
        ("error_total", totals["error"]),
        ("rizhu_hit", totals["rizhu_hit"]),
        ("rizhu_source", rizhu_src),
        ("rizhu_note", "Excel 日柱表优先；次新股/ETF 按上市日期推算 60 甲子补齐"),
        ("exported_at", stamp),
    ]:
        meta.append([k, v])

    wb.save(out)
    return out
