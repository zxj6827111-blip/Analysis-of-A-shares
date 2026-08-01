# -*- coding: utf-8 -*-
"""Export one or more backtest configurations as a task-record workbook."""

from __future__ import annotations

import math
import re
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from ..config import AStockConfig
from .runs import load_run_summary


_WEEKDAY_LABELS = {
    1: "周一",
    2: "周二",
    3: "周三",
    4: "周四",
    5: "周五",
    6: "周六",
    7: "周日",
}

_SESSION_LABELS = {
    "open": "开盘",
    "close": "收盘",
}

_STATUS_LABELS = {
    "ok": "完成",
    "succeeded": "完成",
    "success": "完成",
    "done": "完成",
    "completed": "完成",
    "research_unadjusted": "完成 · 未复权",
    "research_unconfirmed_formula": "完成 · 公式待确认",
    "unsupported_corporate_action": "未通过（公司行为不支持）",
    "no_go": "未通过",
    "failed": "失败",
    "cancelled": "已取消",
}

_SAFE_RUN_ID = re.compile(r"^[A-Za-z0-9_.-]+$")

# 状态判定：下列状态视为"已完成、结果可用"，参与推荐排名
_COMPLETED_STATUSES = {
    "ok",
    "succeeded",
    "success",
    "done",
    "completed",
    "research_unadjusted",
    "research_unconfirmed_formula",
}

# 交易样本阈值只用于导出提示，不等同于统计显著性检验。
_MIN_OBSERVATION_TRADES = 30
_BASE_OBSERVATION_TRADES = 100
_MAX_BEST_DRAWDOWN = 0.25
_BEST_ELIGIBLE_TIERS = {"silver", "bronze", "caution", "average"}

# 复验优先级 → 行底色（仅对较高优先级着色，其余保持默认）
_TIER_FILL = {
    "silver": "D9EAD3",   # 优先复验：浅绿
    "bronze": "E8F5E9",   # 建议复验：淡绿
    "caution": "FFF2CC",  # 观察：浅黄
}

_BEST_ROW_FILL = "FFF200"       # 本批次最优：鲜黄色整行
_BEST_LABEL_FILL = "FF0000"     # 最优标签：鲜红底色
_BEST_RED_FONT = "FFFF0000"     # 摘要强调：红色字体（8位 ARGB）
_BEST_WHITE_FONT = "FFFFFFFF"   # 最优标签：白色字体（8位 ARGB）
_BEST_BORDER_COLOR = "FF0000"
_SUMMARY_COLUMN_WIDTH = 64


def _to_float(value: Any) -> Optional[float]:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_int(value: Any) -> Optional[int]:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def normalize_export_run_ids(run_ids: Iterable[str], *, max_count: int = 200) -> List[str]:
    """Validate and de-duplicate run IDs while preserving request order."""
    normalized: List[str] = []
    seen = set()
    for raw in run_ids:
        run_id = str(raw or "").strip()
        if not run_id or run_id in seen:
            continue
        if Path(run_id).name != run_id or not _SAFE_RUN_ID.fullmatch(run_id):
            raise ValueError(f"invalid run_id: {run_id!r}")
        normalized.append(run_id)
        seen.add(run_id)
    if not normalized:
        raise ValueError("at least one run_id is required")
    if len(normalized) > max_count:
        raise ValueError(f"at most {max_count} run_ids are allowed")
    return normalized


def _first_value(summary: dict, key: str, default: Any = None) -> Any:
    meta = summary.get("meta") if isinstance(summary.get("meta"), dict) else {}
    repro = summary.get("repro") if isinstance(summary.get("repro"), dict) else {}
    request = repro.get("request") if isinstance(repro.get("request"), dict) else {}
    for source in (summary, meta, repro, request):
        value = source.get(key)
        if value not in (None, "", []):
            return value
    return default


def _weekday_label(value: Any) -> str:
    try:
        return _WEEKDAY_LABELS.get(int(value), str(value))
    except (TypeError, ValueError):
        return ""


def _session_label(value: Any) -> str:
    raw = str(value or "").strip().lower()
    return _SESSION_LABELS.get(raw, raw or "")


def _schedule_text(summary: dict, *, side: str) -> str:
    schedule_mode = str(_first_value(summary, "schedule_mode", "") or "").lower()
    if side == "buy":
        weekday = _first_value(summary, "buy_weekday")
        if weekday is not None:
            return _weekday_label(weekday)
        lag = _first_value(summary, "entry_lag")
        return f"T+{lag}" if lag is not None else ""

    weekday = _first_value(summary, "exit_weekday")
    if weekday is not None:
        return _weekday_label(weekday)
    hold = _first_value(summary, "hold")
    if hold is None:
        return ""
    if schedule_mode == "weekday":
        return f"持有{hold}期"
    return f"持有{hold}日"


def _condition_text(summary: dict) -> str:
    names = _first_value(summary, "indicator_names")
    if not names:
        names = _first_value(summary, "indicator_ids")
    if isinstance(names, str):
        names = [names]
    cleaned = []
    for raw in names or []:
        text = str(raw or "").strip()
        for prefix in ("tn6_", "txt_", "user_"):
            if text.startswith(prefix):
                text = text[len(prefix) :]
        if text:
            cleaned.append(text)
    if cleaned:
        return " + ".join(cleaned)
    return str(_first_value(summary, "title", summary.get("run_id") or "回测任务"))


def _gua_text(summary: dict) -> Any:
    enabled = bool(_first_value(summary, "with_bagua", False))
    gua_filter = _first_value(summary, "gua_filter")
    if isinstance(gua_filter, dict):
        enabled = bool(gua_filter.get("enabled", enabled))
    if not enabled:
        return "未启用"

    if isinstance(gua_filter, dict):
        history = (
            gua_filter.get("history_summary")
            if isinstance(gua_filter.get("history_summary"), dict)
            else {}
        )
        tooltip_lines = history.get("tooltip_lines")
        if isinstance(tooltip_lines, list):
            labels = [str(x).strip() for x in tooltip_lines if str(x).strip()]
            if labels:
                return "、".join(labels)
        short = str(history.get("short") or "").strip()
        if short:
            return short
    label = _first_value(summary, "bagua_filter_label")
    if label:
        return str(label)
    if isinstance(gua_filter, dict):
        for key in ("natural_language", "label", "gua_label"):
            if gua_filter.get(key):
                return str(gua_filter[key])
        state_ids = gua_filter.get("selected_state_ids")
        if isinstance(state_ids, list) and state_ids:
            return "、".join(str(x) for x in state_ids)
    return "已启用"


def _format_percent(value: Any) -> Optional[str]:
    try:
        return f"{float(value):.2%}"
    except (TypeError, ValueError):
        return None


def _verdict(
    total: Optional[float],
    dd_abs: Optional[float],
    win: Optional[float],
    trades: Optional[int],
    completed: bool,
) -> str:
    """给出一句审慎的复验判断，避免把回测结果表述为投资建议。"""
    if not completed:
        return "该任务未通过有效性校验，本结果不进入复验排序。"
    if trades is None:
        return "缺少完成交易数，当前结果暂不评级。"
    if trades < _MIN_OBSERVATION_TRADES:
        return "当前交易样本过少，暂不评级；建议扩大样本后重新评估。"
    if total is None:
        return "缺少收益数据，当前结果无法评估。"
    if total <= 0:
        return "本回测区间收益为负，暂缓进入下一轮复验。"
    if dd_abs is None:
        return "缺少最大回撤数据，建议补齐风险指标后再评级。"
    if dd_abs > _MAX_BEST_DRAWDOWN:
        return "当前最大回撤较高，建议先进行仓位、止损和持有期敏感性测试。"
    if trades < _BASE_OBSERVATION_TRADES:
        return "当前收益为正，但交易样本偏少；建议累计至100笔后复评。"
    if win is not None and win < 0.48:
        return "当前收益为正，但胜率偏低；需结合盈亏比、利润因子和收益分布判断稳定性。"
    if total < 0.05:
        return "收益为正但幅度较低，建议继续与基准收益和交易成本比较。"
    if total < 0.15:
        return "收益为正、回撤相对较低；建议继续观察并完成样本外验证。"
    if total < 0.30:
        return "收益表现尚可；建议进入样本外和成本敏感性验证。"
    return "收益表现较强；建议进入下一轮样本外、成本敏感性与稳健性验证。"


def _sample_text(trades: Optional[int]) -> str:
    if trades is None:
        return "完成交易数未知，无法判断样本可信度"
    if trades < _MIN_OBSERVATION_TRADES:
        return (
            f"完成交易 {trades} 笔，样本过少"
            f"（未达到 {_MIN_OBSERVATION_TRADES} 笔最低观察阈值；暂不评级）"
        )
    if trades < _BASE_OBSERVATION_TRADES:
        return (
            f"完成交易 {trades} 笔，样本偏少"
            f"（建议累计至 {_BASE_OBSERVATION_TRADES} 笔后复评）"
        )
    return f"完成交易 {trades} 笔，达到基础观察阈值"


def _conclusion_text(
    summary: dict,
    total: Optional[float],
    dd_abs: Optional[float],
    win: Optional[float],
    trades: Optional[int],
    completed: bool,
) -> str:
    """按状态、绩效、样本和判断四行输出可扫描的回测摘要。"""
    status = str(summary.get("status") or "ok").strip().lower()
    status_label = _STATUS_LABELS.get(status, status or "未知状态")
    metrics = summary.get("metrics") if isinstance(summary.get("metrics"), dict) else {}
    lines = [f"回测状态：{status_label}"]
    performance: List[str] = []

    if total is not None:
        performance.append(f"总收益 {total * 100:+.2f}%")
    ann = _to_float(metrics.get("annual_return"))
    if ann is not None:
        performance.append(f"年化收益 {ann * 100:+.2f}%")

    if dd_abs is not None:
        if dd_abs <= 0.15:
            risk = "回撤较低"
        elif dd_abs <= 0.25:
            risk = "回撤中等"
        else:
            risk = "回撤较高"
        performance.append(f"最大回撤 {dd_abs * 100:.2f}%（{risk}）")

    if win is not None:
        performance.append(f"胜率 {win * 100:.2f}%")

    if performance:
        lines.append("绩效：" + "｜".join(performance))
    lines.append("样本：" + _sample_text(trades))
    lines.append("判断：" + _verdict(total, dd_abs, win, trades, completed))
    return "\n".join(lines)


def _recommendation(
    summary: dict,
    total: Optional[float],
    dd_abs: Optional[float],
    win: Optional[float],
    trades: Optional[int],
    completed: bool,
) -> Tuple[str, str]:
    """返回 (复验优先级文案, 分级key)。本批次最优由调用方单独标记。"""
    if not completed:
        return "未通过", "fail"
    if trades is None:
        return "暂不评级（交易数缺失）", "na"
    if trades < _MIN_OBSERVATION_TRADES:
        return "暂不评级（样本过少）", "invalid"
    if total is None:
        return "暂不评级（数据缺失）", "na"
    if total <= 0:
        return "暂缓复验", "no"
    if dd_abs is None:
        return "暂不评级（回撤缺失）", "na"
    if dd_abs > _MAX_BEST_DRAWDOWN:
        return "暂缓复验（回撤较高）", "no"
    if trades < _BASE_OBSERVATION_TRADES:
        return "暂不评级（样本偏少）", "small"
    if total >= 0.6 and dd_abs <= 0.18:
        return "优先复验", "silver"
    if total >= 0.3:
        if dd_abs <= 0.13:
            return "优先复验", "silver"
        return "建议复验", "bronze"
    if total >= 0.15:
        if dd_abs <= 0.18:
            return "建议复验", "bronze"
        return "观察", "caution"
    if total >= 0.05:
        return "观察", "average"
    return "暂缓复验", "no"


def build_task_export_rows(
    cfg: AStockConfig, run_ids: Sequence[str]
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for index, run_id in enumerate(normalize_export_run_ids(run_ids), 1):
        summary = load_run_summary(cfg, run_id)
        metrics = summary.get("metrics") if isinstance(summary.get("metrics"), dict) else {}
        status = str(summary.get("status") or "ok").strip().lower()
        completed = status in _COMPLETED_STATUSES
        total = _to_float(metrics.get("total_return"))
        dd = _to_float(metrics.get("max_drawdown"))
        dd_abs = abs(dd) if dd is not None else None
        win = _to_float(metrics.get("win_rate"))
        trades = _to_int(metrics.get("n_round_trips"))

        rec_label, tier = _recommendation(summary, total, dd_abs, win, trades, completed)
        rows.append(
            {
                "序号": index,
                "测试条件": _condition_text(summary),
                "买入时间": _schedule_text(summary, side="buy"),
                "买入价格": _session_label(_first_value(summary, "buy_on", "open")),
                "卖出": _schedule_text(summary, side="sell"),
                "卖出价格": _session_label(_first_value(summary, "sell_on", "open")),
                "卦象": _gua_text(summary),
                "测试结论": _conclusion_text(summary, total, dd_abs, win, trades, completed),
                "推荐等级": rec_label,
                "run_id": run_id,
                "任务名称": str(_first_value(summary, "title", _condition_text(summary))),
                "状态": _STATUS_LABELS.get(
                    status,
                    str(summary.get("status") or "ok"),
                ),
                "周期": str(_first_value(summary, "period", "")),
                "回测区间": (
                    f"{_first_value(summary, 'start', '')} ~ "
                    f"{_first_value(summary, 'end', '')}"
                ),
                "信号星期": "、".join(
                    _weekday_label(x)
                    for x in (_first_value(summary, "signal_weekdays", []) or [])
                )
                or "不限",
                "账户模式": str(_first_value(summary, "account_mode", "")),
                "总收益率": metrics.get("total_return"),
                "年化收益率": metrics.get("annual_return"),
                "最大回撤": metrics.get("max_drawdown"),
                "胜率": metrics.get("win_rate"),
                "交易次数": metrics.get("n_round_trips"),
                # 内部字段（不写入表格列，仅供 write_run_task_excel 排版）
                "_completed": completed,
                "_total_return": total,
                "_max_drawdown_abs": dd_abs,
                "_trades": trades,
                "_tier": tier,
                "_is_best": False,
            }
        )

    # 本批次最优必须先满足有效性、样本和回撤门槛，再按总收益选择。
    best_idx = -1
    best_ret: Optional[float] = None
    for i, row in enumerate(rows):
        if not row["_completed"] or row["_tier"] not in _BEST_ELIGIBLE_TIERS:
            continue
        ret = row["_total_return"]
        trades = row["_trades"]
        dd_abs = row["_max_drawdown_abs"]
        if (
            ret is None
            or ret <= 0
            or trades is None
            or trades < _BASE_OBSERVATION_TRADES
            or dd_abs is None
            or dd_abs > _MAX_BEST_DRAWDOWN
        ):
            continue
        if best_ret is None or ret > best_ret:
            best_ret = ret
            best_idx = i
    if best_idx >= 0:
        rows[best_idx]["_is_best"] = True
        rows[best_idx]["推荐等级"] = "★ 本批次最优\n优先复验"
    return rows


def write_run_task_excel(
    cfg: AStockConfig,
    run_ids: Sequence[str],
    path: Optional[Path] = None,
) -> Path:
    """Create an Excel task record modeled after the user's reference sheet."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    rows = build_task_export_rows(cfg, run_ids)
    if path is None:
        export_dir = Path(cfg.output_root) / ".task_exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        path = export_dir / f"回测数据记录_{int(time.time() * 1000)}.xlsx"
    else:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "回测任务"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    thin = Side(style="thin", color="B7C4CE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    best_side = Side(style="medium", color=_BEST_BORDER_COLOR)
    best_border = Border(
        left=best_side,
        right=best_side,
        top=best_side,
        bottom=best_side,
    )
    top_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    top_font = Font(name="宋体", size=11, bold=True, color="FFFFFF")
    header_font = Font(name="宋体", size=11, bold=True, color="1F2937")
    body_font = Font(name="宋体", size=11, color="111827")

    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:G1")
    ws.merge_cells("I1:I2")
    ws["A1"] = "序号"
    ws["B1"] = "回测设置"
    ws["I1"] = "复验优先级"
    for column, value in enumerate(
        ["策略条件", "买入日", "买入时点", "卖出日", "卖出时点", "卦象筛选", "回测摘要"],
        start=2,
    ):
        ws.cell(2, column, value)

    for cell in ws[1]:
        cell.fill = top_fill
        cell.font = top_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws["H2"].comment = Comment(
        "样本提示按完成交易数计算：少于30笔为样本过少，30至99笔为样本偏少，"
        "100笔及以上仅表示达到基础观察阈值，不代表统计显著性。",
        "A股回测系统",
    )
    ws["I1"].comment = Comment(
        "本批次最优判定：任务已完成、收益为正、完成交易不少于100笔、"
        "最大回撤不超过25%，再从符合条件的任务中按总收益选择。",
        "A股回测系统",
    )

    for row in rows:
        ws.append(
            [
                row["序号"],
                row["测试条件"],
                row["买入时间"],
                row["买入价格"],
                row["卖出"],
                row["卖出价格"],
                row["卦象"],
                row["测试结论"],
                row["推荐等级"],
            ]
        )
        row_index = ws.max_row
        is_best = row.get("_is_best")
        tier = row.get("_tier")
        for cell in ws[row_index]:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if cell.column in (2, 7, 8) else "center",
                vertical="center",
                wrap_text=cell.column in (2, 7, 8, 9),
            )
        # 分级着色 / 最佳行高亮
        if is_best:
            for cell in ws[row_index]:
                cell.fill = PatternFill("solid", fgColor=_BEST_ROW_FILL)
                cell.font = Font(name="宋体", size=11, bold=True, color="111827")
                cell.border = best_border
            # 摘要红字、最优标签红底白字，方便快速定位。
            ws.cell(row_index, 8).font = Font(
                name="宋体", size=11, bold=True, color=_BEST_RED_FONT
            )
            ws.cell(row_index, 9).fill = PatternFill(
                "solid", fgColor=_BEST_LABEL_FILL
            )
            ws.cell(row_index, 9).font = Font(
                name="宋体", size=11, bold=True, color=_BEST_WHITE_FONT
            )
            ws.cell(row_index, 9).comment = Comment(
                "本批次最优：通过样本与回撤门槛后，按总收益在本次导出任务中排名第一。",
                "A股回测系统",
            )
        elif tier in _TIER_FILL:
            fill = PatternFill("solid", fgColor=_TIER_FILL[tier])
            for cell in ws[row_index]:
                cell.fill = fill
        # 依据换行和结论长度动态设定行高，保证长文本不被截断。
        conclusion_text = str(ws.cell(row_index, 8).value or "")
        chars_per_line = max(24, int(_SUMMARY_COLUMN_WIDTH * 0.58))
        visual_lines = sum(
            max(1, math.ceil(len(part) / chars_per_line))
            for part in (conclusion_text.splitlines() or [""])
        )
        ws.row_dimensions[row_index].height = min(
            104,
            max(76, visual_lines * 14 + 6),
        )
        ws.cell(row_index, 2).comment = Comment(
            f"run_id={row['run_id']}\n任务名称={row['任务名称']}",
            "A股回测系统",
        )

    for column, width in {
        "A": 10.2,
        "B": 30,
        "C": 12,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 32,
        "H": _SUMMARY_COLUMN_WIDTH,
        "I": 18,
    }.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.auto_filter.ref = f"A2:I{ws.max_row}"
    ws.print_title_rows = "1:2"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_orientation = "landscape"

    detail = wb.create_sheet("任务参数")
    detail.sheet_view.showGridLines = False
    detail.freeze_panes = "A2"
    detail_headers = [
        "序号",
        "run_id",
        "任务名称",
        "状态",
        "周期",
        "回测区间",
        "信号星期",
        "买入规则",
        "卖出规则",
        "账户模式",
        "总收益率",
        "年化收益率",
        "最大回撤",
        "胜率",
        "交易次数",
    ]
    detail.append(detail_headers)
    for cell in detail[1]:
        cell.fill = top_fill
        cell.font = top_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in rows:
        detail.append(
            [
                row["序号"],
                row["run_id"],
                row["任务名称"],
                row["状态"],
                row["周期"],
                row["回测区间"],
                row["信号星期"],
                f"{row['买入时间']} {row['买入价格']}".strip(),
                f"{row['卖出']} {row['卖出价格']}".strip(),
                row["账户模式"],
                row["总收益率"],
                row["年化收益率"],
                row["最大回撤"],
                row["胜率"],
                row["交易次数"],
            ]
        )
        for cell in detail[detail.max_row]:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for column in ("K", "L", "M", "N"):
        for cell in detail[column][1:]:
            cell.number_format = "0.00%"
    for column, width in {
        "A": 8,
        "B": 26,
        "C": 42,
        "D": 22,
        "E": 10,
        "F": 24,
        "G": 16,
        "H": 18,
        "I": 18,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 12,
        "O": 12,
    }.items():
        detail.column_dimensions[column].width = width
    detail.auto_filter.ref = f"A1:O{detail.max_row}"

    tmp_path = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp_path)
    tmp_path.replace(path)
    return path
