"""Independent forecast knowledge base (64 gua / 384 yao judgements)."""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from ..bagua.calculator import TRIGRAMS, YAO_ORDER_MAP, rebuild_knowledge_from_excel
from .name_norm import normalize_gua_name, strip_gua_symbol

_GUA_ORDER_RE = re.compile(r"[\u4dc0-\u4dff]")


class ForecastKnowledgeBase:
    """In-memory indices over bagua_384-style forecast JSON."""

    def __init__(self, data: dict):
        self.raw = data
        self.entries: List[dict] = list(data.get("entries") or [])
        self.source_file = data.get("source_file")
        self.source_sha256 = data.get("source_sha256")
        self.version_id = data.get("version_id")
        self._by_gua_yao: Dict[Tuple[str, int], dict] = {}
        self._by_gua: Dict[str, List[dict]] = {}
        for e in self.entries:
            gname = normalize_gua_name(e.get("gua_name") or e.get("full_name") or "")
            yo = e.get("yao_order")
            if not gname or yo is None:
                continue
            yo_i = int(yo)
            self._by_gua_yao[(gname, yo_i)] = e
            self._by_gua.setdefault(gname, []).append(e)

    @classmethod
    def from_json_path(cls, path: Path | str) -> "ForecastKnowledgeBase":
        path = Path(path)
        data = json.loads(path.read_text(encoding="utf-8"))
        return cls(data)

    def lookup_ben_yao(self, ben_gua: str, yao_order: int) -> Optional[dict]:
        key = (normalize_gua_name(ben_gua), int(yao_order))
        hit = self._by_gua_yao.get(key)
        if hit:
            return hit
        # try without 为X variants already normalized
        return None

    def gua_names(self) -> List[str]:
        return sorted(self._by_gua.keys())


def import_xlsx_to_kb(
    xlsx_path: Path | str,
    *,
    out_json: Path,
    version_id: Optional[str] = None,
) -> dict:
    """
    Import 简判 Excel into forecast KB JSON.

    Reuses rebuild_knowledge_from_excel when layout matches bagua authority sheet;
    otherwise parses openpyxl with forward-fill (same columns as 豆包表).
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(str(xlsx_path))
    version_id = version_id or time.strftime("%Y%m%d_%H%M%S")
    out_json = Path(out_json)
    out_json.parent.mkdir(parents=True, exist_ok=True)

    # Prefer full parser that includes 操作信号 when the sheet has that column
    try:
        kb = _parse_judgement_xlsx(xlsx_path, out_json)
        # if rebuild structure better for missing fields, fill from calculator rebuild
        if not kb.get("entries"):
            raise ValueError("empty entries")
    except Exception:
        try:
            kb = rebuild_knowledge_from_excel(xlsx_path, out_json)
        except Exception:
            kb = _parse_judgement_xlsx(xlsx_path, out_json)
    # Ensure operation_signal key exists on all entries
    for e in kb.get("entries") or []:
        e.setdefault("operation_signal", e.get("operation_signal") or "")

    kb["version_id"] = version_id
    kb["imported_at"] = time.strftime("%Y-%m-%dT%H:%M:%S")
    out_json.write_text(json.dumps(kb, ensure_ascii=False, indent=2), encoding="utf-8")
    return kb


def _parse_judgement_xlsx(xlsx_path: Path, out_json: Path) -> dict:
    """Fallback parser for 卦象卦名/爻位/变卦/个股行情简判 layout."""
    import openpyxl

    xlsx_path = Path(xlsx_path)
    sha = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    # header row 1
    headers = {}
    for c in range(1, 20):
        v = ws.cell(1, c).value
        if v:
            headers[str(v).strip()] = c

    def col(*names: str) -> Optional[int]:
        for n in names:
            if n in headers:
                return headers[n]
        return None

    c_full = col("卦象卦名", "full_name") or 1
    c_gci = col("卦辞原文") or 2
    c_gang = col("卦核心总纲") or 3
    c_yao = col("爻位") or 4
    c_bian = col("变卦") or 5
    c_yaoci = col("爻辞原文") or 6
    c_pan = col("个股行情简判") or 7
    c_note = col("备注&实操总结", "备注") or 8
    c_sig = col("操作信号", "信号", "操作") or 9

    self_map = {
        "乾为天": (1, 1),
        "坤为地": (8, 8),
        "震为雷": (4, 4),
        "艮为山": (7, 7),
        "坎为水": (6, 6),
        "离为火": (3, 3),
        "巽为风": (5, 5),
        "兑为泽": (2, 2),
    }
    alias_to_id = {v["alias"]: k for k, v in TRIGRAMS.items()}

    def parse_ul(name: str):
        name = normalize_gua_name(name)
        if name in self_map:
            return self_map[name]
        if len(name) >= 2:
            a0, a1 = name[0], name[1]
            if a0 in alias_to_id and a1 in alias_to_id:
                return alias_to_id[a0], alias_to_id[a1]
        return 0, 0

    entries: List[dict] = []
    current: Optional[dict] = None
    gua_order = 0
    for r in range(2, ws.max_row + 1):
        full = ws.cell(r, c_full).value
        gci = ws.cell(r, c_gci).value
        gang = ws.cell(r, c_gang).value
        yao_pos = ws.cell(r, c_yao).value
        biangua = ws.cell(r, c_bian).value
        yao_ci = ws.cell(r, c_yaoci).value
        pan = ws.cell(r, c_pan).value
        note = ws.cell(r, c_note).value
        sig = ws.cell(r, c_sig).value if c_sig else None
        if full and str(full).strip():
            gua_order += 1
            full_s = str(full).strip()
            symbol = full_s[0] if full_s and _GUA_ORDER_RE.match(full_s[0]) else ""
            rest = strip_gua_symbol(full_s)
            current = {
                "gua_order": gua_order,
                "full_name": full_s,
                "gua_symbol": symbol,
                "gua_name": rest,
                "gua_ci": str(gci).strip() if gci else "",
                "core_gang": str(gang).strip() if gang else "",
            }
        elif current is None:
            continue
        else:
            if gci and str(gci).strip():
                current["gua_ci"] = str(gci).strip()
            if gang and str(gang).strip():
                current["core_gang"] = str(gang).strip()

        yao = str(yao_pos).strip() if yao_pos else ""
        u, l = parse_ul(current["gua_name"])
        entries.append(
            {
                "excel_row": r,
                "gua_order": current["gua_order"],
                "gua_symbol": current["gua_symbol"],
                "gua_name": current["gua_name"],
                "full_name": current["full_name"],
                "gua_ci": current["gua_ci"],
                "core_gang": current["core_gang"],
                "yao_order": YAO_ORDER_MAP.get(yao),
                "yao_name": yao,
                "biangua": str(biangua).strip() if biangua else "",
                "yao_ci": str(yao_ci).strip() if yao_ci else "",
                "market_judgement": str(pan).strip() if pan else "",
                "note": str(note).strip() if note else "",
                "operation_signal": str(sig).strip() if sig else "",
                "upper": u,
                "lower": l,
                "upper_name": TRIGRAMS.get(u, {}).get("name", ""),
                "lower_name": TRIGRAMS.get(l, {}).get("name", ""),
                "upper_alias": TRIGRAMS.get(u, {}).get("alias", ""),
                "lower_alias": TRIGRAMS.get(l, {}).get("alias", ""),
                "upper_symbol": TRIGRAMS.get(u, {}).get("symbol", ""),
                "lower_symbol": TRIGRAMS.get(l, {}).get("symbol", ""),
            }
        )

    # propagate group fields
    by_go: Dict[int, List[dict]] = {}
    for e in entries:
        by_go.setdefault(int(e["gua_order"]), []).append(e)
    for go, rows in by_go.items():
        gang = next((x["core_gang"] for x in rows if x["core_gang"]), "")
        gci = next((x["gua_ci"] for x in rows if x["gua_ci"]), "")
        for x in rows:
            x["core_gang"] = gang
            x["gua_ci"] = gci

    kb = {
        "source_file": xlsx_path.name,
        "source_sha256": sha,
        "source_path": str(xlsx_path.resolve()),
        "trigrams": TRIGRAMS,
        "entries": entries,
        "count_gua": len(by_go),
        "count_yao": len(entries),
    }
    out_json.write_text(json.dumps(kb, ensure_ascii=False, indent=2), encoding="utf-8")
    return kb


def seed_from_existing_json(src: Path, dest: Path, version_id: str = "seed") -> dict:
    """Copy backtest bagua_384.json into forecast KB (optional bootstrap)."""
    src, dest = Path(src), Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    data = json.loads(src.read_text(encoding="utf-8"))
    data["version_id"] = version_id
    data["imported_at"] = time.strftime("%Y-%m-%dT%H:%M:%S")
    data["seeded_from"] = str(src)
    dest.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data


def save_version_copy(active_json: Path, versions_dir: Path, version_id: str) -> Path:
    versions_dir = Path(versions_dir)
    versions_dir.mkdir(parents=True, exist_ok=True)
    dest = versions_dir / f"{version_id}.json"
    shutil.copy2(active_json, dest)
    meta = {
        "version_id": version_id,
        "path": str(dest),
        "sha256": hashlib.sha256(dest.read_bytes()).hexdigest(),
        "saved_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
    }
    (versions_dir / f"{version_id}.meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return dest
